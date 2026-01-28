#!/usr/bin/env pythonw
"""
Task Manager GUI - A graphical application to manage daily tasks.
"""
import os
import json
import datetime
import tkinter as tk
from tkinter import ttk, messagebox, simpledialog
from tkcalendar import DateEntry
from typing import Dict, List, Optional, Any, Tuple
import subprocess  # Add this import for the subprocess module
import webbrowser
import logging
import threading
# OPTIMIZED: Heavy imports moved to lazy loading (inside functions)
# import pymssql  # Changed: Import pymssql instead of pyodbc
import getpass # Import the 'getpass' module to get username
import tkinter.messagebox as messagebox # Import messagebox
import platform
import sys # <<< Add sys
import appdirs # <<< Add appdirs (requires pip install appdirs)
# OPTIMIZED: Keyring and smartdb_login imported lazily when needed
# import keyring # Added for secure password storage
# import smartdb_login # Added for scraping functionality
from tkinter import filedialog # Added for askdirectory
import time # Added for timer
import shutil # Added for moving files
import configparser # Added for configuration file support

# Excel import library (lazy loaded when needed)
# import openpyxl

# Import enhanced features module
try:
    from task_manager_enhancements import (
        TimeTrackingManager,
        SubtaskManager,
        TemplateManager,
        AdvancedSearchManager,
        DashboardManager
    )
    ENHANCED_FEATURES_AVAILABLE = True
    logging.info("Enhanced features module loaded successfully")
except ImportError as e:
    ENHANCED_FEATURES_AVAILABLE = False
    logging.warning(f"Enhanced features not available: {e}")
    logging.info("Run 'sql/create_enhanced_features.sql' in SSMS first!")

# --- Configuration Loader ---
class ConfigManager:
    """Manages application configuration from config.ini file."""
    
    def __init__(self, config_file='config.ini'):
        self.config = configparser.ConfigParser()
        self.config_file = config_file
        self._load_config()
    
    def _load_config(self):
        """Load configuration file with fallback to defaults."""
        try:
            # Determine config file path (same directory as script/exe)
            if getattr(sys, 'frozen', False):
                base_path = os.path.dirname(sys.executable)
            else:
                base_path = os.path.dirname(os.path.abspath(__file__))
            
            config_path = os.path.join(base_path, self.config_file)
            
            if os.path.exists(config_path):
                self.config.read(config_path)
                logging.info(f"Configuration loaded from: {config_path}")
            else:
                logging.warning(f"Config file not found: {config_path}. Using defaults.")
                self._set_defaults()
        except Exception as e:
            logging.error(f"Error loading configuration: {e}", exc_info=True)
            self._set_defaults()
    
    def _set_defaults(self):
        """Set default configuration values."""
        # Database defaults
        if 'Database' not in self.config:
            self.config['Database'] = {}
        self.config['Database'].setdefault('server', '10.195.96.58:1433')
        self.config['Database'].setdefault('database', 'TaskManager1')
        self.config['Database'].setdefault('timeout', '10')
        self.config['Database'].setdefault('enable_retry', 'true')
        self.config['Database'].setdefault('max_retry_attempts', '3')
        self.config['Database'].setdefault('retry_delay_seconds', '2')
        
        # Update defaults
        if 'Update' not in self.config:
            self.config['Update'] = {}
        self.config['Update'].setdefault('server', r'\\10.195.103.198\shared\TaskManager\updates')
        self.config['Update'].setdefault('version_file', 'version.json')
        self.config['Update'].setdefault('auto_check', 'true')
        
        # Paths defaults
        if 'Paths' not in self.config:
            self.config['Paths'] = {}
        self.config['Paths'].setdefault('standard_location', r'\\srb096154\01_CESSD_SCG_CAD\01_Projects')
        
        # Application defaults
        if 'Application' not in self.config:
            self.config['Application'] = {}
        self.config['Application'].setdefault('version', '0.23.0')
        self.config['Application'].setdefault('log_level', 'INFO')
    
    def get(self, section, key, fallback=None):
        """Get configuration value with fallback."""
        try:
            return self.config.get(section, key, fallback=fallback)
        except (configparser.NoSectionError, configparser.NoOptionError) as e:
            logging.debug(f"Config key not found: [{section}] {key}. Using fallback: {fallback}")
            return fallback
    
    def getint(self, section, key, fallback=0):
        """Get integer configuration value."""
        try:
            return self.config.getint(section, key, fallback=fallback)
        except (configparser.NoSectionError, configparser.NoOptionError, ValueError) as e:
            logging.debug(f"Config key not found or invalid: [{section}] {key}. Using fallback: {fallback}")
            return fallback
    
    def getboolean(self, section, key, fallback=False):
        """Get boolean configuration value."""
        try:
            return self.config.getboolean(section, key, fallback=fallback)
        except (configparser.NoSectionError, configparser.NoOptionError, ValueError) as e:
            logging.debug(f"Config key not found or invalid: [{section}] {key}. Using fallback: {fallback}")
            return fallback

# Initialize global configuration
app_config = ConfigManager()

# --- Determine Log File Path ---
def get_log_file_path():
    """Determines the appropriate absolute path for the log file, prioritizing the script/exe directory."""
    app_name = "TaskManager"
    # author_name = "TaskManagerApp" # Not needed if appdirs isn't used

    # <<< REMOVED appdirs logic >>>
    # try:
    #     # Preferred: User's AppData directory (log subfolder)
    #     log_dir = appdirs.user_log_dir(app_name, author_name)
    #     # ... (rest of appdirs block removed) ...
    # except Exception as e_appdata:
    #     print(f"[WARNING] Could not use AppData directory ({log_dir}): {e_appdata}. Trying executable directory.")

    # <<< MODIFIED: Make executable/script directory the primary attempt >>>
    try:
        if getattr(sys, 'frozen', False):
            # Running as bundled executable
            base_path = os.path.dirname(sys.executable)
        else:
            # Running as a script
            base_path = os.path.dirname(os.path.abspath(__file__))
        log_path = os.path.join(base_path, 'task_manager.log')
        # Basic write test
        try:
            # <<< Use 'a' append mode for testing to avoid clearing existing file if test fails later >>>
            with open(log_path, 'a') as f: 
                f.write("") # Test write permission
        except Exception as e:
            print(f"[WARNING] Could not write to log file in executable/script directory: {e}")
            raise # Re-raise to trigger fallback
        # <<< MODIFIED: Log message indicates this is the intended location >>>
        print(f"[INFO] Attempting to log to executable/script directory: {log_path}") 
        return log_path
    except Exception as e_exe_dir:
        print(f"[ERROR] Could not use executable/script directory: {e_exe_dir}")
        # Final fallback: Current working directory (less reliable)
        log_path = 'task_manager.log'
        print(f"[CRITICAL] Falling back to current working directory for log: {os.path.abspath(log_path)}")
        return log_path

# Configure logging *using the absolute path*
log_file_path = get_log_file_path() # <<< Call the function (no change here)

# <<< Get the root logger and remove existing handlers >>>
# This prevents adding handlers multiple times if the setup code is run again
root_logger = logging.getLogger()
# Remove all handlers associated with the root logger object.
for handler in root_logger.handlers[:]:
    root_logger.removeHandler(handler)
    handler.close() # Ensure handlers are closed before removing

# <<< Create handlers explicitly >>>
log_formatter = logging.Formatter('%(asctime)s - %(levelname)s - [%(threadName)s] - %(message)s')

# File Handler
try:
    file_handler = logging.FileHandler(log_file_path, mode='w') # Overwrite log each run
    file_handler.setFormatter(log_formatter)
    root_logger.addHandler(file_handler)
except Exception as e:
     print(f"[ERROR] Failed to create file handler for log file '{log_file_path}': {e}")
     file_handler = None # Ensure it's None if creation failed

# Stream Handler (Console)
stream_handler = logging.StreamHandler(sys.stdout)
stream_handler.setFormatter(log_formatter)
root_logger.addHandler(stream_handler)

# Set level on the root logger
root_logger.setLevel(logging.INFO)

# Log the chosen path after basicConfig is set up
logging.info(f"--- Log Start ---")
logging.info(f"Log file path determined: {log_file_path}") # (no change here)
# <<< Add an immediate flush >>>
if file_handler:
    file_handler.flush()
# <<< Add an immediate test message >>>
logging.info("--- Initial logging setup complete. This message should appear in the log. ---")
if file_handler:
    file_handler.flush()

# User mapping with TCN numbers and SQL credentials
USER_MAPPING = {
    "Jay": {"tcn": "a0011071", "sql_user": "TaskUser1"},
    "Jude": {"tcn": "a0010756", "sql_user": "TaskUser2"},
    "Jorgen": {"tcn": "a0012923", "sql_user": "TaskUser3"},
    "Earl": {"tcn": "a0010751", "sql_user": "TaskUser4"},
    "Philip": {"tcn": "a0012501", "sql_user": "TaskUser5"},
    "Sam": {"tcn": "a0008432", "sql_user": "TaskUser6"},
    "Glenn": {"tcn": "a0003878", "sql_user": "TaskUser7"}
}

# List of users that can be assigned tasks (for backwards compatibility with comboboxes)
USERS = ["All"] + sorted(USER_MAPPING.keys())

class TaskManagerApp:
    """GUI application for managing tasks."""
    
    def __init__(self, root, task_manager, is_embedded=False, back_callback=None):
        """Initialize the GUI application."""
        self.root = root
        self.is_embedded = is_embedded
        self.back_callback = back_callback
        
        # Only set window properties if not embedded
        if not self.is_embedded:
            self.root.title("Task Manager")
            self.root.geometry("900x600")
            self.root.minsize(800, 500)
        
        # Set application icon
        try:
            # Only set icon if not embedded (icon should be set by parent)
            if not self.is_embedded:
                # Determine base path (works for both script and frozen executable)
                if getattr(sys, 'frozen', False):
                    # Running as compiled executable
                    base_path = os.path.dirname(sys.executable)
                else:
                    # Running as script
                    base_path = os.path.dirname(os.path.abspath(__file__))
                
                # Try PNG first (more reliable in tkinter)
                icon_path = os.path.join(base_path, 'TaskManager_main.png')
                if os.path.exists(icon_path):
                    self.main_icon_image = tk.PhotoImage(file=icon_path)
                    self.root.iconphoto(True, self.main_icon_image)
                    logging.info(f"Icon loaded from PNG: {icon_path}")
                else:
                    # Fallback: Try .ico file (Windows)
                    icon_path = os.path.join(base_path, 'TaskManager_main.ico')
                    if os.path.exists(icon_path):
                        self.root.iconbitmap(icon_path)
                        logging.info(f"Icon loaded from ICO: {icon_path}")
                    else:
                        logging.warning(f"No icon file found. Searched: {base_path}")
        except Exception as e:
            logging.warning(f"Could not set application icon: {e}")
        
        # Configure styles
        self.style = ttk.Style()
        
        # Try to set a modern theme if available
        try:
            self.root.tk.call("source", "azure.tcl")
            self.style.theme_use("azure")
        except tk.TclError:
            # If azure theme is not available, try other themes
            available_themes = self.style.theme_names()
            for theme in ["clam", "alt", "vista"]:
                if theme in available_themes:
                    self.style.theme_use(theme)
                    break
        
        # Configure styles AFTER theme is set to override theme defaults
        self.style.configure("TButton", padding=6, font=('Helvetica', 10))
        self.style.configure("Compact.TButton", padding=0, font=('Helvetica', 10))  # Zero padding to match entry
        self.style.configure("TLabel", font=('Helvetica', 10))
        self.style.configure("Header.TLabel", font=('Helvetica', 12, 'bold'))
        self.style.configure("Title.TLabel", font=('Helvetica', 16, 'bold'))
        
        # Get current user
        self.current_user = self._get_current_user()
        


        # Connection string - Not used directly with pymssql's connect()
        # self.CONNECTION_STRING = f'DRIVER={{SQL Server}};SERVER={self.SERVER};DATABASE={self.DATABASE};Trusted_Connection=yes;'
        
        # Initialize connection_error flag and task manager
        self.connection_error = False
        try:
            self.task_manager = task_manager
        except Exception as e:
            self.connection_error = True
            messagebox.showerror("Connection Error", 
                                f"Could not connect to the SQL Server database.\n\n"
                                f"Error: {str(e)}\n\n"
                                f"The application will run without database connection.")
            logging.error("Connection error at startup", exc_info=True) # Log the exception
            self.task_manager = None
        
        # Initialize deleted tasks stack
        self.deleted_tasks_stack = []
        self.max_undo_stack_size = 10  # Maximum number of undo operations to keep
        
        # Initialize refresh lock to prevent multiple simultaneous refresh operations
        self.refresh_lock = threading.Lock()
        self.refresh_in_progress = False
        
        # Create main frame
        self.main_frame = ttk.Frame(self.root, padding="10")
        self.main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Create header
        self.create_header()
        
        # Create load status banner
        self.create_load_status_banner()
        
        # Create task list frame
        self.create_task_list()
        
        # Create control panel
        self.create_control_panel()
        
        # Create status bar
        self.create_status_bar()
        
        # Set initial status
        self.set_status("Loading tasks...")
        
        # Schedule task loading after UI is displayed
        def initial_load():
            self.refresh_task_list()
            self.hide_loading_indicator()  # Ensure loading indicator is hidden
        
        self.root.after(100, initial_load)
        
        # Bind the window close event to cleanup method (only if not embedded)
        if not self.is_embedded:
            self.root.protocol("WM_DELETE_WINDOW", self.on_close)
    
    def _get_current_user(self):
        """Get the current user's username."""
        try:
            username = os.getenv('USERNAME') or os.getenv('USER') or 'Unknown User'
            logging.info(f"Current user: {username}") # Log the username
            return username
        except (OSError, KeyError) as e:
            logging.warning(f"Error getting current user: {e}")
            return 'Unknown User'
    
    def on_close(self):
        """Handle application close event - clean up resources and close database connections."""
        try:
            logging.info("Application shutdown sequence started.") # Log shutdown start
            # Close database connection if it exists
            if hasattr(self, 'task_manager') and self.task_manager:
                if hasattr(self.task_manager, 'conn') and self.task_manager.conn:
                    try:
                        # Commit any pending changes
                        self.task_manager.conn.commit()
                        # Close the connection
                        self.task_manager.conn.close()
                        logging.info("Database connection closed properly")
                    except Exception as e:
                        logging.error(f"Error closing database connection: {str(e)}", exc_info=True)
        except Exception as e:
            logging.error(f"Error during application shutdown: {str(e)}", exc_info=True)
        finally:
            logging.info("Flushing log handlers before exit...")
            # <<< Explicitly flush and close handlers on exit >>>
            for handler in logging.getLogger().handlers:
                try:
                    handler.flush()
                    handler.close()
                except Exception as e:
                    print(f"Error closing handler {handler}: {e}") # Print error during close
            # Close the application
            self.root.destroy()
    
    def create_load_status_banner(self):
        """Create banner showing load status - Not needed anymore since we load all tasks."""
        # Banner removed - we now load all tasks by default
        pass
    
    def load_all_tasks(self):
        """Load all tasks from database."""
        if self.task_manager.all_tasks_loaded:
            return
        
        self.set_status("Loading all tasks...")
        self.show_loading_indicator()
        
        # Load in background thread
        def load_thread():
            try:
                self.task_manager.reload_tasks(load_all=True)
                # Update UI in main thread
                self.root.after(0, self._on_all_tasks_loaded)
            except Exception as e:
                logging.error(f"Error loading all tasks: {e}", exc_info=True)
                self.root.after(0, lambda: messagebox.showerror("Error", f"Failed to load all tasks: {e}"))
                self.root.after(0, self.hide_loading_indicator)
        
        threading.Thread(target=load_thread, daemon=True).start()
    
    def _on_all_tasks_loaded(self):
        """Called after all tasks are loaded."""
        # Hide banner
        if hasattr(self, 'banner_frame'):
            self.banner_frame.pack_forget()
        
        # Refresh task list
        self.refresh_task_list()
        
        total = len(self.task_manager.tasks)
        self.set_status(f"All {total} tasks loaded successfully.")
    
    def create_header(self):
        """Create the application header."""
        header_frame = ttk.Frame(self.main_frame)
        header_frame.pack(fill=tk.X, pady=(0, 10))
        
        # Row 1: Title and storage indicator (compact)
        title_row = ttk.Frame(header_frame)
        title_row.pack(fill=tk.X, pady=(0, 5))
        
        # Title with storage indicator inline
        title_text = "Task Manager "
        if self.connection_error:
            storage_text = "(No Database)"
            storage_color = "red"
        else:
            storage_text = "(SQL Server)"
            storage_color = "green"
        
        title_label = ttk.Label(title_row, text=title_text, style="Title.TLabel")
        title_label.pack(side=tk.LEFT)
        
        self.storage_indicator = ttk.Label(title_row, text=storage_text, foreground=storage_color, font=('Helvetica', 11))
        self.storage_indicator.pack(side=tk.LEFT)
        
        # Row 2: Filters and Back button (all in one row)
        filter_row = ttk.Frame(header_frame)
        filter_row.pack(fill=tk.X)
        
        # Category filter
        ttk.Label(filter_row, text="Category:").pack(side=tk.LEFT, padx=(0, 5))
        
        self.category_var = tk.StringVar(value="All")
        self.category_combo = ttk.Combobox(filter_row, textvariable=self.category_var, width=12, state="readonly")
        self.category_combo.pack(side=tk.LEFT, padx=(0, 10))
        self.update_category_filter()
        self.category_combo.bind("<<ComboboxSelected>>", lambda e: self.refresh_task_list())
        
        # Main Staff filter
        ttk.Label(filter_row, text="Main Staff:").pack(side=tk.LEFT, padx=(0, 5))
        
        self.main_staff_var = tk.StringVar(value="All")
        self.main_staff_combo = ttk.Combobox(filter_row, textvariable=self.main_staff_var, width=12, values=USERS, state="readonly")
        self.main_staff_combo.pack(side=tk.LEFT, padx=(0, 10))
        self.main_staff_combo.bind("<<ComboboxSelected>>", lambda e: self.refresh_task_list())
        
        # User filter
        ttk.Label(filter_row, text="User:").pack(side=tk.LEFT, padx=(0, 5))
        
        self.user_var = tk.StringVar(value="All")
        self.user_combo = ttk.Combobox(filter_row, textvariable=self.user_var, width=12, values=USERS, state="readonly")
        self.user_combo.pack(side=tk.LEFT, padx=(0, 10))
        self.user_combo.bind("<<ComboboxSelected>>", lambda e: self.refresh_task_list())
        
        # Show completed checkbox
        ttk.Label(filter_row, text="Status:").pack(side=tk.LEFT, padx=(0, 5))
        
        self.show_completed_var = tk.BooleanVar(value=False)
        show_completed_check = ttk.Checkbutton(
            filter_row, 
            text="Show Completed", 
            variable=self.show_completed_var,
            command=self.refresh_task_list
        )
        show_completed_check.pack(side=tk.LEFT, padx=(0, 10))
        
        # Back to Menu button (only show if embedded with callback)
        if self.is_embedded and self.back_callback:
            back_btn = tk.Button(
                filter_row,
                text="← Menu",
                command=self.back_callback,
                bg='#607D8B',
                fg='white',
                font=('Arial', 9, 'bold'),
                cursor='hand2',
                relief=tk.RAISED,
                bd=1,
                padx=10,
                pady=2
            )
            back_btn.pack(side=tk.RIGHT, padx=5)
    
    def create_task_list(self):
        """Create the task list with a treeview."""
        # Create frame with scrollbar
        task_frame = ttk.Frame(self.main_frame)
        task_frame.pack(fill=tk.BOTH, expand=True, pady=5)
        
        # Create treeview with both vertical and horizontal scrollbars
        columns = (
            "status", "title", "rev", "applied_vessel",
            "priority", "main_staff", "assigned_to",
            "qtd_mhr", "actual_mhr",
            "target_start", "target_finish"  # Target date columns
        )
        
        # Create a frame to hold the treeview and scrollbars
        tree_frame = ttk.Frame(task_frame)
        tree_frame.pack(fill=tk.BOTH, expand=True)
        
        # Create the treeview
        self.task_tree = ttk.Treeview(tree_frame, columns=columns, show="headings")
        
        # Define standard font for all treeview elements - make it bold
        standard_font = ('TkDefaultFont', 9, 'bold')
        
        # Configure the treeview to use the standard font
        self.style.configure("Treeview", font=standard_font)
        self.style.configure("Treeview.Heading", font=('TkDefaultFont', 9, 'bold'))
        
        # Define headings with proper command for toggling sort direction
        self.task_tree.heading("status", text="Status", command=lambda: self.sort_treeview("status", True))
        self.task_tree.heading("title", text="Equipment Name", command=lambda: self.sort_treeview("title", True))
        self.task_tree.heading("rev", text="Rev", command=lambda: self.sort_treeview("rev", True))
        self.task_tree.heading("applied_vessel", text="Applied Vessel", command=lambda: self.sort_treeview("applied_vessel", True))
        self.task_tree.heading("priority", text="Priority", command=lambda: self.sort_treeview("priority", True))
        self.task_tree.heading("main_staff", text="Main Staff", command=lambda: self.sort_treeview("main_staff", True))
        self.task_tree.heading("assigned_to", text="Assigned To", command=lambda: self.sort_treeview("assigned_to", True))
        self.task_tree.heading("qtd_mhr", text="Qtd Mhr", command=lambda: self.sort_treeview("qtd_mhr", True))  # Heading for Qtd Mhr
        self.task_tree.heading("actual_mhr", text="Actual Mhr", command=lambda: self.sort_treeview("actual_mhr", True))  # Heading for Actual Mhr
        self.task_tree.heading("target_start", text="Target Start", command=lambda: self.sort_treeview("target_start", True))  # Heading for Target Start
        self.task_tree.heading("target_finish", text="Target Finish", command=lambda: self.sort_treeview("target_finish", True))  # Heading for Target Finish
        
        # Define optimized column widths - REDUCED WIDTHS FOR BETTER VISIBILITY
        self.task_tree.column("status", width=120, minwidth=100, anchor=tk.W)  # Reduced width
        self.task_tree.column("title", width=250, minwidth=200) # Reduced width
        self.task_tree.column("rev", width=40, minwidth=40, anchor=tk.CENTER) # Reduced width
        self.task_tree.column("applied_vessel", width=100, minwidth=100) # Reduced width
        self.task_tree.column("priority", width=70, minwidth=70, anchor=tk.CENTER) # Reduced width
        self.task_tree.column("main_staff", width=80, minwidth=80, anchor=tk.CENTER) # Reduced width
        self.task_tree.column("assigned_to", width=80, minwidth=80, anchor=tk.CENTER) # Reduced width
        self.task_tree.column("qtd_mhr", width=70, minwidth=70, anchor=tk.CENTER)  # Width for Qtd Mhr - slightly reduced
        self.task_tree.column("actual_mhr", width=70, minwidth=70, anchor=tk.CENTER)  # Width for Actual Mhr - slightly reduced
        self.task_tree.column("target_start", width=100, minwidth=100, anchor=tk.CENTER)  # Width for Target Start
        self.task_tree.column("target_finish", width=100, minwidth=100, anchor=tk.CENTER)  # Width for Target Finish
        
        # Initialize sorting variables
        self.sort_column = None
        self.sort_reverse = False
        
        # Add vertical scrollbar
        v_scrollbar = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=self.task_tree.yview)
        self.task_tree.configure(yscrollcommand=v_scrollbar.set)
        
        # Add horizontal scrollbar
        h_scrollbar = ttk.Scrollbar(tree_frame, orient=tk.HORIZONTAL, command=self.task_tree.xview)
        self.task_tree.configure(xscrollcommand=h_scrollbar.set)
        
        # Grid layout for treeview and scrollbars to ensure they're always visible
        self.task_tree.grid(row=0, column=0, sticky='nsew')
        v_scrollbar.grid(row=0, column=1, sticky='ns')
        h_scrollbar.grid(row=1, column=0, sticky='ew')
        
        # Configure the grid to expand properly
        tree_frame.columnconfigure(0, weight=1)
        tree_frame.rowconfigure(0, weight=1)
        
        # Bind keyboard shortcuts
        self.task_tree.bind("<Control-a>", self._select_all_tasks)
        self.task_tree.bind("<Delete>", self._delete_task_via_keyboard)
        
        # Create a loading indicator overlay with better visibility
        self.loading_frame = tk.Frame(tree_frame, bg='white', relief=tk.FLAT, bd=2)
        
        # Add a subtle border
        border_frame = tk.Frame(self.loading_frame, bg='#e0e0e0', bd=1, relief=tk.SOLID)
        border_frame.pack(fill=tk.BOTH, expand=True, padx=2, pady=2)
        
        inner_frame = tk.Frame(border_frame, bg='white')
        inner_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # Animated loading text
        self.loading_label = tk.Label(
            inner_frame, 
            text="⏳ Loading tasks...", 
            font=('Helvetica', 14, 'bold'),
            bg='white',
            fg='#1976d2'
        )
        self.loading_label.pack(pady=10)
        
        # Progress info label
        self.loading_info = tk.Label(
            inner_frame,
            text="Please wait...",
            font=('Helvetica', 10),
            bg='white',
            fg='#666'
        )
        self.loading_info.pack()
        
        # Bind double-click to view task details
        self.task_tree.bind("<Double-1>", self.view_task_details)
        
        # Bind right-click to show context menu
        self.task_tree.bind("<Button-3>", self.show_context_menu)
        
        # Bind single-click to check for link clicks
        self.task_tree.bind("<ButtonRelease-1>", self.check_link_click)
        
        # Bind mouse motion to change cursor over links
        self.task_tree.bind("<Motion>", self.update_cursor)
        
        # Configure tag for links - use bold font for consistency
        self.task_tree.tag_configure("link", foreground="#0066cc", font=standard_font)
        
        # Configure tag for link hover effect
        self.task_tree.tag_configure("link_hover", foreground="#0099ff", font=standard_font)
        
        # Configure tag for link click effect
        self.task_tree.tag_configure("link_click", foreground="#003366", font=standard_font)
        
        # Configure tag for completed tasks with gray text and background
        self.task_tree.tag_configure("completed", foreground="#9e9e9e", background="#e0e0e0", font=standard_font)  # Gray text and light gray background
    
    def show_loading_indicator(self):
        """Show the loading indicator over the treeview."""
        if hasattr(self, 'loading_frame'):
            self.loading_frame.place(relx=0.5, rely=0.5, anchor=tk.CENTER, width=300, height=120)
            self.root.update_idletasks()
            self._animate_loading()
    
    def _animate_loading(self):
        """Animate the loading indicator."""
        if not hasattr(self, 'loading_animation_active'):
            self.loading_animation_active = True
            self.loading_dots = 0
        
        if self.loading_animation_active and hasattr(self, 'loading_label'):
            dots = '.' * (self.loading_dots % 4)
            self.loading_label.config(text=f"⏳ Loading tasks{dots}")
            self.loading_dots += 1
            self.loading_animation_timer = self.root.after(500, self._animate_loading)
    
    def hide_loading_indicator(self):
        """Hide the loading indicator."""
        self.loading_animation_active = False
        if hasattr(self, 'loading_animation_timer'):
            self.root.after_cancel(self.loading_animation_timer)
        if hasattr(self, 'loading_frame'):
            self.loading_frame.place_forget()
            self.root.update_idletasks()

    def _get_due_date_color(self, due_date_str):
        """Calculate background color based on due date proximity."""
        if not due_date_str:
            return "#ffffff"  # White for no due date

        try:
            due_date = datetime.datetime.strptime(due_date_str, "%Y-%m-%d").date()
            today = datetime.date.today()
            days_until_due = (due_date - today).days

            if days_until_due < 0:
                return "#ffcccc"  # More intense light red for overdue
            elif days_until_due == 0:
                return "#ffd9d9"  # More intense red for due today
            elif days_until_due <= 3:
                return "#ffe6e6"  # Light red for near due
            elif days_until_due <= 7:
                return "#fff2cc"  # More intense light yellow for upcoming
            elif days_until_due <= 14:
                return "#e6ffcc"  # More intense light yellow-green
            else:
                return "#ccffcc"  # More intense light green for far future
        except ValueError:
            return "#ffffff"  # White for invalid date format

    def create_control_panel(self):
        """Create the control panel with buttons."""
        # Create a container with a subtle border
        control_frame = ttk.Frame(self.main_frame, style="TFrame")
        control_frame.pack(fill=tk.X, pady=10)
        
        # Main button frame
        button_frame = ttk.Frame(control_frame, style="TFrame")
        button_frame.pack(side=tk.LEFT)
        
        # Add task button
        add_btn = ttk.Button(
            button_frame, 
            text=" Add Task", 
            command=self.add_task,
            style="TButton"
        )
        add_btn.pack(side=tk.LEFT, padx=5)
        
        # Undo button
        self.undo_btn = ttk.Button(
            button_frame,
            text=" Undo",
            command=self.undo_delete,
            state=tk.DISABLED,  # Initially disabled
            style="TButton"
        )
        self.undo_btn.pack(side=tk.LEFT, padx=5)
        
        # Refresh button (reload from database)
        refresh_btn = ttk.Button(
            button_frame, 
            text="Reload", 
            command=lambda: self.refresh_task_list(reload_from_db=True),
            style="TButton"
        )
        refresh_btn.pack(side=tk.LEFT, padx=5)
        
        # Scheduler button
        scheduler_btn = ttk.Button(
            button_frame,
            text="Scheduler",
            command=self.open_scheduler,
            style="TButton"
        )
        scheduler_btn.pack(side=tk.LEFT, padx=5)
        
        # Search bar frame
        search_frame = ttk.Frame(control_frame, style="TFrame")
        search_frame.pack(side=tk.RIGHT, padx=(20, 0))
        
        # Search label
        search_label = ttk.Label(search_frame, text="Search:", style="TLabel")
        search_label.pack(side=tk.LEFT, padx=(0, 5))
        
        # Search entry
        self.search_var = tk.StringVar()
        self.search_entry = ttk.Entry(
            search_frame, 
            textvariable=self.search_var, 
            width=20,
            font=('Helvetica', 10)
        )
        self.search_entry.pack(side=tk.LEFT)
        
        # Bind search events
        self.search_var.trace('w', self.on_search_change)
        self.search_entry.bind('<Return>', lambda e: self.refresh_task_list())
        self.search_entry.bind('<Escape>', self.clear_search)
        
        # Clear search button - smaller size to match entry height
        clear_search_btn = ttk.Button(
            search_frame,
            text="✕",
            command=self.clear_search,
            width=2,  # Reduced width
            style="Compact.TButton"
        )
        clear_search_btn.pack(side=tk.LEFT, padx=(2, 0))
    
    def create_status_bar(self):
        """Create the status bar at the bottom of the main window."""
        status_frame = ttk.Frame(self.root, relief=tk.FLAT, padding="2 2 2 2")
        status_frame.pack(side=tk.BOTTOM, fill=tk.X)
        
        separator = ttk.Separator(status_frame, orient=tk.HORIZONTAL)
        separator.pack(fill=tk.X, pady=(0, 2))
        
        # Status bar with task statistics
        self.status_var = tk.StringVar()
        status_bar = ttk.Label(
            status_frame, 
            textvariable=self.status_var, 
            anchor=tk.W,
            style="Status.TLabel"
        )
        status_bar.pack(side=tk.LEFT, fill=tk.X, expand=True)
        
        # Add user info to the right side of the status bar
        # Get Windows username (TCN)
        windows_username = self.task_manager._get_current_user()
        
        # Look up display name from USER_MAPPING or database
        display_name = windows_username  # Default to Windows username if not found
        
        # First try USER_MAPPING (hardcoded users)
        for name, data in USER_MAPPING.items():
            if data.get("tcn") == windows_username:
                display_name = name
                break
        
        # If not found in USER_MAPPING, try loading from database
        if display_name == windows_username and self.task_manager.conn:
            try:
                users_from_db = self.task_manager.load_users_from_db()
                for name, data in users_from_db.items():
                    if data.get("tcn") == windows_username:
                        display_name = name
                        break
            except Exception as e:
                logging.debug(f"Could not load users from database for display: {e}")

        user_label = ttk.Label(
            status_frame, 
            text=f"User: {display_name}",
            anchor=tk.E,
            style="Status.TLabel"
        )
        user_label.pack(side=tk.RIGHT, padx=10)
        
        # Initialize status
        self.update_status()
    
    def update_status(self):
        """Update the status bar with task statistics."""
        # Get filtered non-deleted tasks based on active filters
        category = None if self.category_var.get() == "All" else self.category_var.get()
        main_staff = None if self.main_staff_var.get() == "All" else self.main_staff_var.get()
        assigned_to = None if self.user_var.get() == "All" else self.user_var.get()
        
        # Get non-deleted tasks with current filters applied
        filtered_tasks = self.task_manager.get_filtered_tasks(
            show_completed=True,
            category=category,
            main_staff=main_staff,
            assigned_to=assigned_to,
            show_deleted=False  # Explicitly exclude deleted tasks
        )
        
        total = len(filtered_tasks)
        completed = sum(1 for task in filtered_tasks if task["completed"])
        pending = total - completed
        
        # Create a more informative status message (total now excludes completed tasks)
        status_text = f"Total Tasks: {pending} | Completed: {completed} | Pending: {pending}"
        
        # Add filter information if filters are active
        filters_active = []
        
        if self.category_var.get() != "All":
            filters_active.append(f"Category: {self.category_var.get()}")
            
        if self.main_staff_var.get() != "All":
            filters_active.append(f"Main Staff: {self.main_staff_var.get()}")
            
        if self.user_var.get() != "All":
            filters_active.append(f"Assigned To: {self.user_var.get()}")
        
        if filters_active:
            status_text += " | Filters: " + ", ".join(filters_active)
        
        # Add search information if search is active
        if hasattr(self, 'search_var') and self.search_var.get().strip():
            search_term = self.search_var.get().strip()
            status_text += f" | Search: '{search_term}'"
        
        self.status_var.set(status_text)
    
    def refresh_task_list(self, reload_from_db=False):
        """Refresh the task list in the treeview.
        
        Args:
            reload_from_db: If True, reload tasks from database first (shows loading indicator)
        """
        if reload_from_db:
            # Only show loading indicator when actually reloading from database
            with self.refresh_lock:
                if self.refresh_in_progress:
                    return
                self.refresh_in_progress = True
            
            self.set_status("Reloading from database...")
            self.show_loading_indicator()
            threading.Thread(target=self._perform_task_refresh_threaded, daemon=True).start()
        else:
            # Instant refresh from memory - no loading indicator needed
            self._perform_task_refresh()

    def _perform_task_refresh_threaded(self):
        """Threaded method to perform task refresh with database reload."""
        try:
            self._perform_task_refresh(reload_from_db=True) # Reload from database
        except Exception as e:
            logging.error("Error during task refresh in thread", exc_info=True)
            # Handle errors that occur during task refresh in the background thread
            self.root.after(0, lambda: messagebox.showerror("Error", f"Error refreshing task list: {str(e)}", parent=self.root)) # Use lambda for deferred execution
        finally:
            # Reset the refresh flag and ensure loading indicator is hidden even if there's an error
            with self.refresh_lock:
                self.refresh_in_progress = False
            self.root.after(0, self.hide_loading_indicator) # Use root.after to run in main thread
            self.root.after(0, self.update_status) # Update status bar in main thread

    def _perform_task_refresh(self, reload_from_db=False):
        """Perform the actual task refresh after UI has updated.
        
        Args:
            reload_from_db: If True, reload from database first
        """
        try:
            # Only reload from database if explicitly requested
            if reload_from_db:
                try:
                    self.task_manager.reload_tasks()
                except Exception as e:
                    logging.error(f"Error reloading tasks: {str(e)}", exc_info=True)
                    raise Exception(f"Database error: {str(e)}")

            # Clear existing items in the treeview
            for item in self.task_tree.get_children():
                self.task_tree.delete(item)

            # Get filtered tasks
            show_completed = self.show_completed_var.get()
            category = None if self.category_var.get() == "All" else self.category_var.get()
            main_staff = None if self.main_staff_var.get() == "All" else self.main_staff_var.get()
            assigned_to = None if self.user_var.get() == "All" else self.user_var.get()

            # Get search term if available
            search_term = ""
            if hasattr(self, 'search_var'):
                search_term = self.search_var.get().lower().strip()

            # Get filtered tasks
            filtered_tasks = self.task_manager.get_filtered_tasks(
                show_completed=show_completed,
                category=category,
                main_staff=main_staff,
                assigned_to=assigned_to
            )

            # Apply search filter if needed
            if search_term:
                filtered_tasks = [
                    task for task in filtered_tasks if self._task_matches_search(task, search_term)
                ]

            # Apply sorting if a sort column is set
            if self.sort_column:
                self.apply_sorting(filtered_tasks)

            # Track the maximum status text length to adjust column width
            max_status_length = 0

            # Define a standard font for all task items - make it bold
            standard_font = ('TkDefaultFont', 9, 'bold')

            # Add tasks to treeview with improved visual indicators
            for task in filtered_tasks:
                # Create status with time remaining - using standardized format for better alignment
                if task["completed"]:
                    status = "✓ Completed"
                    status_tags = ["completed"]
                else:
                    # Add due date information for pending tasks
                    if task["due_date"]:
                        try:
                            due_date = datetime.datetime.strptime(task["due_date"], "%Y-%m-%d").date()
                            today = datetime.date.today()
                            days_until_due = (due_date - today).days
                            
                            if days_until_due < 0:
                                # Use consistent format for overdue
                                status = f"! {abs(days_until_due)} Days Overdue"
                                status_tags = ["overdue"]
                            elif days_until_due == 0:
                                status = "! Due Today"
                                status_tags = ["due_today"]
                            elif days_until_due == 1:
                                status = "! Due Tomorrow"
                                status_tags = ["pending_soon"]
                            else:
                                # Use consistent format for days left
                                status = f"• {days_until_due} Days Left"
                                status_tags = ["pending"]
                                if days_until_due <= 3:
                                    status_tags = ["pending_soon"]
                        except ValueError:
                            status = "• No Due Date"
                            status_tags = ["pending"]
                    else:
                        status = "• No Due Date"
                        status_tags = ["pending"]
                
                # Update max status length for column width adjustment
                max_status_length = max(max_status_length, len(status))
                
                # Format due date
                due_date = task["due_date"] if task["due_date"] else "No due date"
                
                # Format staff assignments
                main_staff = task.get("main_staff", "Unassigned")
                assigned_to = task.get("assigned_to", "Unassigned")
                
                # Get values for new columns (with defaults if not present)
                applied_vessel = task.get("applied_vessel", "")
                rev = task.get("rev", "")
                
                # Set row tags for styling based on priority and completion
                if task["completed"]:
                    # For completed tasks, only use the completed tag to ensure it overrides all other styles
                    tags = ["completed"]
                else:
                    # For pending tasks, use priority and status tags
                    tags = [task["priority"]] + status_tags
                
                # Format priority with colored symbols
                priority_display = {
                    "high": "🔴 High",
                    "medium": "🟠 Medium",
                    "low": "🟢 Low"
                }.get(task["priority"].lower(), task["priority"].capitalize())
                
                # Create link display
                link_display = "🔗" if task.get("link") else ""
                
                # Create a unique tag for this task's due date color
                due_date_tag = f"due_date_{task['id']}"
                tags.append(due_date_tag)
                
                # Configure the tag with the appropriate background color
                bg_color = self._get_due_date_color(task["due_date"])
                self.task_tree.tag_configure(due_date_tag, background=bg_color)
                
                # For completed tasks, make sure "completed" is the first tag to ensure it has priority
                if task["completed"]:
                    # Remove any status tags that might override the completed style
                    tags = [tag for tag in tags if tag not in ["overdue", "due_today", "pending_soon", "pending"]]
                    # Put completed tag first to ensure it has priority
                    tags.insert(0, "completed")
                
                # Add link tag if there's a link
                if task.get("link"):
                    tags.append("link")
                
                self.task_tree.insert(
                    "", tk.END, 
                    values=(
                        status,
                        task["title"],
                        str(task.get("rev", "")) if task.get("rev") is not None else "",  # Show "0" if rev is 0
                        str(task.get("applied_vessel", "")),
                        priority_display,
                        str(task.get("main_staff", "")),
                        str(task.get("assigned_to", "")),
                        str(task.get("qtd_mhr", "")) if task.get("qtd_mhr") is not None else "",  # Show 0 if it's 0
                        str(task.get("actual_mhr", "")) if task.get("actual_mhr") is not None else "",  # Show 0 if it's 0
                        task.get("target_start", "") or "",  # Target Start date
                        task.get("target_finish", "") or ""  # Target Finish date
                    ),
                    tags=tags,
                    iid=str(task["id"])  # Use task ID as the item ID for direct lookup
                )
            
            # Adjust the status column width based on content
            # Calculate pixel width: approximate 8 pixels per character plus some padding
            if max_status_length > 0:
                # Use a multiplier for average character width (depends on font)
                char_width = 8  # Approximate width in pixels per character
                padding = 20    # Extra padding for the column
                new_width = max(150, min(300, max_status_length * char_width + padding))
                self.task_tree.column("status", width=new_width, minwidth=150)
            
            # Configure tag colors for status indicators with consistent bold font
            self.task_tree.tag_configure("completed", foreground="#9e9e9e", font=standard_font)  # Gray text
            self.task_tree.tag_configure("overdue", foreground="#d32f2f", font=standard_font)  # Red text
            self.task_tree.tag_configure("due_today", foreground="#f57c00", font=standard_font)  # Orange text
            self.task_tree.tag_configure("pending_soon", foreground="#7b1fa2", font=standard_font)  # Purple text
            self.task_tree.tag_configure("pending", foreground="#1976d2", font=standard_font)  # Blue text
            
            # Configure priority tags with consistent bold font
            self.task_tree.tag_configure("high", font=standard_font)
            self.task_tree.tag_configure("medium", font=standard_font)
            self.task_tree.tag_configure("low", font=standard_font)
            
            # Configure link tag with consistent bold font but still make it stand out
            self.task_tree.tag_configure("link", foreground="#0066cc", font=standard_font)
            self.task_tree.tag_configure("link_hover", foreground="#0099ff", font=standard_font)
            self.task_tree.tag_configure("link_click", foreground="#003366", font=standard_font)
            
            # Show "no results" message if no tasks match the filters
            if not filtered_tasks:
                self.task_tree.insert("", tk.END, values=("No tasks match your filters", "", "", "", "", "", "", "", ""), tags=["no_results"])
                self.task_tree.tag_configure("no_results", foreground="#9e9e9e", font=standard_font)
            
        except Exception as e:
            logging.error("Error during _perform_task_refresh", exc_info=True)
            # Re-raise to be caught in _perform_task_refresh_threaded - No, handle here and show error
            self.root.after(0, lambda: messagebox.showerror("Error", "Error refreshing task list. See log for details.", parent=self.root)) # Show error in main thread
        finally:
            # Hide loading indicator and update status are now handled in _perform_task_refresh_threaded
            pass # No changes needed here
    
    def update_category_filter(self):
        """Update the category filter dropdown with available categories."""
        categories = set(task["category"] for task in self.task_manager.tasks)
        categories = sorted(list(categories))
        
        # Add "All" option
        categories = ["All"] + categories
        
        # Update combobox values
        current = self.category_var.get()
        self.category_combo["values"] = categories
        
        # If current category is not in the list, reset to "All"
        if current not in categories:
            self.category_var.set("All")
    
    def get_selected_task_id(self):
        """Get the ID of the selected task (single selection)."""
        selection = self.task_tree.selection()
        if not selection:
            messagebox.showinfo("Information", "Please select a task first.")
            return None
        
        # Get the task ID directly from the item ID
        try:
            task_id = int(selection[0])
            return task_id
        except (ValueError, TypeError):
            # Fallback to the old method if the item ID is not a valid task ID
            item = selection[0]
            values = self.task_tree.item(item, "values")
            
            # Find the task with matching values
            for task in self.task_manager.tasks:
                # Match by title, rev, and applied_vessel
                if (task["title"] == values[1] and
                    task.get("rev", "") == values[2] and
                    task.get("applied_vessel", "") == values[3]):
                    return task["id"]
            
            messagebox.showerror("Error", "Could not identify the selected task.")
            return None
    
    def get_selected_task_ids(self):
        """Get the IDs of the selected tasks."""
        selection = self.task_tree.selection()
        if not selection:
            messagebox.showinfo("Information", "Please select at least one task.")
            return None
        
        # Get the IDs directly from the item IDs
        task_ids = []
        for item in selection:
            try:
                task_id = int(item)
                task_ids.append(task_id)
            except (ValueError, TypeError):
                # Fallback to the old method if the item ID is not a valid task ID
                values = self.task_tree.item(item, "values")
                
                # Find the task with matching values
                for task in self.task_manager.tasks:
                    # Match by title, rev, and applied_vessel
                    if (task["title"] == values[1] and
                        task.get("rev", "") == values[2] and
                        task.get("applied_vessel", "") == values[3]):
                        task_ids.append(task["id"])
                        break
        
        if not task_ids:
            messagebox.showerror("Error", "Could not identify the selected tasks.")
            return None
            
        return task_ids
    
    def add_task(self):
        """Open dialog to add a new task."""
        dialog = TaskDialog(self.root, "Add Task")
        if dialog.result:
            # Attempt to add the task using the TaskManager
            # TaskManager.add_task now returns the new ID on success, or None on failure
            new_task_id = self.task_manager.add_task(**dialog.result)

            if new_task_id is not None:
                # Success: Task was added to DB and local list - instant refresh from memory
                self._perform_task_refresh()
                self.set_status(f"Task '{dialog.result['title']}' (ID: {new_task_id}) added successfully.")
                self.update_category_filter() # Update filters if a new category was added
            else:
                # Failure: TaskManager.add_task already logged the error
                messagebox.showerror("Error Adding Task",
                                     "Failed to save the task to the database.\n"
                                     "This might be due to a temporary network issue or database contention.\n\n"
                                     "Please try adding the task again.",
                                     parent=self.root)
                # No refresh needed as the task wasn't added successfully
    
    def edit_task(self):
        """Edit the selected task."""

        task_id = self.get_selected_task_id()
        if not task_id:
            return
        
        # Get the task
        task = self.task_manager._find_task_by_id(task_id)
        if not task:
            messagebox.showerror("Error", f"Task with ID {task_id} not found.")
            return
        
        # Open dialog with task data
        dialog = TaskDialog(self.root, "Edit Task", task)
        if dialog.result:
            # Update the task
            self.task_manager.update_task(task_id, **dialog.result)
            
            # Instant refresh from memory
            self._perform_task_refresh()
            
            # Show confirmation message
            self.set_status(f"Task {task_id} updated successfully.")
    
    def complete_task(self):
        """Mark the selected task(s) as completed or pending."""
        task_ids = self.get_selected_task_ids()
        if not task_ids:
            return
        
        # Group tasks by current completion status and toggle each group
        tasks = [self.task_manager._find_task_by_id(task_id) for task_id in task_ids]
        
        completed_task_ids = [t["id"] for t in tasks if t and t["completed"]]
        pending_task_ids = [t["id"] for t in tasks if t and not t["completed"]]
        
        # Toggle each group independently
        if completed_task_ids:
            self.batch_update_completion_status(completed_task_ids, completed=False)
        if pending_task_ids:
            self.batch_update_completion_status(pending_task_ids, completed=True)
        
        count = len(task_ids)
        self.set_status(f"{count} task{'s' if count > 1 else ''} toggled.")
        
        # Refresh immediately from memory (not from database)
        self._perform_task_refresh()
    
    def batch_update_completion_status(self, task_ids, completed):
        """Update completion status for multiple tasks at once."""
        for task_id in task_ids:
            self.task_manager.update_task(task_id, completed=completed)
        
        # No need to call refresh_task_list() here as it will be called by the calling method
    
    def delete_task(self):
        """Delete the selected task(s) using soft delete."""
        task_ids = self.get_selected_task_ids()
        if not task_ids:
            return

        count = len(task_ids)
        message = (f"Are you sure you want to delete {count} task{'s' if count > 1 else ''}?\n\n"
                   "Deleted tasks can be recovered later from the 'View Deleted Tasks' option.")

        if not messagebox.askyesno("Confirm Deletion", message, parent=self.root):
            return

        # For single task deletion:
        if count == 1:
            task_id_to_delete = task_ids[0]
            # Call the newly added delete_task method in TaskManager
            if self.task_manager.delete_task(task_id_to_delete):
                # Successfully soft-deleted - instant refresh from memory
                self._perform_task_refresh()
                self.set_status(f"Task {task_id_to_delete} marked as deleted.")
                # Enable the Undo button after deletion
                if hasattr(self, 'undo_btn'):
                    self.undo_btn.config(state=tk.NORMAL)
            else:
                # Deletion failed (error message likely shown by TaskManager)
                 self.set_status(f"Failed to delete task {task_id_to_delete}.")

        else:
            # For batch deletion:
            # Call the newly added batch_delete_tasks method
            succeeded, failed = self.task_manager.batch_delete_tasks(task_ids)
            if succeeded:
                # Instant refresh from memory
                self._perform_task_refresh()
                self.set_status(f"Marked {len(succeeded)} task{'s' if len(succeeded) > 1 else ''} as deleted.")
                # Enable the Undo button after batch deletion
                if hasattr(self, 'undo_btn'):
                    self.undo_btn.config(state=tk.NORMAL)
            if failed:
                messagebox.showwarning("Partial Deletion",
                                     f"{len(failed)} task{'s' if len(failed) > 1 else ''} could not be deleted. Check logs for details.",
                                     parent=self.root)
            if not succeeded and not failed:
                 # This case might happen if all task_ids were invalid or a major DB error occurred
                 self.set_status(f"Failed to delete selected tasks.")
    
    def _select_all_tasks(self, event=None):
        """Select all tasks in the task tree."""
        # Get all items in the treeview
        all_items = self.task_tree.get_children()
        if all_items:
            # Select all items
            self.task_tree.selection_set(all_items)
        return "break"  # Prevent default behavior
    
    def _delete_task_via_keyboard(self, event=None):
        """Delete selected tasks via Delete key."""
        self.delete_task()
        return "break"  # Prevent default behavior
    
    def view_task_details(self, event):
        """View details of the selected task."""
        # If called from an event, get the task ID from the clicked item
        if event:
            item = self.task_tree.identify_row(event.y)
            if not item:
                return
            self.task_tree.selection_set(item)
        
        task_id = self.get_selected_task_id()
        if not task_id:
            return
        
        # Get the task - reload tasks first to ensure we have the latest data
        self.task_manager.reload_tasks()
        task = self.task_manager._find_task_by_id(task_id)
        if not task:
            messagebox.showerror("Error", f"Task with ID {task_id} not found.")
            return
        
        # Create a fixed-size dialog window
        details_window = tk.Toplevel(self.root)
        
        # Hide window during setup to prevent flash
        details_window.withdraw()
        
        details_window.title("Task Details")
        details_window.geometry("500x600")
        details_window.minsize(500, 600)
        details_window.maxsize(500, 600)
        details_window.transient(self.root)
        details_window.grab_set()
        
        # Set icon for details window
        try:
            icon_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'TaskManager_main.png')
            if os.path.exists(icon_path):
                icon_image = tk.PhotoImage(file=icon_path)
                details_window.iconphoto(False, icon_image)
                details_window.details_icon = icon_image
        except Exception as e:
            logging.warning(f"Could not set task details window icon: {e}")
        
        # Create main frame with padding
        main_frame = ttk.Frame(details_window, padding=15)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Create a canvas with scrollbar for content
        canvas = tk.Canvas(main_frame, highlightthickness=0)
        scrollbar = ttk.Scrollbar(main_frame, orient=tk.VERTICAL, command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas, padding=(0, 0, 15, 0))  # Add right padding to avoid scrollbar overlap
        
        def on_frame_configure(e):
            canvas.configure(scrollregion=canvas.bbox("all"))
            # Update the width of the window to account for scrollbar
            canvas.itemconfig(canvas_window, width=canvas.winfo_width() - 5)
        
        scrollable_frame.bind("<Configure>", on_frame_configure)
        
        canvas_window = canvas.create_window((0, 0), window=scrollable_frame, anchor=tk.NW)
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Pack canvas and scrollbar
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Format dates for display
        request_date = task.get('requested_date', 'Not specified')
        due_date = task.get('due_date', 'No due date')
        created_date = task.get('created_date', 'Unknown')
        last_modified = task.get('last_modified', 'Not modified')
        modified_by = task.get('modified_by', 'Not modified')
        
        # Add details with labels in scrollable frame
        details_data = [
            ("Equipment Name:", task['title']),
            ("Request No:", task.get('request_no', 'Not specified')),
            ("Rev:", task.get('rev', 'Not specified')),
            ("Applied Vessel:", task.get('applied_vessel', 'Not specified')),
            ("Drawing No:", task.get('drawing_no', 'Not specified')),
            ("Description:", task['description'] or 'No description'),
            ("Request Date:", request_date),
            ("Due Date:", due_date),
            ("Target Start:", task.get('target_start', 'Not specified')),
            ("Target Finish:", task.get('target_finish', 'Not specified')),
            ("Priority:", task.get('priority', 'Not specified')),
            ("Category:", task['category']),
            ("Main Staff:", task.get('main_staff', 'Not specified')),
            ("Assigned To:", task.get('assigned_to', 'Not specified')),
            ("Status:", "Completed" if task.get('completed') else "Active"),
            ("Qtd Mhr:", str(task.get('qtd_mhr', 'Not specified'))),
            ("Actual Mhr:", str(task.get('actual_mhr', 'Not specified'))),
            ("Link:", task.get('link', 'Not specified')),
            ("SDB Link:", task.get('sdb_link', 'Not specified')),
            ("Created Date:", created_date),
            ("Last Modified:", last_modified),
            ("Modified By:", modified_by)
        ]
        
        for label_text, value_text in details_data:
            # Create frame for each field
            field_frame = ttk.Frame(scrollable_frame)
            field_frame.pack(fill=tk.X, pady=5)
            
            # Label
            label = ttk.Label(field_frame, text=label_text, font=('Helvetica', 10, 'bold'))
            label.pack(anchor=tk.W)
            
            # Value (use Text widget for long content)
            if label_text in ("Description:", "Link:", "SDB Link:"):
                value_text_widget = tk.Text(field_frame, height=3, wrap=tk.WORD, font=('Helvetica', 9))
                value_text_widget.insert("1.0", str(value_text))
                value_text_widget.config(state=tk.DISABLED)
                value_text_widget.pack(fill=tk.X, pady=(2, 0))
            else:
                value_label = ttk.Label(field_frame, text=str(value_text), font=('Helvetica', 9))
                value_label.pack(anchor=tk.W, pady=(2, 0))
        
        # Add Close button at bottom
        button_frame = ttk.Frame(details_window, padding=10)
        button_frame.pack(fill=tk.X, side=tk.BOTTOM)
        ttk.Button(button_frame, text="Close", command=details_window.destroy).pack(side=tk.RIGHT)
        
        # Center the window relative to the main application window
        details_window.update_idletasks()
        
        # Get the toplevel parent window (main application window, not just self.root)
        parent_window = self.root.winfo_toplevel()
        
        # Calculate center position relative to parent
        parent_x = parent_window.winfo_x()
        parent_y = parent_window.winfo_y()
        parent_width = parent_window.winfo_width()
        parent_height = parent_window.winfo_height()
        
        x = parent_x + (parent_width - 500) // 2
        y = parent_y + (parent_height - 600) // 2
        
        details_window.geometry(f"+{x}+{y}")
        
        # Bind mousewheel for scrolling
        def on_mousewheel(event):
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        
        canvas.bind_all("<MouseWheel>", on_mousewheel)
        details_window.protocol("WM_DELETE_WINDOW", lambda: [canvas.unbind_all("<MouseWheel>"), details_window.destroy()])
        
        # Show window after everything is set up
        details_window.deiconify()
    
    def show_context_menu(self, event):
        """Show the context menu for tasks."""
        # Get the item under cursor
        item = self.task_tree.identify_row(event.y)
        if not item:
            return
        
        # Select the item if it's not already selected
        if item not in self.task_tree.selection():
            self.task_tree.selection_set(item)
        
        # Get the number of selected items
        selection_count = len(self.task_tree.selection())
        
        # Create context menu
        context_menu = tk.Menu(self.root, tearoff=0)
        
        # Direct menu items (no submenu)
        if selection_count == 1:
            context_menu.add_command(label="View Details", command=lambda: self.view_task_details(None))
            context_menu.add_command(label="Edit", command=self.edit_task)
            
            # Dynamic Complete/Pending label based on task status
            try:
                task_id = int(item)
                task = self.task_manager._find_task_by_id(task_id)
                if task:
                    if task.get("completed"):
                        context_menu.add_command(label="Pending", command=self.complete_task)
                    else:
                        context_menu.add_command(label="Complete", command=self.complete_task)
                else:
                    context_menu.add_command(label="Toggle Complete/Pending", command=self.complete_task)
            except (ValueError, TypeError):
                context_menu.add_command(label="Toggle Complete/Pending", command=self.complete_task)
        else:
            context_menu.add_command(label=f"Toggle {selection_count} Tasks", command=self.complete_task)
        
        context_menu.add_separator()
        
        # Add Assign To submenu
        assign_menu = tk.Menu(context_menu, tearoff=0)
        for staff in USERS[1:]:  # Skip "All" from the USERS list
            assign_menu.add_command(
                label=staff,
                command=lambda s=staff: self.assign_task_to(s)
            )
        context_menu.add_cascade(label="Assign To", menu=assign_menu)
        
        # Copy options submenu (only for single selection)
        if selection_count == 1:
            try:
                task_id = int(item)
                task = self.task_manager._find_task_by_id(task_id)
                if task:
                    # Create Copy submenu
                    copy_menu = tk.Menu(context_menu, tearoff=0)
                    has_copy_items = False
                    
                    # Add copy options for hidden fields
                    if task.get("drawing_no"):
                        copy_menu.add_command(
                            label="Copy Drawing No.", 
                            command=lambda: self.copy_to_clipboard(task.get("drawing_no", ""))
                        )
                        has_copy_items = True
                    
                    if task.get("request_no"):
                        copy_menu.add_command(
                            label="Copy Request No.", 
                            command=lambda: self.copy_to_clipboard(task.get("request_no", ""))
                        )
                        has_copy_items = True
                    
                    if task.get("link"):
                        copy_menu.add_command(
                            label="Copy Folder Link", 
                            command=lambda: self.copy_to_clipboard(task.get("link", ""))
                        )
                        has_copy_items = True
                    
                    if task.get("sdb_link"):
                        copy_menu.add_command(
                            label="Copy SDB Link", 
                            command=lambda: self.copy_to_clipboard(task.get("sdb_link", ""))
                        )
                        has_copy_items = True
                    
                    # Add dates submenu if any dates are present
                    dates = {
                        "Requested Date": task.get("requested_date"),
                        "Due Date": task.get("due_date"),
                        "Target Start": task.get("target_start"),
                        "Target Finish": task.get("target_finish")
                    }
                    
                    if any(dates.values()):
                        dates_menu = tk.Menu(copy_menu, tearoff=0)
                        for label, date in dates.items():
                            if date:
                                dates_menu.add_command(
                                    label=f"Copy {label}", 
                                    command=lambda d=date: self.copy_to_clipboard(d)
                                )
                        copy_menu.add_cascade(label="Copy Dates", menu=dates_menu)
                        has_copy_items = True
                    
                    # Only add the Copy menu if there are items in it
                    if has_copy_items:
                        context_menu.add_cascade(label="Copy", menu=copy_menu)
                    
                    # Create Open submenu
                    open_menu = tk.Menu(context_menu, tearoff=0)
                    has_open_items = False
                    
                    if task.get("link"):
                        open_menu.add_command(
                            label="Open Folder Link", 
                            command=lambda: self.open_link(task.get("link", ""))
                        )
                        has_open_items = True
                    
                    if task.get("sdb_link"):
                        open_menu.add_command(
                            label="Open SDB Link", 
                            command=lambda: self.open_link(task.get("sdb_link", ""))
                        )
                        has_open_items = True
                    
                    # Only add the Open menu if there are items in it
                    if has_open_items:
                        context_menu.add_cascade(label="Open", menu=open_menu)
                    
            except (ValueError, TypeError):
                # If we can't get the task ID directly, try to find the task by values
                values = self.task_tree.item(item, "values")
                if len(values) >= 4:  # Make sure we have enough values
                    for task in self.task_manager.tasks:
                        # Match by title, rev, and applied_vessel
                        if (task["title"] == values[1] and
                            task.get("rev", "") == values[2] and
                            task.get("applied_vessel", "") == values[3]):
                            
                            # Create Copy submenu
                            copy_menu = tk.Menu(context_menu, tearoff=0)
                            has_copy_items = False
                            
                            # Add copy options for hidden fields
                            if task.get("drawing_no"):
                                copy_menu.add_command(
                                    label="Copy Drawing No.", 
                                    command=lambda: self.copy_to_clipboard(task.get("drawing_no", ""))
                                )
                                has_copy_items = True
                            
                            if task.get("request_no"):
                                copy_menu.add_command(
                                    label="Copy Request No.", 
                                    command=lambda: self.copy_to_clipboard(task.get("request_no", ""))
                                )
                                has_copy_items = True
                            
                            if task.get("link"):
                                copy_menu.add_command(
                                    label="Copy Folder Link", 
                                    command=lambda: self.copy_to_clipboard(task.get("link", ""))
                                )
                                has_copy_items = True
                            
                            if task.get("sdb_link"):
                                copy_menu.add_command(
                                    label="Copy SDB Link", 
                                    command=lambda: self.copy_to_clipboard(task.get("sdb_link", ""))
                                )
                                has_copy_items = True
                            
                            # Add dates submenu if any dates are present
                            dates = {
                                "Requested Date": task.get("requested_date"),
                                "Due Date": task.get("due_date"),
                                "Target Start": task.get("target_start"),
                                "Target Finish": task.get("target_finish")
                            }
                            
                            if any(dates.values()):
                                dates_menu = tk.Menu(copy_menu, tearoff=0)
                                for label, date in dates.items():
                                    if date:
                                        dates_menu.add_command(
                                            label=f"Copy {label}", 
                                            command=lambda d=date: self.copy_to_clipboard(d)
                                        )
                                copy_menu.add_cascade(label="Copy Dates", menu=dates_menu)
                                has_copy_items = True
                            
                            # Only add the Copy menu if there are items in it
                            if has_copy_items:
                                context_menu.add_cascade(label="Copy", menu=copy_menu)
                            
                            # Create Open submenu
                            open_menu = tk.Menu(context_menu, tearoff=0)
                            has_open_items = False
                            
                            if task.get("link"):
                                open_menu.add_command(
                                    label="Open Folder Link", 
                                    command=lambda: self.open_link(task.get("link", ""))
                                )
                                has_open_items = True
                            
                            if task.get("sdb_link"):
                                open_menu.add_command(
                                    label="Open SDB Link", 
                                    command=lambda: self.open_link(task.get("sdb_link", ""))
                                )
                                has_open_items = True
                            
                            # Only add the Open menu if there are items in it
                            if has_open_items:
                                context_menu.add_cascade(label="Open", menu=open_menu)
                            
                            break
        
        context_menu.add_separator()
        
        # Delete option shows count of items to be deleted
        if selection_count == 1:
            context_menu.add_command(label="Delete Task", command=self.delete_task)
        else:
            context_menu.add_command(label=f"Delete {selection_count} Tasks", command=self.delete_task)
        
        # Add a 'View Deleted Tasks' option
        context_menu.add_separator()
        context_menu.add_command(label="View Deleted Tasks", command=self.view_deleted_tasks)
        
        # Display the context menu
        context_menu.tk_popup(event.x_root, event.y_root)
    
    def check_link_click(self, event):
        """Check if a link cell was clicked and open the link if so."""
        region = self.task_tree.identify_region(event.x, event.y)
        if region == "cell":
            # Get the item and column that was clicked
            item = self.task_tree.identify_row(event.y)
            column = self.task_tree.identify_column(event.x)
            
            # Check if the link column was clicked (column #8)
            if column == "#8":  # Link column
                # Get the values from the display
                values = self.task_tree.item(item, "values")
                link_display = values[7] if len(values) > 7 else ""
                
                # Only proceed if it contains a link indicator
                if link_display == "🔗":
                    # Get the actual task to find the real link
                    try:
                        task_id = int(item)
                        task = self.task_manager._find_task_by_id(task_id)
                        if task and task.get("link"):
                            # Apply click effect
                            tags = list(self.task_tree.item(item, "tags"))
                            if "link" in tags:
                                tags.remove("link")
                            if "link_hover" in tags:
                                tags.remove("link_hover")
                            tags.append("link_click")
                            self.task_tree.item(item, tags=tags)
                            
                            # Schedule to revert the effect after a short delay
                            self.root.after(150, lambda i=item: self.revert_link_click_effect(i))
                            
                            # Open the link
                            self.open_link(task["link"])
                    except (ValueError, TypeError):
                        # If we can't get the task ID directly, try to find the task
                        for task in self.task_manager.tasks:
                            # Check if this is the right task by matching other values
                            if (task["title"] == values[1] and
                                task.get("rev", "") == values[2] and
                                task.get("applied_vessel", "") == values[3]):
                                if task.get("link"):
                                    # Open the link
                                    self.open_link(task["link"])
                                break
    
    def revert_link_click_effect(self, item):
        """Revert the link click effect after a short delay."""
        if item in self.task_tree.get_children():
            tags = list(self.task_tree.item(item, "tags"))
            if "link_click" in tags:
                tags.remove("link_click")
            if "link" not in tags:
                tags.append("link")
            self.task_tree.item(item, tags=tags)
    
    def open_link(self, link):
        """Open a link which can be a URL, file path, or network path."""
        
        if not link:
            messagebox.showerror("Error", "No link provided.")
            return
            
        try:
            # Check if it's a network path or local file path
            if link.startswith("\\\\") or os.path.exists(link) or link.startswith(("C:", "D:", "E:", "F:", "\\")):
                # It's a file path - use the appropriate command based on OS
                if os.name == 'nt':  # Windows
                    # For network paths, ensure they're properly formatted
                    if link.startswith("//"):
                        # Convert forward slashes to backslashes for Windows
                        link = link.replace("/", "\\")
                    
                    # Check if it's a directory
                    if os.path.isdir(link) or (link.startswith("\\\\") and not os.path.splitext(link)[1]):
                        # Open directory in Explorer
                        subprocess.run(['explorer', link], shell=True)
                    else:
                        # Open file with default application
                        os.startfile(link)
                elif os.name == 'posix':  # macOS and Linux
                    subprocess.run(['xdg-open', link] if os.name == 'posix' else ['open', link])
            else:
                # It's a web URL - add http:// if needed
                if not link.startswith(("http://", "https://", "www.")):
                    link = f"http://{link}"
                webbrowser.open(link)
                
            self.set_status(f"Opening: {link}")
            
        except Exception as e:
            messagebox.showerror("Error", f"Could not open link: {link}\n\nError: {str(e)}")
            # Log more detailed error information
            print(f"Error opening link: {link}")
            print(f"Error details: {str(e)}")
            print(f"Error type: {type(e)}")
    
    def set_status(self, message):
        """Set a temporary status message."""
        self.status_var.set(message)
        # Schedule to reset after 5 seconds
        self.root.after(5000, self.update_status)
    
    def open_scheduler(self):
        """Open the task scheduler window with auto-assignment capabilities."""
        SchedulerDialog(self.root, self.task_manager, callback=lambda: self._perform_task_refresh())
    
    def on_search_change(self, *args):
        """Handle search text changes with debouncing."""
        # Cancel any existing search timer
        if hasattr(self, 'search_timer'):
            self.root.after_cancel(self.search_timer)
        
        # Set a new timer to refresh after 300ms of no typing
        self.search_timer = self.root.after(300, self.refresh_task_list)
    
    def clear_search(self):
        """Clear the search field and refresh the task list."""
        self.search_var.set("")
        self.refresh_task_list()
    
    def _task_matches_search(self, task, search_term):
        """Check if a task matches the search term across all relevant fields."""
        search_term = search_term.lower().strip()
        
        # Fields to search in (convert to lowercase for case-insensitive search)
        searchable_fields = [
            task.get("title", ""),
            task.get("description", ""),
            task.get("applied_vessel", ""),
            task.get("category", ""),
            task.get("main_staff", ""),
            task.get("assigned_to", ""),
            task.get("rev", ""),
            task.get("drawing_no", ""),
            task.get("request_no", ""),
            task.get("link", ""),
            task.get("sdb_link", ""),
            str(task.get("id", "")),  # Convert ID to string for searching
            str(task.get("qtd_mhr", "")),  # Convert to string for searching
            str(task.get("actual_mhr", "")),  # Convert to string for searching
        ]
        
        # Check if search term matches any field
        for field in searchable_fields:
            if field and search_term in str(field).lower():
                return True
        
        # Also check date fields if they exist
        date_fields = ["due_date", "requested_date", "target_start", "target_finish"]
        for date_field in date_fields:
            date_value = task.get(date_field)
            if date_value and search_term in str(date_value).lower():
                return True
        
        return False
    
    def update_cursor(self, event):
        """Update cursor based on whether it's over a link."""
        region = self.task_tree.identify_region(event.x, event.y)
        if region == "cell":
            column = self.task_tree.identify_column(event.x)
            item = self.task_tree.identify_row(event.y)
            
            # Reset all link items to normal link style
            for all_item in self.task_tree.get_children():
                tags = list(self.task_tree.item(all_item, "tags"))
                if "link_hover" in tags:
                    tags.remove("link_hover")
                    if "link" not in tags:
                        tags.append("link")
                    self.task_tree.item(all_item, tags=tags)
            
            # Check if over the link column (column #8)
            if column == "#8" and item:
                # Get the values from the display
                values = self.task_tree.item(item, "values")
                link_display = values[7] if len(values) > 7 else ""
                
                # Only change cursor if it contains a link indicator
                if link_display == "🔗":
                    # Get the actual task to confirm it has a link
                    try:
                        task_id = int(item)
                        task = self.task_manager._find_task_by_id(task_id)
                        if task and task.get("link"):
                            self.task_tree.config(cursor="hand2")
                            
                            # Apply hover effect to this link
                            tags = list(self.task_tree.item(item, "tags"))
                            if "link" in tags:
                                tags.remove("link")
                            if "link_hover" not in tags:
                                tags.append("link_hover")
                            self.task_tree.item(item, tags=tags)
                            return
                    except (ValueError, TypeError):
                        # Try to find the task by values
                        for task in self.task_manager.tasks:
                            if (task["title"] == values[1] and
                                task.get("rev", "") == values[2] and
                                task.get("applied_vessel", "") == values[3]):
                                if task.get("link"):
                                    self.task_tree.config(cursor="hand2")
                                    
                                    # Apply hover effect to this link
                                    tags = list(self.task_tree.item(item, "tags"))
                                    if "link" in tags:
                                        tags.remove("link")
                                    if "link_hover" not in tags:
                                        tags.append("link_hover")
                                    self.task_tree.item(item, tags=tags)
                                    return
                                break
            
            # Reset cursor if not over a link
            self.task_tree.config(cursor="")
    
    def copy_to_clipboard(self, text):
        """Copy text to clipboard."""
        self.root.clipboard_clear()
        self.root.clipboard_append(text)
        self.set_status(f"Copied to clipboard: {text}")
    
    def sort_treeview(self, column, reset=True):
        """
        Sort treeview by a column.
        
        Args:
            column: Column to sort by
            reset: Whether to reset the sort direction if clicking the same column
        """
        # Update sort indicators in column headings
        for col in self.task_tree["columns"]:
            # Remove any existing sort indicators
            heading_text = self.task_tree.heading(col)["text"]
            heading_text = heading_text.replace(" ▲", "").replace(" ▼", "")
            self.task_tree.heading(col, text=heading_text)
        
        # Determine sort direction
        if self.sort_column == column and reset:
            # Toggle sort direction if clicking the same column
            self.sort_reverse = not self.sort_reverse
        else:
            # Default to ascending for a new column
            self.sort_column = column
            self.sort_reverse = False
        
        # Add sort indicator to column heading
        heading_text = self.task_tree.heading(column)["text"]
        indicator = " ▲" if not self.sort_reverse else " ▼"
        self.task_tree.heading(column, text=heading_text + indicator)
        
        # Refresh the task list to apply the sorting
        self.refresh_task_list()
    
    def apply_sorting(self, tasks):
        """
        Apply sorting to the task list.
        
        Args:
            tasks: List of tasks to sort
        """
        # Define custom key functions for special columns
        def priority_key(task):
            # Sort by priority level (high, medium, low)
            priority_order = {"high": 0, "medium": 1, "low": 2}
            return priority_order.get(task.get("priority", "").lower(), 3)
        
        def status_key(task):
            # Sort by completion status and due date
            if task.get("completed", False):
                return (2, "")  # Completed tasks at the bottom
            
            if not task.get("due_date"):
                return (1, "")  # Tasks without due date in the middle
            
            # Tasks with due date sorted by days until due
            try:
                due_date = datetime.datetime.strptime(task["due_date"], "%Y-%m-%d").date()
                today = datetime.date.today()
                days_until_due = (due_date - today).days
                
                # Return a tuple for proper sorting
                if days_until_due < 0:
                    return (0, f"{abs(days_until_due):04d}")  # Overdue tasks first
                else:
                    return (0, f"{days_until_due+10000:04d}")  # Then by days until due
            except ValueError:
                return (1, "")  # Invalid date format in the middle
        
        # Sort the tasks based on the selected column
        if self.sort_column == "status":
            tasks.sort(key=status_key, reverse=self.sort_reverse)
        elif self.sort_column == "priority":
            tasks.sort(key=priority_key, reverse=self.sort_reverse)
        elif self.sort_column == "rev":
            # Special handling for revision numbers (numeric sorting when possible)
            def rev_key(task):
                rev = task.get("rev", "")
                # Try to extract numeric part for numeric sorting
                try:
                    # If it's a pure number, convert to int for proper sorting
                    if rev.isdigit():
                        return int(rev)
                    # If it has a format like "Rev 1" or "R2", extract the number
                    import re
                    match = re.search(r'(\d+)', rev)
                    if match:
                        return int(match.group(1))
                except (ValueError, AttributeError, TypeError) as e:
                    logging.debug(f"Revision parsing error: {e}")
                # Fall back to string sorting
                return rev.lower()
            
            tasks.sort(key=rev_key, reverse=self.sort_reverse)
        else:
            # For other columns, sort by the column value
            # Use a case-insensitive sort for text columns
            tasks.sort(
                key=lambda task: str(task.get(self.sort_column, "")).lower(),
                reverse=self.sort_reverse
            )

    def assign_task_to(self, staff_name):
        """Assign the selected task(s) to a staff member."""
        task_ids = self.get_selected_task_ids()
        if not task_ids:
            return
            
        count = len(task_ids)
        # Update each selected task
        for task_id in task_ids:
            self.task_manager.update_task(task_id, assigned_to=staff_name)
        
        # Refresh the task list and show confirmation
        self.refresh_task_list()
        self.set_status(f"{count} task{'s' if count > 1 else ''} assigned to {staff_name}")
        
        # Disable the Undo button after any other action
        if hasattr(self, 'undo_btn'):
            self.undo_btn.config(state=tk.DISABLED)

    def undo_delete(self):
        """Undo the most recent deletion (which might be multiple tasks)"""
        if not self.task_manager.deleted_tasks_stack:
            self.set_status("Nothing to undo")
            return
        
        # Get the most recent batch of deleted tasks
        recent_deletion = self.task_manager.deleted_tasks_stack.pop()
        
        # Recover all tasks in this batch
        recovered_ids = self.task_manager.recover_batch(recent_deletion)
        
        # Update the UI
        if recovered_ids:
            self.refresh_task_list()
            if len(recovered_ids) == 1:
                self.set_status(f"Restored task {recovered_ids[0]}")
            else:
                self.set_status(f"Restored {len(recovered_ids)} tasks")
        else:
            self.set_status("Failed to restore tasks")
        
        # Disable the Undo button if no more items in the stack
        if hasattr(self, 'undo_btn'):
            if not self.task_manager.deleted_tasks_stack:
                self.undo_btn.config(state=tk.DISABLED)
            else:
                self.undo_btn.config(state=tk.NORMAL)
    
    def position_window_center(self):
        """Positions the application window in the center of the screen."""
        self.root.update_idletasks()
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        window_width = self.root.winfo_width()
        window_height = self.root.winfo_height()
        x = int((screen_width / 2) - (window_width / 2))
        y = int((screen_height / 2) - (window_height / 2))
        self.root.geometry(f"+{x}+{y}")

    def bind_mousewheel_to_treeview(self):
        """Binds mousewheel scrolling to the task treeview for different platforms."""
        if platform.system() == 'Windows':
            self.task_tree.bind("<MouseWheel>", self.on_mousewheel_windows)
        elif platform.system() == 'Darwin':  # macOS
            self.task_tree.bind("<MouseWheel>", self.on_mousewheel_macos)
        elif platform.system() == 'Linux':
            self.task_tree.bind("<Button-4>", self.on_mousewheel_linux)
            self.task_tree.bind("<Button-5>", self.on_mousewheel_linux)

    def on_mousewheel_windows(self, event):
        """Handles mousewheel scrolling on Windows."""
        self.task_tree.yview_scroll(int(-1*(event.delta/120)), "units")

    def on_mousewheel_macos(self, event):
        """Handles mousewheel scrolling on macOS."""
        self.task_tree.yview_scroll(event.delta, "units")

    def on_mousewheel_linux(self, event):
        """Handles mousewheel scrolling on Linux."""
        if event.num == 4:
            self.task_tree.yview_scroll(-1, "units")
        elif event.num == 5:
            self.task_tree.yview_scroll(1, "units")

    def run(self): # ADD THIS run METHOD
        """Starts the Tkinter main event loop."""
        self.root.mainloop()

    def view_deleted_tasks(self):
        """Show a dialog with all deleted tasks for recovery or permanent deletion."""
        # Get all deleted tasks
        deleted_tasks = self.task_manager.get_filtered_tasks(show_completed=True, show_deleted=True)
        deleted_tasks = [t for t in deleted_tasks if t.get("deleted", False)]
        
        if not deleted_tasks:
            messagebox.showinfo("Deleted Tasks", "No deleted tasks found")
            return
            
        # Create a new toplevel window
        deleted_window = tk.Toplevel(self.root)
        deleted_window.title("Deleted Tasks")
        deleted_window.geometry("800x500")
        deleted_window.transient(self.root)
        deleted_window.grab_set()
        
        # Set icon for deleted tasks window
        try:
            icon_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'deleted.png')
            if os.path.exists(icon_path):
                icon_image = tk.PhotoImage(file=icon_path)
                deleted_window.iconphoto(False, icon_image)  # False = only apply to this window
                # Keep reference to prevent garbage collection
                deleted_window.deleted_icon = icon_image
        except Exception as e:
            logging.warning(f"Could not set deleted tasks window icon: {e}")
        
        # Create main frame
        main_frame = ttk.Frame(deleted_window, padding=10)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Add info label
        ttk.Label(main_frame, 
                  text="These tasks have been deleted but can be recovered. Select tasks and use the buttons below.",
                  wraplength=780).pack(pady=(0, 10))
        
        # Create treeview for deleted tasks
        columns = ("id", "title", "due_date", "priority", "category", "assigned_to")
        tree = ttk.Treeview(main_frame, columns=columns, show='headings')
        
        # Set column headings
        tree.heading('id', text='ID')
        tree.heading('title', text='Title')
        tree.heading('due_date', text='Due Date')
        tree.heading('priority', text='Priority')
        tree.heading('category', text='Category')
        tree.heading('assigned_to', text='Assigned To')
        
        # Set column widths
        tree.column('id', width=50, anchor=tk.CENTER)
        tree.column('title', width=250)
        tree.column('due_date', width=100, anchor=tk.CENTER)
        tree.column('priority', width=80, anchor=tk.CENTER)
        tree.column('category', width=100)
        tree.column('assigned_to', width=150)
        
        # Add scrollbar
        scrollbar = ttk.Scrollbar(main_frame, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        
        # Pack tree and scrollbar
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Add button frame
        button_frame = ttk.Frame(deleted_window, padding=10)
        button_frame.pack(fill=tk.X)
        
        # Populate the tree
        for task in deleted_tasks:
            tree.insert('', tk.END, values=(
                task['id'],
                task['title'],
                task.get('due_date', ''),
                task.get('priority', ''),
                task.get('category', ''),
                task.get('assigned_to', '')
            ))
        
        # Function to recover selected tasks
        def recover_selected():
            selected_items = tree.selection()
            if not selected_items:
                messagebox.showinfo("Recover Task", "Please select at least one task to recover")
                return
                
            recovered_count = 0
            for item in selected_items:
                task_id = int(tree.item(item, 'values')[0])
                if self.task_manager.recover_task(task_id):
                    recovered_count += 1
                    
            if recovered_count > 0:
                self.refresh_task_list()
                self.set_status(f"Recovered {recovered_count} task(s)")
                # Refresh the deleted tasks window
                for item in tree.get_children():
                    tree.delete(item)
                    
                # Re-populate with remaining deleted tasks
                remaining_deleted = [t for t in self.task_manager.get_filtered_tasks(
                    show_completed=True, show_deleted=True) if t.get("deleted", False)]
                    
                for task in remaining_deleted:
                    tree.insert('', tk.END, values=(
                        task['id'],
                        task['title'],
                        task.get('due_date', ''),
                        task.get('priority', ''),
                        task.get('category', ''),
                        task.get('assigned_to', '')
                    ))
                    
                if not remaining_deleted:
                    deleted_window.destroy()
        
        # Function to permanently delete selected tasks
        def permanently_delete_selected():
            selected_items = tree.selection()
            if not selected_items:
                messagebox.showinfo("Delete Task", "Please select at least one task to permanently delete")
                return
                
            task_count = len(selected_items)
            confirm = messagebox.askyesno(
                "Confirm Permanent Deletion",
                f"Are you sure you want to permanently delete {task_count} task(s)?\n\n"
                "This action cannot be undone.",
                icon=messagebox.WARNING
            )
            
            if not confirm:
                return
                
            deleted_count = 0
            for item in selected_items:
                task_id = int(tree.item(item, 'values')[0])
                if self.task_manager.permanently_delete_task(task_id):
                    deleted_count += 1
                    
            if deleted_count > 0:
                self.set_status(f"Permanently deleted {deleted_count} task(s)")
                # Refresh the deleted tasks window
                for item in tree.get_children():
                    tree.delete(item)
                    
                # Re-populate with remaining deleted tasks
                remaining_deleted = [t for t in self.task_manager.get_filtered_tasks(
                    show_completed=True, show_deleted=True) if t.get("deleted", False)]
                    
                for task in remaining_deleted:
                    tree.insert('', tk.END, values=(
                        task['id'],
                        task['title'],
                        task.get('due_date', ''),
                        task.get('priority', ''),
                        task.get('category', ''),
                        task.get('assigned_to', '')
                    ))
                    
                if not remaining_deleted:
                    deleted_window.destroy()
        
        # Function to recover all tasks
        def recover_all():
            confirm = messagebox.askyesno(
                "Confirm Recovery",
                "Are you sure you want to recover all deleted tasks?",
                icon=messagebox.QUESTION
            )
            
            if not confirm:
                return
                
            recovered = self.task_manager.recover_all_deleted_tasks()
            if recovered > 0:
                self.refresh_task_list()
                self.set_status(f"Recovered all {recovered} deleted task(s)")
                deleted_window.destroy()
        
        # Function to permanently delete all tasks
        def permanently_delete_all():
            confirm = messagebox.askyesno(
                "Confirm Permanent Deletion",
                "Are you sure you want to PERMANENTLY DELETE ALL deleted tasks?\n\n"
                "This action cannot be undone.",
                icon=messagebox.WARNING
            )
            
            if not confirm:
                return
                
            deleted = self.task_manager.permanently_delete_all_deleted_tasks()
            if deleted > 0:
                self.set_status(f"Permanently deleted all {deleted} task(s)")
                deleted_window.destroy()
        
        # Add buttons
        ttk.Button(button_frame, text="Recover Selected", command=recover_selected).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Permanently Delete Selected", command=permanently_delete_selected).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Recover All", command=recover_all).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Permanently Delete All", command=permanently_delete_all).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Close", command=deleted_window.destroy).pack(side=tk.RIGHT, padx=5)


class UserManagementApp:
    """User Management interface for managing users."""
    
    def __init__(self, root, main_menu, task_manager):
        """Initialize the User Management interface."""
        self.root = root
        self.main_menu = main_menu
        self.task_manager = task_manager
        self.main_frame = None
        
        # Load user mapping from database if available
        if self.task_manager:
            db_users = self.task_manager.load_users_from_db()
            if db_users:
                self.user_mapping = db_users
            else:
                # Fallback to global USER_MAPPING
                self.user_mapping = USER_MAPPING.copy()
        else:
            self.user_mapping = USER_MAPPING.copy()
    
    def show(self):
        """Show the user management interface."""
        # Create main frame with white background
        self.main_frame = tk.Frame(self.root, bg='white')
        self.main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Header section with color
        header_frame = tk.Frame(self.main_frame, bg='#2c3e50', height=80)
        header_frame.pack(fill=tk.X)
        header_frame.pack_propagate(False)
        
        # Title and back button container
        header_content = tk.Frame(header_frame, bg='#2c3e50')
        header_content.pack(fill=tk.BOTH, expand=True, padx=15)
        
        # Title label in header
        title_label = tk.Label(
            header_content,
            text="User Management",
            font=('Segoe UI', 20, 'bold'),
            bg='#2c3e50',
            fg='white'
        )
        title_label.pack(side=tk.LEFT, expand=True)
        
        # Back button in header
        back_btn = tk.Button(
            header_content,
            text="← Back",
            command=self.back_to_menu,
            bg='#34495e',
            fg='white',
            font=('Segoe UI', 9, 'bold'),
            cursor='hand2',
            relief=tk.FLAT,
            bd=0,
            padx=15,
            pady=8,
            activebackground='#1c2833',
            activeforeground='white'
        )
        back_btn.pack(side=tk.RIGHT, pady=20)
        
        # Content frame
        content_frame = tk.Frame(self.main_frame, bg='white')
        content_frame.pack(fill=tk.BOTH, expand=True, padx=25, pady=20)
        
        # Action buttons frame
        button_frame = tk.Frame(content_frame, bg='white')
        button_frame.pack(fill=tk.X, pady=(0, 15))
        
        # Modern button style - more compact
        btn_config = {
            'font': ('Segoe UI', 9),
            'cursor': 'hand2',
            'relief': tk.FLAT,
            'bd': 0,
            'padx': 12,
            'pady': 8
        }
        
        # Add User button
        add_btn = tk.Button(
            button_frame,
            text="➕ Add",
            command=self.add_user,
            bg='#27ae60',
            fg='white',
            activebackground='#229954',
            activeforeground='white',
            **btn_config
        )
        add_btn.pack(side=tk.LEFT, padx=(0, 5))
        
        # Edit User button
        edit_btn = tk.Button(
            button_frame,
            text="✏️ Edit",
            command=self.edit_user,
            bg='#3498db',
            fg='white',
            activebackground='#2980b9',
            activeforeground='white',
            **btn_config
        )
        edit_btn.pack(side=tk.LEFT, padx=5)
        
        # Delete User button
        delete_btn = tk.Button(
            button_frame,
            text="🗑️ Delete",
            command=self.delete_user,
            bg='#e74c3c',
            fg='white',
            activebackground='#c0392b',
            activeforeground='white',
            **btn_config
        )
        delete_btn.pack(side=tk.LEFT, padx=5)
        
        # Refresh button
        refresh_btn = tk.Button(
            button_frame,
            text="🔄 Refresh",
            command=self.refresh_user_list,
            bg='#95a5a6',
            fg='white',
            activebackground='#7f8c8d',
            activeforeground='white',
            **btn_config
        )
        refresh_btn.pack(side=tk.LEFT, padx=5)
        
        # User list frame with border
        list_container = tk.Frame(content_frame, bg='#ecf0f1', relief=tk.SOLID, bd=1)
        list_container.pack(fill=tk.BOTH, expand=True)
        
        # Create treeview for users
        columns = ("display_name", "tcn", "status")
        self.user_tree = ttk.Treeview(list_container, columns=columns, show="headings", height=12)
        
        # Configure treeview style for better alignment
        style = ttk.Style()
        style.configure("Treeview", font=('Segoe UI', 9), rowheight=30)
        style.configure("Treeview.Heading", font=('Segoe UI', 10, 'bold'))
        
        # Define headings
        self.user_tree.heading("display_name", text="Display Name", anchor=tk.CENTER)
        self.user_tree.heading("tcn", text="TCN Number", anchor=tk.CENTER)
        self.user_tree.heading("status", text="Status", anchor=tk.CENTER)
        
        # Define column widths with center alignment for all
        self.user_tree.column("display_name", width=150, anchor=tk.CENTER)
        self.user_tree.column("tcn", width=120, anchor=tk.CENTER)
        self.user_tree.column("status", width=100, anchor=tk.CENTER)
        
        # Add scrollbar
        scrollbar = ttk.Scrollbar(list_container, orient=tk.VERTICAL, command=self.user_tree.yview)
        self.user_tree.configure(yscrollcommand=scrollbar.set)
        
        # Pack treeview and scrollbar
        self.user_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Bind double-click to edit
        self.user_tree.bind("<Double-Button-1>", lambda e: self.edit_user())
        
        # Refresh user list
        self.refresh_user_list()
    
    def refresh_user_list(self):
        """Refresh the user list display with real-time status."""
        # Clear existing items
        for item in self.user_tree.get_children():
            self.user_tree.delete(item)
        
        # Get active sessions from database if available
        active_users = set()
        if self.task_manager and self.task_manager.cursor:
            try:
                self.task_manager.cursor.execute("""
                    SELECT DISTINCT username
                    FROM UserSessions
                    WHERE is_active = 1
                    AND last_heartbeat > DATEADD(MINUTE, -5, GETDATE())
                """)
                active_users = {row[0] for row in self.task_manager.cursor.fetchall()}
            except Exception as e:
                logging.warning(f"Could not fetch active sessions: {e}")
        
        # Add users to tree
        for display_name in sorted(self.user_mapping.keys()):
            user_data = self.user_mapping[display_name]
            tcn = user_data.get("tcn", "N/A")
            
            # Check if user is online (tcn is the Windows username)
            status = "Online" if tcn in active_users else "Offline"
            
            self.user_tree.insert("", tk.END, values=(display_name, tcn, status), iid=display_name)
            
            # Color code the status
            item_id = display_name
            if status == "Online":
                self.user_tree.item(item_id, tags=('online',))
            else:
                self.user_tree.item(item_id, tags=('offline',))
        
        # Configure tag colors
        self.user_tree.tag_configure('online', foreground='#27ae60', font=('Segoe UI', 9, 'bold'))
        self.user_tree.tag_configure('offline', foreground='#95a5a6')
    
    def add_user(self):
        """Add a new user."""
        dialog = UserDialog(self.root, "Add User")
        if dialog.result:
            display_name = dialog.result["display_name"].strip()
            tcn = dialog.result["tcn"].strip()
            
            if display_name and tcn:
                if display_name in self.user_mapping:
                    messagebox.showerror("Error", f"User '{display_name}' already exists.", parent=self.root)
                elif display_name == "All":
                    messagebox.showerror("Error", "'All' is a reserved name.", parent=self.root)
                else:
                    # Auto-generate SQL username (TaskUser + next number)
                    existing_numbers = []
                    for user_data in self.user_mapping.values():
                        sql_user = user_data.get("sql_user", "")
                        if sql_user.startswith("TaskUser"):
                            try:
                                num = int(sql_user.replace("TaskUser", ""))
                                existing_numbers.append(num)
                            except ValueError:
                                pass
                    
                    next_number = max(existing_numbers) + 1 if existing_numbers else 1
                    sql_username = f"TaskUser{next_number}"
                    
                    # Save to database
                    if self.task_manager:
                        success = self.task_manager.save_user_to_db(display_name, tcn, sql_username)
                        if not success:
                            messagebox.showerror("Database Error", 
                                f"Failed to save user '{display_name}' to the database.\n\n"
                                f"Please check the logs for details.",
                                parent=self.root)
                            return
                    else:
                        messagebox.showwarning("No Database Connection",
                            f"Cannot save user without database connection.",
                            parent=self.root)
                        return
                    
                    # Add to local mapping
                    self.user_mapping[display_name] = {
                        "tcn": tcn,
                        "sql_user": sql_username
                    }
                    self.update_global_users()
                    self.refresh_user_list()
                    
                    messagebox.showinfo("Success", 
                        f"User '{display_name}' added successfully!\n\n"
                        f"Display Name: {display_name}\n"
                        f"TCN: {tcn}\n"
                        f"SQL Username: {sql_username}\n\n"
                        f"✓ User saved to database\n\n"
                        f"Note: To use this account, a DBA must create\n"
                        f"the SQL Server login '{sql_username}' with password 'pass1'.",
                        parent=self.root)
    
    def edit_user(self):
        """Edit selected user."""
        selection = self.user_tree.selection()
        if not selection:
            messagebox.showwarning("No Selection", "Please select a user to edit.", parent=self.root)
            return
        
        # Get selected user
        old_display_name = selection[0]  # Using iid which is the display name
        old_data = self.user_mapping.get(old_display_name)
        
        if not old_data:
            messagebox.showerror("Error", "User data not found.", parent=self.root)
            return
        
        # Show edit dialog
        dialog = UserDialog(self.root, "Edit User", old_display_name, old_data.get("tcn", ""))
        if dialog.result:
            new_display_name = dialog.result["display_name"].strip()
            new_tcn = dialog.result["tcn"].strip()
            
            if new_display_name and new_tcn:
                if new_display_name != old_display_name and new_display_name in self.user_mapping:
                    messagebox.showerror("Error", f"User '{new_display_name}' already exists.", parent=self.root)
                elif new_display_name == "All":
                    messagebox.showerror("Error", "'All' is a reserved name.", parent=self.root)
                else:
                    # Update in database
                    if self.task_manager:
                        success = self.task_manager.update_user_in_db(old_display_name, new_display_name, new_tcn)
                        if not success:
                            messagebox.showerror("Database Error", "Failed to update user in database.", parent=self.root)
                            return
                    
                    # Update local mapping
                    sql_user = old_data.get("sql_user", "")  # Keep same SQL username
                    
                    if new_display_name != old_display_name:
                        # Remove old entry and add new one
                        del self.user_mapping[old_display_name]
                    
                    self.user_mapping[new_display_name] = {
                        "tcn": new_tcn,
                        "sql_user": sql_user
                    }
                    
                    self.update_global_users()
                    self.refresh_user_list()
                    messagebox.showinfo("Success", f"User updated successfully.", parent=self.root)
    
    def delete_user(self):
        """Delete selected user."""
        selection = self.user_tree.selection()
        if not selection:
            messagebox.showwarning("No Selection", "Please select a user to delete.", parent=self.root)
            return
        
        # Get selected user (using iid which is display_name)
        display_name = selection[0]
        
        # Confirm deletion
        if messagebox.askyesno("Confirm Delete", f"Are you sure you want to delete user '{display_name}'?\n\nNote: This will not remove their assignments from existing tasks.", parent=self.root):
            # Delete from database
            if self.task_manager:
                success = self.task_manager.delete_user_from_db(display_name)
                if not success:
                    messagebox.showerror("Database Error", "Failed to delete user from database.", parent=self.root)
                    return
            
            # Remove from local mapping
            del self.user_mapping[display_name]
            self.update_global_users()
            self.refresh_user_list()
            messagebox.showinfo("Success", f"User '{display_name}' deleted successfully.", parent=self.root)
    
    def update_global_users(self):
        """Update the global USERS and USER_MAPPING."""
        global USERS, USER_MAPPING
        USER_MAPPING = self.user_mapping
        USERS = ["All"] + sorted(USER_MAPPING.keys())
    
    def back_to_menu(self):
        """Return to main menu."""
        self.main_menu.show_main_menu()


class UserDialog:
    """Dialog for adding or editing a user."""
    
    def __init__(self, parent, title, display_name=None, tcn=None):
        """Initialize the dialog."""
        self.result = None
        self.parent = parent
        
        # Create dialog
        self.dialog = tk.Toplevel(parent)
        self.dialog.withdraw()
        self.dialog.title(title)
        self.dialog.transient(parent)
        self.dialog.resizable(False, False)
        
        # Create form
        main_frame = ttk.Frame(self.dialog, padding=20)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Display Name field
        ttk.Label(main_frame, text="Display Name:", font=('TkDefaultFont', 10)).pack(anchor=tk.W, pady=(0, 5))
        
        self.display_name_var = tk.StringVar(value=display_name if display_name else "")
        display_name_entry = ttk.Entry(main_frame, textvariable=self.display_name_var, width=30, font=('TkDefaultFont', 10))
        display_name_entry.pack(fill=tk.X, pady=(0, 15))
        display_name_entry.focus()
        
        # TCN Number field
        ttk.Label(main_frame, text="TCN Number (e.g., a0011071):", font=('TkDefaultFont', 10)).pack(anchor=tk.W, pady=(0, 5))
        
        self.tcn_var = tk.StringVar(value=tcn if tcn else "")
        tcn_entry = ttk.Entry(main_frame, textvariable=self.tcn_var, width=30, font=('TkDefaultFont', 10))
        tcn_entry.pack(fill=tk.X, pady=(0, 15))
        
        # Buttons
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X)
        
        ttk.Button(button_frame, text="Cancel", command=self.cancel).pack(side=tk.RIGHT, padx=(5, 0))
        ttk.Button(button_frame, text="Save", command=self.save).pack(side=tk.RIGHT)
        
        # Bind Enter key to save
        self.dialog.bind("<Return>", lambda e: self.save())
        self.dialog.bind("<Escape>", lambda e: self.cancel())
        
        # Position dialog
        self.dialog.update_idletasks()
        x = parent.winfo_rootx() + (parent.winfo_width() - self.dialog.winfo_width()) // 2
        y = parent.winfo_rooty() + (parent.winfo_height() - self.dialog.winfo_height()) // 2
        self.dialog.geometry(f"+{x}+{y}")
        
        # Show dialog
        self.dialog.deiconify()
        self.dialog.grab_set()
        self.dialog.wait_window()
    
    def save(self):
        """Save the user data."""
        display_name = self.display_name_var.get().strip()
        tcn = self.tcn_var.get().strip()
        
        if not display_name:
            messagebox.showerror("Validation Error", "Display Name is required.", parent=self.dialog)
            return
        
        if not tcn:
            messagebox.showerror("Validation Error", "TCN Number is required.", parent=self.dialog)
            return
        
        if not tcn.startswith("a") or len(tcn) != 8:
            messagebox.showerror("Validation Error", "TCN Number must start with 'a' and be 8 characters long (e.g., a0011071).", parent=self.dialog)
            return
        
        self.result = {
            "display_name": display_name,
            "tcn": tcn
        }
        self.dialog.destroy()
    
    def cancel(self):
        """Cancel the dialog."""
        self.result = None
        self.dialog.destroy()


class TaskDialog:
    """Dialog for adding or editing a task."""
    
    def __init__(self, parent, title, task=None):
        """Initialize the dialog."""
        self.result = None
        self.parent = parent
        
        # Create dialog window but keep it hidden until fully built
        self.dialog = tk.Toplevel(parent)
        self.dialog.withdraw()  # Hide the window during setup
        self.dialog.title(title)
        
        # Set icon for Add/Edit task dialog
        if "Add" in title:
            try:
                icon_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'Add.png')
                if os.path.exists(icon_path):
                    icon_image = tk.PhotoImage(file=icon_path)
                    self.dialog.iconphoto(False, icon_image)
                    self.dialog.add_icon = icon_image  # Keep reference
            except Exception as e:
                logging.warning(f"Could not set add task dialog icon: {e}")
        
        # Set fixed dialog size - ADJUSTED HEIGHT FOR SCRAPE FIELD
        self.dialog_width = 630
        self.dialog_height = 670 # Increased height
        self.dialog.geometry(f"{self.dialog_width}x{self.dialog_height}")
        self.dialog.minsize(self.dialog_width, self.dialog_height)
        self.dialog.maxsize(self.dialog_width, self.dialog_height)  # Prevent resizing
        self.dialog.transient(parent)
        
        # Create main container with fixed padding
        main_container = ttk.Frame(self.dialog)
        main_container.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Create canvas with scrollbar - use exact width calculation
        canvas_width = self.dialog_width - 40  # Account for dialog padding (20px on each side)
        self.canvas = tk.Canvas(main_container, width=canvas_width, highlightthickness=0)
        scrollbar = ttk.Scrollbar(main_container, orient="vertical", command=self.canvas.yview) # Define scrollbar here
        
        # --- ADD FIELD MAPPING ---
        # Copy the field mapping from smartdb_login.py for easy access
        self.FIELDS_TO_EXTRACT = {
            "Request No": "Application_number",
            "Requested Date": "id_shinseibi",
            "Deadline": "id_yoteibi",
            "Revision No": "REV_No",
            "Applied Vessel": "Ship_No",
            "Drawing No": "DWG_No",
            "Description": "Drawing_Name_Common_1", # Often used for 'Equipment Name'
            "Note": "id_bikouran",
            "Modeling Staff": "acModeler", # Could map to Main Staff or Assigned To
            # We also need the file path if downloaded
            "Downloaded File Path": "Downloaded File Path" # Key added by login_to_smartdb
        }
        # Define which dialog field variable corresponds to which scraped field name
        # --- MOVED THIS BLOCK AFTER create_form CALL ---
        # self.FIELD_VAR_MAP = {
        #     "Request No": self.request_no_var,
        #     "Equipment Name": self.title_var, # Mapping 'Description' from scrape to 'Equipment Name' dialog field
        #     "Applied Vessel": self.applied_vessel_var,
        #     "Revision No": self.rev_var,
        #     "Drawing No": self.drawing_no_var,
        #     # Dates need special handling (parsing and setting DateEntry)
        #     "Requested Date": self.requested_date_picker,
        #     "Due Date": self.due_date_picker, # Mapping 'Deadline' from scrape
        #     # Links
        #     "Link": self.link_var, # Mapping 'Downloaded File Path' to 'Link' dialog field
        #     # Text Area
        #     "Description": self.description_text, # Mapping 'Note' from scrape to 'Description' dialog field
        #     # Comboboxes
        #     "Main Staff": self.main_staff_var, # Mapping 'Modeling Staff' from scrape
        #     "Assigned To": self.assigned_to_var # Can also map 'Modeling Staff' here if desired
        # }
        # --- END FIELD MAPPING ---
        
        # Create a frame inside the canvas for the form - fill entire canvas width
        self.form_width = canvas_width  # Fill to scrollbar
        self.form_frame = ttk.Frame(self.canvas, width=self.form_width)
        
        # Pack scrollbar and canvas AFTER they are both defined
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        # Configure canvas - use exact coordinates to prevent shifting
        self.canvas.configure(yscrollcommand=scrollbar.set)
        self.canvas_frame = self.canvas.create_window((0, 0), window=self.form_frame, anchor=tk.NW, width=self.form_width)
        
        # Create form elements
        self.create_form(self.form_frame, task)
        
        # --- DEFINE FIELD MAPPING AFTER WIDGETS ARE CREATED ---
        self.FIELD_VAR_MAP = {
            "Request No": self.request_no_var,
            "Equipment Name": self.title_var, # Mapping 'Description' from scrape to 'Equipment Name' dialog field
            "Applied Vessel": self.applied_vessel_var,
            "Revision No": self.rev_var,
            "Drawing No": self.drawing_no_var,
            # Dates need special handling (parsing and setting DateEntry)
            "Requested Date": self.requested_date_picker,
            "Due Date": self.due_date_picker, # Mapping 'Deadline' from scrape
            # Links
            "Link": self.link_var, # Mapping 'Downloaded File Path' to 'Link' dialog field
            # Text Area
            "Description": self.description_text, # Mapping 'Note' from scrape to 'Description' dialog field
            # Comboboxes
            "Main Staff": self.main_staff_var, # Mapping 'Modeling Staff' from scrape
            "Assigned To": self.assigned_to_var # Can also map 'Modeling Staff' here if desired
        }
        # --- END FIELD MAPPING DEFINITION ---
        
        # Create button frame at the bottom (outside the scrollable area)
        button_frame = ttk.Frame(self.dialog)
        button_frame.pack(fill=tk.X, padx=10, pady=10)
        
        ttk.Button(button_frame, text="Cancel", command=self.cancel).pack(side=tk.RIGHT, padx=5)
        ttk.Button(button_frame, text="Save", command=self.save).pack(side=tk.RIGHT, padx=5)
        
        # Bind canvas configuration and mouse wheel
        self.form_frame.bind('<Configure>', self.on_frame_configure)
        self.canvas.bind('<Configure>', self.on_canvas_configure)
        
        # Use a more reliable way to bind the mousewheel
        self.bind_mousewheel()
        
        # Bind dialog close event to ensure proper cleanup
        self.dialog.protocol("WM_DELETE_WINDOW", self.cancel)
        
        # Position the dialog
        self.position_dialog()
        
        # Pre-configure the canvas scroll region and size before showing the dialog
        self.dialog.update_idletasks()
        self.canvas.update_idletasks()
        self.on_frame_configure()
        # Force canvas to set correct width immediately
        self.canvas.itemconfig(self.canvas_frame, width=self.form_width)
        self.canvas.update_idletasks()
        
        # Now show the dialog
        self.dialog.deiconify()
        
        # Grab focus and wait
        self.dialog.grab_set()
        self.dialog.focus_force()
        self.dialog.wait_window()
    
    def bind_mousewheel(self):
        """Bind mousewheel events to the canvas."""
        # Windows and macOS use different events and deltas
        if sys.platform.startswith('win'):
            self.canvas.bind_all("<MouseWheel>", self.on_mousewheel_windows)
        elif sys.platform.startswith('darwin'):
            self.canvas.bind_all("<MouseWheel>", self.on_mousewheel_macos)
        else:
            # Linux
            self.canvas.bind_all("<Button-4>", self.on_mousewheel_linux)
            self.canvas.bind_all("<Button-5>", self.on_mousewheel_linux)
    
    def unbind_mousewheel(self):
        """Unbind mousewheel events when dialog is closed."""
        try:
            # Unbind all mousewheel events
            self.canvas.unbind_all("<MouseWheel>")
            self.canvas.unbind_all("<Button-4>")
            self.canvas.unbind_all("<Button-5>")
        except (tk.TclError, AttributeError) as e:
            logging.debug(f"Mousewheel unbind error (expected if canvas destroyed): {e}")
    
    def on_mousewheel_windows(self, event):
        """Handle mousewheel scrolling on Windows."""
        try:
            self.canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        except (tk.TclError, AttributeError) as e:
            logging.debug(f"Mousewheel scroll error (expected if canvas destroyed): {e}")
    
    def on_mousewheel_macos(self, event):
        """Handle mousewheel scrolling on macOS."""
        try:
            self.canvas.yview_scroll(int(-1 * event.delta), "units")
        except (tk.TclError, AttributeError) as e:
            logging.debug(f"Mousewheel scroll error (expected if canvas destroyed): {e}")
    
    def on_mousewheel_linux(self, event):
        """Handle mousewheel scrolling on Linux."""
        try:
            if event.num == 4:
                self.canvas.yview_scroll(-1, "units")
            elif event.num == 5:
                self.canvas.yview_scroll(1, "units")
        except (tk.TclError, AttributeError) as e:
            logging.debug(f"Mousewheel scroll error (expected if canvas destroyed): {e}")
    
    def on_frame_configure(self, event=None):
        """Reset the scroll region to encompass the inner frame"""
        try:
            # Set the scroll region to the entire canvas
            self.canvas.configure(scrollregion=self.canvas.bbox("all"))
            
            # Ensure the form frame stays at the left edge
            self.canvas.coords(self.canvas_frame, 0, 0)
        except (tk.TclError, AttributeError) as e:
            logging.debug(f"Frame configure error (expected if canvas destroyed): {e}")
    
    def on_canvas_configure(self, event):
        """When the canvas is resized, maintain fixed form width"""
        try:
            # Keep the form at a fixed width - do NOT resize with canvas
            self.canvas.itemconfig(self.canvas_frame, width=self.form_width)
            
            # Ensure the form frame stays at the left edge
            self.canvas.coords(self.canvas_frame, 0, 0)
        except (tk.TclError, AttributeError) as e:
            logging.debug(f"Canvas configure error (expected if canvas destroyed): {e}")
    
    def cancel(self):
        """Cancel the dialog and clean up resources."""
        # Unbind mousewheel events before destroying the dialog
        self.unbind_mousewheel()
        self.dialog.destroy()
    
    def position_dialog(self):
        """Position the dialog on the same monitor as the parent window."""
        # Update the dialog's geometry
        self.dialog.update_idletasks()
        
        # Get parent window geometry
        parent_x = self.parent.winfo_rootx()
        parent_y = self.parent.winfo_rooty()
        parent_width = self.parent.winfo_width()
        parent_height = self.parent.winfo_height()
        
        # Calculate position
        x = parent_x + (parent_width - self.dialog_width) // 2
        y = parent_y + (parent_height - self.dialog_height) // 2
        
        # Set the position
        self.dialog.geometry(f"+{x}+{y}")
    
    def create_form(self, parent, task=None):
        """Create the form fields in a single frame."""
        # Configure grid
        parent.columnconfigure(0, weight=0, minsize=120)  # Label column - fixed width
        parent.columnconfigure(1, weight=1)  # Entry column - expandable
        parent.columnconfigure(2, weight=0, minsize=20)  # Button column - fixed width
        
        # Request No. and Equipment Name
        row = 0
        # --- ADD SCRAPE ROW ---
        ttk.Label(parent, text="Scrape URL:").grid(row=row, column=0, sticky=tk.W, pady=5, padx=(0,5))
        scrape_frame = ttk.Frame(parent)
        scrape_frame.grid(row=row, column=1, sticky=tk.EW, pady=5)
        scrape_frame.columnconfigure(0, weight=1) # Make entry expand

        self.scrape_url_var = tk.StringVar() # Variable for scrape URL
        scrape_entry = ttk.Entry(scrape_frame, textvariable=self.scrape_url_var)
        scrape_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))

        # Add Scrape button
        scrape_btn = ttk.Button(scrape_frame, text="Scrape", command=self.scrape_data, style="BrowseButton.TButton") # Reuse browse button style for size consistency
        scrape_btn.pack(side=tk.RIGHT, fill=tk.Y)
        
        # --- END SCRAPE ROW ---
        
        row += 1 # Increment row index for next element
        ttk.Label(parent, text="Request No.:").grid(row=row, column=0, sticky=tk.W, pady=5, padx=(0,5))
        self.request_no_var = tk.StringVar(value=task.get("request_no", "") if task else "")
        ttk.Entry(parent, textvariable=self.request_no_var).grid(row=row, column=1, sticky=tk.EW, pady=5)
        
        row += 1
        ttk.Label(parent, text="Equipment Name:*").grid(row=row, column=0, sticky=tk.W, pady=5, padx=(0,5))
        self.title_var = tk.StringVar(value=task["title"] if task else "")
        ttk.Entry(parent, textvariable=self.title_var).grid(row=row, column=1, sticky=tk.EW, pady=5)
        
        # Vessel Information
        row += 1
        ttk.Separator(parent, orient=tk.HORIZONTAL).grid(row=row, column=0, columnspan=2, sticky=tk.EW, pady=10)
        
        row += 1
        ttk.Label(parent, text="Applied Vessel:").grid(row=row, column=0, sticky=tk.W, pady=5, padx=(0,5))
        self.applied_vessel_var = tk.StringVar(value=task.get("applied_vessel", "") if task else "")
        ttk.Entry(parent, textvariable=self.applied_vessel_var).grid(row=row, column=1, sticky=tk.EW, pady=5)
        
        row += 1
        ttk.Label(parent, text="Rev:").grid(row=row, column=0, sticky=tk.W, pady=5, padx=(0,5))
        self.rev_var = tk.StringVar(value=task.get("rev", "") if task else "")
        ttk.Entry(parent, textvariable=self.rev_var).grid(row=row, column=1, sticky=tk.EW, pady=5)
        
        row += 1
        ttk.Label(parent, text="Drawing No.:").grid(row=row, column=0, sticky=tk.W, pady=5, padx=(0,5))
        self.drawing_no_var = tk.StringVar(value=task.get("drawing_no", "") if task else "")
        ttk.Entry(parent, textvariable=self.drawing_no_var).grid(row=row, column=1, sticky=tk.EW, pady=5)
        
        # Dates
        row += 1
        ttk.Separator(parent, orient=tk.HORIZONTAL).grid(row=row, column=0, columnspan=2, sticky=tk.EW, pady=10)
        
        row += 1
        # Create a frame for dates that spans both columns and uses grid
        date_frame = ttk.Frame(parent)
        date_frame.grid(row=row, column=0, columnspan=2, sticky=tk.EW, pady=5)
        date_frame.columnconfigure(1, weight=1)  # Make the date entries expand
        
        # Configure date frame columns for proper alignment
        date_frame.columnconfigure(0, minsize=120)  # Same as label column
        date_frame.columnconfigure(1, weight=1)     # First date picker
        date_frame.columnconfigure(2, minsize=80)   # Second label
        date_frame.columnconfigure(3, weight=1)     # Second date picker
        date_frame.columnconfigure(4, minsize=80)   # Third label
        date_frame.columnconfigure(3, weight=1)     # Second date picker
        
        # Row 1: Requested Date and Due Date
        ttk.Label(date_frame, text="Requested Date:").grid(row=0, column=0, sticky=tk.W)
        self.requested_date_picker = DateEntry(date_frame, width=12, date_pattern='yyyy-mm-dd')
        self.requested_date_picker.grid(row=0, column=1, sticky=tk.EW, padx=(0,10))
        
        ttk.Label(date_frame, text="Due Date:").grid(row=0, column=2, sticky=tk.W)
        self.due_date_picker = DateEntry(date_frame, width=12, date_pattern='yyyy-mm-dd')
        self.due_date_picker.grid(row=0, column=3, sticky=tk.EW)
        
        # Row 2: Target Start and Target Finish
        ttk.Label(date_frame, text="Target Start:").grid(row=1, column=0, sticky=tk.W, pady=(5,0))
        self.target_start_picker = DateEntry(date_frame, width=12, date_pattern='yyyy-mm-dd')
        self.target_start_picker.grid(row=1, column=1, sticky=tk.EW, padx=(0,10), pady=(5,0))
        
        ttk.Label(date_frame, text="Target Finish:").grid(row=1, column=2, sticky=tk.W, pady=(5,0))
        self.target_finish_picker = DateEntry(date_frame, width=12, date_pattern='yyyy-mm-dd')
        self.target_finish_picker.grid(row=1, column=3, sticky=tk.EW, padx=(0,10), pady=(5,0))
        
        # Set existing dates if available
        if task:
            for picker, date_key in [(self.requested_date_picker, "requested_date"),
                                   (self.due_date_picker, "due_date"),
                                   (self.target_start_picker, "target_start"),
                                   (self.target_finish_picker, "target_finish")]:
                if task.get(date_key):
                    try:
                        # Handle both string and datetime.date objects
                        date_val = task[date_key]
                        if isinstance(date_val, datetime.date):
                            date_value = date_val
                        else:
                            date_value = datetime.datetime.strptime(date_val, "%Y-%m-%d").date()
                        picker.set_date(date_value)
                    except (ValueError, TypeError):
                        pass
        
        # Links
        row += 1
        ttk.Separator(parent, orient=tk.HORIZONTAL).grid(row=row, column=0, columnspan=2, sticky=tk.EW, pady=10)
        
        row += 1
        ttk.Label(parent, text="Link:").grid(row=row, column=0, sticky=tk.W, pady=5)
        
        # Create a frame for the link field and browse button
        link_frame = ttk.Frame(parent)
        link_frame.grid(row=row, column=1, sticky=tk.EW, pady=5)
        link_frame.columnconfigure(0, weight=1)  # Make the entry expand
        
        self.link_var = tk.StringVar(value=task.get("link", "") if task else "")
        
        # Create a custom style for the browse button to match entry height
        # FIX: Use the style from the parent window instead of the dialog
        style = ttk.Style()
        style.configure("BrowseButton.TButton", padding=0)
        
        # Create the entry
        link_entry = ttk.Entry(link_frame, textvariable=self.link_var)
        link_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
        
        # Add browse button - using pack instead of grid for better height matching
        browse_btn = ttk.Button(link_frame, text="Browse", command=self.browse_folder, style="BrowseButton.TButton")
        browse_btn.pack(side=tk.RIGHT, fill=tk.Y)
        
        row += 1
        ttk.Label(parent, text="SDB Link:").grid(row=row, column=0, sticky=tk.W, pady=5)
        self.sdb_link_var = tk.StringVar(value=task.get("sdb_link", "") if task else "")
        sdb_link_entry = ttk.Entry(parent, textvariable=self.sdb_link_var)
        sdb_link_entry.grid(row=row, column=1, sticky=tk.EW, pady=5)
        
        # Description
        row += 1
        ttk.Separator(parent, orient=tk.HORIZONTAL).grid(row=row, column=0, columnspan=2, sticky=tk.EW, pady=10)
        
        row += 1
        ttk.Label(parent, text="Description:").grid(row=row, column=0, sticky=tk.W, pady=5, padx=(0,5))
        
        # Create a frame to contain both text widget and scrollbar in the same row as the label
        desc_frame = ttk.Frame(parent)
        desc_frame.grid(row=row, column=1, sticky=tk.EW, pady=5)
        desc_frame.columnconfigure(0, weight=1)  # Make text widget expand
        
        self.description_text = tk.Text(desc_frame, height=4, wrap=tk.WORD)
        self.description_text.grid(row=0, column=0, sticky=tk.EW)
        if task and task.get("description"):
            self.description_text.insert("1.0", task["description"])
        
        # Add scrollbar to description inside the frame
        desc_scrollbar = ttk.Scrollbar(desc_frame, orient="vertical", command=self.description_text.yview)
        desc_scrollbar.grid(row=0, column=1, sticky=tk.NS)
        self.description_text.configure(yscrollcommand=desc_scrollbar.set)
        
        # Priority
        row += 1
        ttk.Label(parent, text="Priority:").grid(row=row, column=0, sticky=tk.W, pady=5)
        priority_frame = ttk.Frame(parent)
        priority_frame.grid(row=row, column=1, sticky=tk.W, pady=5)
        
        self.priority_var = tk.StringVar(value=task["priority"] if task else "medium")
        ttk.Radiobutton(priority_frame, text="Low", variable=self.priority_var, value="low").pack(side=tk.LEFT, padx=5)
        ttk.Radiobutton(priority_frame, text="Medium", variable=self.priority_var, value="medium").pack(side=tk.LEFT, padx=5)
        ttk.Radiobutton(priority_frame, text="High", variable=self.priority_var, value="high").pack(side=tk.LEFT, padx=5)
        
        # Category
        row += 1
        ttk.Label(parent, text="Category:").grid(row=row, column=0, sticky=tk.W, pady=5)
        self.category_var = tk.StringVar(value=task["category"] if task else "general")
        ttk.Entry(parent, textvariable=self.category_var).grid(row=row, column=1, sticky=tk.EW, pady=5)
        
        # Main Staff
        row += 1
        ttk.Label(parent, text="Main Staff:").grid(row=row, column=0, sticky=tk.W, pady=5)
        self.main_staff_var = tk.StringVar(value=task.get("main_staff", "") if task else "")
        self.main_staff_combo = ttk.Combobox(parent, textvariable=self.main_staff_var, 
                                           values=USERS[1:], state="readonly")
        self.main_staff_combo.grid(row=row, column=1, sticky=tk.EW, pady=5)
        self.disable_combobox_mousewheel(self.main_staff_combo)  # Disable mousewheel scrolling
        row += 1
        
        # Assigned To
        ttk.Label(parent, text="Assigned To:").grid(row=row, column=0, sticky=tk.W, pady=5)
        self.assigned_to_var = tk.StringVar(value=task.get("assigned_to", "") if task else "")
        self.assigned_to_combo = ttk.Combobox(parent, textvariable=self.assigned_to_var, 
                                            values=USERS[1:], state="readonly")
        self.assigned_to_combo.grid(row=row, column=1, sticky=tk.EW, pady=5)
        self.disable_combobox_mousewheel(self.assigned_to_combo)  # Disable mousewheel scrolling
        row += 1
        
        # Qtd Mhr
        ttk.Label(parent, text="Qtd Mhr:").grid(row=row, column=0, sticky=tk.W, pady=5)
        self.qtd_mhr_var = tk.StringVar(value=str(task.get("qtd_mhr", "")) if task and task.get("qtd_mhr") is not None else "") # Ensure default is empty string
        ttk.Entry(parent, textvariable=self.qtd_mhr_var).grid(row=row, column=1, sticky=tk.EW, pady=5)
        row += 1

        # Actual Mhr
        ttk.Label(parent, text="Actual Mhr:").grid(row=row, column=0, sticky=tk.W, pady=5)
        self.actual_mhr_var = tk.StringVar(value=str(task.get("actual_mhr", "")) if task and task.get("actual_mhr") is not None else "") # Ensure default is empty string
        ttk.Entry(parent, textvariable=self.actual_mhr_var).grid(row=row, column=1, sticky=tk.EW, pady=5)
        row += 1

        # Status (only for editing)
        if task is not None:
            row += 1
            ttk.Label(parent, text="Status:").grid(row=row, column=0, sticky=tk.W, pady=5)
            self.completed_var = tk.BooleanVar(value=task["completed"])
            ttk.Checkbutton(parent, text="Completed", variable=self.completed_var).grid(row=row, column=1, sticky=tk.W, pady=5)
    
    def disable_combobox_mousewheel(self, combobox):
        """Disable mousewheel scrolling for a combobox to prevent accidental changes."""
        def block_mousewheel(event):
            return "break"
        
        # Bind mousewheel events for different platforms
        combobox.bind("<MouseWheel>", block_mousewheel)  # Windows
        combobox.bind("<Button-4>", block_mousewheel)    # Linux
        combobox.bind("<Button-5>", block_mousewheel)    # Linux
        # Remove the incorrect macOS binding that was causing the error
    
    def browse_folder(self):
        """Browse for a folder and set it as the link."""
        # Get the applied vessel value
        applied_vessel = self.applied_vessel_var.get().strip()
        
        # Check if applied vessel is provided
        if not applied_vessel:
            messagebox.showwarning("Missing Information", 
                                 "Please enter a value for Applied Vessel before browsing.",
                                 parent=self.dialog)
            return
        
        # Define the base project path
        base_path = r"\\srb096154\01_CESSD_SCG_CAD\01_Projects"
        vessel_path = os.path.join(base_path, applied_vessel)
        
        # Check if the vessel folder exists
        if not os.path.exists(vessel_path):
            # Ask if the user wants to create the folder
            create_folder = messagebox.askyesno(
                "Create Folder",
                f"Folder for vessel '{applied_vessel}' does not exist.\n\nDo you want to create it?",
                parent=self.dialog
            )
            
            if create_folder:
                try:
                    # Create the folder
                    os.makedirs(vessel_path, exist_ok=True)
                except Exception as e:
                    messagebox.showerror(
                        "Error",
                        f"Could not create folder: {vessel_path}\n\nError: {str(e)}",
                        parent=self.dialog
                    )
                    return
            else:
                return
        
        # Open folder dialog
        try:
            from tkinter import filedialog
            
            # Use the vessel path as the initial directory if it exists
            initial_dir = vessel_path if os.path.exists(vessel_path) else os.path.dirname(vessel_path)
            
            folder_path = filedialog.askdirectory(
                title="Select Folder",
                initialdir=initial_dir,
                parent=self.dialog
            )
            
            # Update the link field if a folder was selected
            if folder_path:
                self.link_var.set(folder_path)
        except Exception as e:
            messagebox.showerror(
                "Error",
                f"An error occurred while browsing for folders: {str(e)}",
                parent=self.dialog
            )
    
    def save(self):
        """Save the form data."""
        # Validate required fields
        validation_errors = []
        
        # Request No validation
        request_no = self.request_no_var.get().strip()
        if not request_no:
            validation_errors.append("Request No. is required")
        
        # Equipment Name validation
        title = self.title_var.get().strip()
        if not title:
            validation_errors.append("Equipment Name is required")
        
        # Applied Vessel validation
        applied_vessel = self.applied_vessel_var.get().strip()
        if not applied_vessel:
            validation_errors.append("Applied Vessel is required")
        
        # Rev validation
        rev = self.rev_var.get().strip()
        if not rev:
            validation_errors.append("Rev is required")
        else:
            if not rev.isdigit(): # Check if rev is numeric
                validation_errors.append("Rev must be a valid integer")
            else:
                try:
                    rev_num = int(rev)
                    if rev_num < 0: # Check if rev is non-negative
                        validation_errors.append("Rev must be a non-negative integer")
                    else:
                        self.rev_var.set(str(rev_num)) # Store the converted integer back
                except ValueError: # Redundant, but kept for clarity
                    validation_errors.append("Rev must be a valid integer")
        
        # Link validation
        link = self.link_var.get().strip()
        if not link:
            validation_errors.append("Link is required")
        
        # SDB Link validation
        sdb_link = self.sdb_link_var.get().strip()
        if not sdb_link:
            validation_errors.append("SDB Link is required")
        
        # Main Staff validation
        main_staff = self.main_staff_var.get().strip()
        if not main_staff:
            validation_errors.append("Main Staff is required")
        
        # Priority validation
        priority = self.priority_var.get().lower()
        if priority not in ["low", "medium", "high"]:
            validation_errors.append("Priority must be one of: Low, Medium, High")
        
        # Category validation (example: alphanumeric and spaces only)
        category = self.category_var.get().strip()
        if not category:
            validation_errors.append("Category is required")
        elif not category.replace(" ", "").isalnum(): # Allow alphanumeric and spaces
            validation_errors.append("Category must be alphanumeric and spaces only")
        
        # Main Staff and Assigned To validation (check against USERS list)
        main_staff = self.main_staff_var.get()
        if main_staff and main_staff not in USERS[1:]: # USERS[1:] to exclude "All"
            validation_errors.append(f"Main Staff must be one of: {', '.join(USERS[1:])}")

        assigned_to = self.assigned_to_var.get()
        if assigned_to and assigned_to not in USERS[1:]: # USERS[1:] to exclude "All"
            validation_errors.append(f"Assigned To must be one of: {', '.join(USERS[1:])}")
        
        # Qtd Mhr validation
        qtd_mhr_str = self.qtd_mhr_var.get().strip()
        if qtd_mhr_str: # Only validate if not empty
            if not qtd_mhr_str.isdigit():
                validation_errors.append("Qtd Mhr must be an integer")
            else:
                try:
                    qtd_mhr = int(qtd_mhr_str)
                    if qtd_mhr < 0:
                        validation_errors.append("Qtd Mhr must be a non-negative integer")
                except ValueError: # Redundant, but kept for clarity
                    validation_errors.append("Qtd Mhr must be an integer")
        else:
            qtd_mhr = 0 # Default to 0 if empty

        # Actual Mhr validation
        actual_mhr_str = self.actual_mhr_var.get().strip()
        if actual_mhr_str: # Only validate if not empty
            if not actual_mhr_str.isdigit():
                validation_errors.append("Actual Mhr must be an integer")
            else:
                try:
                    actual_mhr = int(actual_mhr_str)
                    if actual_mhr < 0:
                        validation_errors.append("Actual Mhr must be a non-negative integer")
                except ValueError: # Redundant, but kept for clarity
                    validation_errors.append("Actual Mhr must be an integer")
        else:
            actual_mhr = 0 # Default to 0 if empty

        # Show all validation errors if any
        if validation_errors:
            error_message = "Please correct the following errors:\n\n" + "\n".join(f"• {error}" for error in validation_errors)
            messagebox.showerror("Validation Error", error_message, parent=self.dialog)
            return
        
        # Get dates - ensure we get the raw string value from the picker
        try:
            requested_date = self.requested_date_picker.get()
            if requested_date:
                # Store as string in YYYY-MM-DD format
                requested_date = requested_date
        except Exception as e:
            print(f"Error getting requested date: {str(e)}")
            requested_date = None
            
        try:
            due_date = self.due_date_picker.get()
            if due_date:
                # Store as string in YYYY-MM-DD format
                due_date = due_date
        except Exception as e:
            print(f"Error getting due date: {str(e)}")
            due_date = None
        
        try:
            target_start = self.target_start_picker.get()
            if target_start:
                # Store as string in YYYY-MM-DD format
                target_start = target_start
        except Exception as e:
            print(f"Error getting target start date: {str(e)}")
            target_start = None
        
        try:
            target_finish = self.target_finish_picker.get()
            if target_finish:
                # Store as string in YYYY-MM-DD format
                target_finish = target_finish
        except Exception as e:
            print(f"Error getting target finish date: {str(e)}")
            target_finish = None
        
        # Create result with all fields
        self.result = {
            "request_no": request_no,
            "title": title,
            "description": self.description_text.get("1.0", tk.END).strip() or None,  # Make None if empty
            "requested_date": requested_date,
            "due_date": due_date,
            "target_start": target_start,
            "target_finish": target_finish,
            "priority": self.priority_var.get(),
            "category": self.category_var.get().strip() or "general",
            "main_staff": main_staff,
            "assigned_to": self.assigned_to_var.get().strip() or None,
            "applied_vessel": applied_vessel,
            "rev": rev,
            "drawing_no": self.drawing_no_var.get().strip() or None,  # Make None if empty
            "link": link,
            "sdb_link": sdb_link,
            "qtd_mhr": qtd_mhr if qtd_mhr_str else 0, # Use validated integer value, default to 0 if empty
            "actual_mhr": actual_mhr if actual_mhr_str else 0, # Use validated integer value, default to 0 if empty
        }
        
        # Add completed status if editing
        if hasattr(self, "completed_var"):
            self.result["completed"] = self.completed_var.get()
        
        # Unbind mousewheel events before destroying the dialog
        self.unbind_mousewheel()
        
        # Close dialog
        self.dialog.destroy()

    def scrape_data(self):
        """Handles scraping data from SmartDB using the provided URL."""
        logging.info("Starting scrape_data method.") # Log start
        url = self.scrape_url_var.get().strip()
        if not url:
            messagebox.showwarning("Missing URL", "Please enter a SmartDB URL to scrape.", parent=self.dialog)
            logging.warning("Scrape cancelled: No URL provided.")
            return

        # --- Set SDB Link field with the Scrape URL ---
        self.sdb_link_var.set(url)
        logging.info(f"Set SDB Link field to: {url}")
        # ----------------------------------------------

        # --- Check Applied Vessel and prompt if needed ---
        applied_vessel = self.applied_vessel_var.get().strip()
        if not applied_vessel:
            logging.info("Applied Vessel is missing, prompting user.")
            # Prompt user for Applied Vessel
            applied_vessel = simpledialog.askstring("Applied Vessel Required",
                                                  "Please enter the Applied Vessel number:",
                                                  parent=self.dialog)
            if not applied_vessel:  # User cancelled or entered empty string
                messagebox.showwarning("Missing Information",
                                     "Applied Vessel is required to proceed.",
                                     parent=self.dialog)
                logging.warning("Scrape cancelled: User did not provide Applied Vessel.")
                return
            # Set the Applied Vessel field
            self.applied_vessel_var.set(applied_vessel)
            logging.info(f"Applied Vessel set by user: {applied_vessel}")

        # --- Define standard location from config and construct base path ---
        standard_location = app_config.get('Paths', 'standard_location', 
                                          fallback=r'\\srb096154\01_CESSD_SCG_CAD\01_Projects')
        base_vessel_path = os.path.join(standard_location, applied_vessel)
        logging.info(f"Base vessel path defined: {base_vessel_path}")

        # Check if the vessel folder exists
        if not os.path.exists(base_vessel_path):
            logging.warning(f"Base vessel folder does not exist: {base_vessel_path}")
            create_folder = messagebox.askyesno(
                "Create Folder",
                f"Folder for vessel '{applied_vessel}' does not exist at:\n{base_vessel_path}\n\nDo you want to create it?",
                parent=self.dialog
            )

            if create_folder:
                try:
                    os.makedirs(base_vessel_path, exist_ok=True)
                    logging.info(f"Created base vessel folder: {base_vessel_path}")
                except Exception as e:
                    messagebox.showerror(
                        "Error",
                        f"Could not create folder: {base_vessel_path}\n\nError: {str(e)}",
                        parent=self.dialog
                    )
                    logging.error(f"Failed to create base vessel folder: {e}", exc_info=True)
                    return
            else:
                logging.warning("Scrape cancelled: User chose not to create base vessel folder.")
                return

        # Show info message about folder selection
        logging.info("Prompting user to select final folder structure.")
        messagebox.showinfo("Select Final Folder",
                          f"You will now be prompted to select or create the final folder structure.\n\n"
                          f"Please navigate to or create the desired subfolder structure\n"
                          f"(e.g., SmartDB/Equipment/MAIN ENGINE)\n\n"
                          f"Starting from: {base_vessel_path}",
                          parent=self.dialog)

        final_target_folder = filedialog.askdirectory(
            title="Select Final Folder",
            initialdir=base_vessel_path,  # Start from the vessel folder
            parent=self.dialog
        )

        if not final_target_folder:
            messagebox.showwarning("Folder Selection Cancelled",
                                 "No final folder selected. Downloads will be skipped, but scraping will proceed.", # Adjusted message
                                 parent=self.dialog)
            logging.warning("No final folder selected by user. Downloads will be skipped.")
            # Allow proceeding without download if user cancels folder selection
            # final_target_folder will be None, which is handled later
        else:
            logging.info(f"Final target folder selected by user: {final_target_folder}")
            # Set the Link field to the final selected path immediately
            self.link_var.set(final_target_folder)
            logging.info(f"Set Link field to final path: {final_target_folder}")


        # --- Create a LOCAL Temporary Directory for Browser Download ---
        temp_download_dir = None
        if final_target_folder: # Only create if needed for download
            try:
                import tempfile
                temp_download_dir = tempfile.mkdtemp(prefix="smartdb_dl_")
                logging.info(f"Created temporary local download directory: {temp_download_dir}")
            except Exception as temp_err:
                logging.error(f"ERROR: Could not create temporary download directory: {temp_err}")
                messagebox.showerror("Error",
                                   f"Could not create a temporary local folder for download:\n{temp_err}\n\nDownloads cannot proceed.",
                                   parent=self.dialog)
                # If temp dir fails, we cannot download, nullify the temp dir variable
                temp_download_dir = None
                # Do not return here, allow text scraping to continue if possible


        # --- Get Credentials ---
        email = None
        password = None
        try:
            # Lazy import: only load heavy smartdb_login when actually needed
            import smartdb_login
            logging.info("Getting email via smartdb_login.get_email_address...")
            email = smartdb_login.get_email_address(smartdb_login.USER_EMAIL)
            if not email:
                messagebox.showerror("Authentication Error", "Email address is required for scraping.", parent=self.dialog)
                logging.error("Scrape failed: Email address not obtained.")
                if temp_download_dir: shutil.rmtree(temp_download_dir, ignore_errors=True)
                return
            logging.info(f"Using email: {email}")

            logging.info("Getting password via smartdb_login.store_or_get_password...")
            password = smartdb_login.store_or_get_password(smartdb_login.KEYRING_SERVICE_NAME, email)
            if not password:
                messagebox.showerror("Authentication Error", "Password is required for scraping.", parent=self.dialog)
                logging.error("Scrape failed: Password not obtained.")
                if temp_download_dir: shutil.rmtree(temp_download_dir, ignore_errors=True)
                return
            logging.info("Password obtained successfully.")

        except Exception as auth_e:
            messagebox.showerror("Authentication Error", f"Failed to get credentials: {auth_e}", parent=self.dialog)
            logging.error(f"Error getting credentials for scraping: {auth_e}", exc_info=True)
            if temp_download_dir: shutil.rmtree(temp_download_dir, ignore_errors=True)
            return

        # --- Start Timer ---
        start_time = time.monotonic()
        logging.info("Starting scraping process timer.")

        # --- Perform Scraping ---
        messagebox.showinfo("Scraping Started",
                            "Starting web scraping process...\nBrowser may run in the background.\nFiles will download to a temporary folder first.\nThis may take a moment.",
                            parent=self.dialog)
        self.dialog.update_idletasks() # Ensure message box is shown

        extracted_data = None
        success = False
        driver = None # Initialize driver variable
        temp_user_data_dir_path = None # <<< Initialize variable for user data dir path
        moved_files_summary = []
        download_info = {}

        try:
            logging.info(f"Calling smartdb_login.login_to_smartdb with URL: {url}")
            # Pass the LOCAL TEMP directory for downloads if it was created
            # Capture the returned temp_user_data_dir_path
            success, extracted_data, driver, temp_user_data_dir_path = smartdb_login.login_to_smartdb( # <<< Capture path
                url, email, password, target_download_folder=temp_download_dir
            )
            logging.info(f"Login/Scrape process finished. Success: {success}")
            if temp_user_data_dir_path:
                logging.info(f"WebDriver used temporary user data directory: {temp_user_data_dir_path}")
            if extracted_data:
                 logging.info(f"Extracted data keys: {list(extracted_data.keys())}")
                 download_info = extracted_data.get("Downloaded Files", {})
                 logging.info(f"Download info found: {len(download_info)} item(s)")

            # --- Populate Text Fields (if successful) ---
            if success and extracted_data:
                logging.info("Populating dialog text fields from extracted data...")
                self.populate_fields_from_scrape(extracted_data) # Populate text fields first
            elif success and not extracted_data:
                 logging.warning("Login successful, but no text data found to populate fields.")
                 messagebox.showwarning("No Text Data", "Login successful, but no text data found to populate fields.", parent=self.dialog)
            elif not success:
                 logging.error("Login or text data scraping failed.")
                 messagebox.showerror("Scraping Failed", "Login or text data scraping failed. Check logs for details.", parent=self.dialog)
                 # Don't proceed to file move if scraping failed

            # --- Move Downloaded Files from Temp to Final Destination (only if scraping succeeded and folders exist) ---
            if success and final_target_folder and temp_download_dir and download_info:
                logging.info(f"\nMoving downloaded files from '{temp_download_dir}' to '{final_target_folder}'...")

                for itemkey, info in download_info.items():
                    # Check status from smartdb_login (should be "Success" if downloaded to temp)
                    if info.get("status") == "Success" and "path" in info:
                        temp_file_path = info["path"]
                        # Double check the temp file exists before moving
                        if not os.path.exists(temp_file_path):
                             logging.warning(f"  Skipping move: Temp file '{temp_file_path}' not found for itemkey '{itemkey}'.")
                             info["error"] = "Temporary file missing before move"
                             info["status"] = "Failed (Move Error)"
                             continue

                        original_filename = os.path.basename(temp_file_path)
                        final_destination_path = os.path.join(final_target_folder, original_filename)

                        try:
                            logging.info(f"  Moving '{original_filename}' to '{final_destination_path}'...")
                            # Ensure final directory exists (should have been checked, but double-check)
                            os.makedirs(final_target_folder, exist_ok=True)

                            # Handle potential overwrite? For now, move will overwrite by default.
                            # Add check if needed:
                            # if os.path.exists(final_destination_path):
                            #     logging.warning(f"    Destination file '{final_destination_path}' already exists. Overwriting.")

                            shutil.move(temp_file_path, final_destination_path)
                            logging.info(f"  Successfully moved: {original_filename}")
                            moved_files_summary.append(original_filename)

                            # Update the path in extracted_data to the final location
                            info["path"] = final_destination_path
                            info["status"] = "Success (Moved)" # Update status

                        except Exception as move_err:
                            logging.error(f"  ERROR moving file '{original_filename}': {move_err}", exc_info=True)
                            info["error"] = f"Move Failed: {move_err}"
                            info["status"] = "Failed (Move Error)" # Update status if move fails
                    elif info.get("status") == "Failed":
                         logging.warning(f"  Skipping move for itemkey '{itemkey}': Initial download failed.")
                    # Add case for if status is already "Success (Moved)" - shouldn't happen here but good practice
                    elif info.get("status") == "Success (Moved)":
                         logging.info(f"  File for itemkey '{itemkey}' already marked as moved.")


                # --- Update the Link field to the FINAL TARGET FOLDER if files were moved ---
                # This is slightly redundant as we set it earlier, but confirms it reflects the download/move outcome
                if moved_files_summary and final_target_folder: # Check if any files were successfully moved
                    self.link_var.set(final_target_folder) # Set Link to the folder path
                    logging.info(f"Confirmed Link field remains set to target directory after move: {final_target_folder}")
                    # Update the key in extracted_data as well for consistency, pointing to the folder
                    if extracted_data: extracted_data["Primary Downloaded File Path"] = final_target_folder
                elif final_target_folder: # If a folder was selected but no files were moved (or found)
                     self.link_var.set(final_target_folder) # Ensure it's still set
                     logging.info(f"Link field remains set to selected target folder (no files moved): {final_target_folder}")


            # Prepare download result message based on final status (including move results)
            download_result_message = ""
            if final_target_folder: # Check if a target folder was selected initially
                 # Recalculate counts based on final status in download_info
                 success_moved_count = sum(1 for info in download_info.values() if info.get("status") == "Success (Moved)")
                 fail_move_count = sum(1 for info in download_info.values() if info.get("status") == "Failed (Move Error)")
                 fail_dl_count = sum(1 for info in download_info.values() if info.get("status") == "Failed") # Original download failures

                 if success_moved_count > 0:
                     download_result_message += f"\n\nSuccessfully downloaded and moved {success_moved_count} file(s) to:\n{final_target_folder}"
                 if fail_move_count > 0:
                     download_result_message += f"\n\nEncountered errors moving {fail_move_count} downloaded file(s)."
                 if fail_dl_count > 0:
                     download_result_message += f"\n\nFailed to download {fail_dl_count} file(s) initially."
                 # Check if scraping was successful but no downloadable files were found/processed
                 if success and not download_info and success_moved_count==0 and fail_move_count==0 and fail_dl_count==0:
                      download_result_message = "\n\nNo files found to download."
                 # If scraping failed, don't report download status
                 elif not success:
                      download_result_message = ""

            elif not final_target_folder: # User cancelled selection or temp dir failed
                 download_result_message = "\n\nAttachment download skipped (no final folder specified/accessible)."


        except Exception as scrape_e:
            success = False # Ensure success is False
            messagebox.showerror("Scraping Process Error", f"An unexpected error occurred during scraping/moving: {scrape_e}", parent=self.dialog)
            # Log the error including traceback
            logging.error(f"Error during scrape_data execution: {scrape_e}", exc_info=True) # <<< Ensure traceback is logged
        finally:
            # --- Stop Timer ---
            elapsed_time = time.monotonic() - start_time
            elapsed_time_str = f"{elapsed_time:.2f} seconds"
            logging.info(f"Scraping process finished. Total time: {elapsed_time_str}")

            # --- Ensure driver is closed ---
            if driver:
                logging.info("Closing WebDriver...")
                try:
                    driver.quit()
                    logging.info("WebDriver closed.")
                except Exception as quit_e:
                    logging.error(f"Error closing WebDriver: {quit_e}", exc_info=True)

            # --- Clean up Temporary DOWNLOAD Directory ---
            if temp_download_dir and os.path.exists(temp_download_dir): # Check existence
                try:
                    logging.info(f"Cleaning up temporary DOWNLOAD directory: {temp_download_dir}")
                    shutil.rmtree(temp_download_dir, ignore_errors=True)
                    logging.info("Temporary download directory removal attempted.")
                except Exception as clean_err:
                    logging.warning(f"Could not remove temporary download directory '{temp_download_dir}': {clean_err}")

            # --- Clean up Temporary USER DATA Directory ---
            if temp_user_data_dir_path and os.path.exists(temp_user_data_dir_path): # <<< Add cleanup here
                try:
                    logging.info(f"Cleaning up temporary USER DATA directory: {temp_user_data_dir_path}")
                    # Use ignore_errors=True for robustness, as sometimes files might be locked briefly after browser close
                    shutil.rmtree(temp_user_data_dir_path, ignore_errors=True)
                    logging.info("Temporary user data directory removal attempted.")
                except Exception as clean_user_err:
                    # Log warning, but don't stop execution
                    logging.warning(f"Could not remove temporary user data directory '{temp_user_data_dir_path}': {clean_user_err}")
            elif temp_user_data_dir_path:
                 logging.warning(f"Temporary user data directory path was set ('{temp_user_data_dir_path}') but directory not found for cleanup.")


            # --- Show Final Status Message ---
            final_message = ""
            if success:
                final_message = "Scraping process finished."
                # Add download/move info message
                final_message += download_result_message
            else:
                # Check if it was just a login failure or a broader exception
                if extracted_data is None: # Indicates likely exception before or during login call
                     final_message = "Scraping process failed due to an unexpected error before completion."
                else: # Login call returned success=False
                     final_message = "Scraping process failed during login or data extraction."


            final_message += f"\n\nTotal time: {elapsed_time_str}"
            messagebox.showinfo("Scraping Complete", final_message, parent=self.dialog)
            logging.info(f"Final status message shown to user: {final_message}")

    def populate_fields_from_scrape(self, extracted_data):
        """Populates the dialog fields based on the scraped data dictionary."""
        # This function should remain largely the same for populating text fields.
        # The crucial change is HOW the 'Link' field gets its final value, which is now handled
        # in the scrape_data method *after* the file move is attempted.
        # We'll remove the explicit setting of the 'Link' field from here based on "Primary Downloaded File Path"
        # as that key might point to the temp location before the move, or the final location after.
        # scrape_data now handles setting self.link_var directly after the move.

        if not extracted_data:
            return

        populated_fields = []
        print("Populating fields with:", extracted_data)
        try:
            # Handle Request No
            req_no_val = extracted_data.get("Request No", "")
            if req_no_val != "Not Found" and req_no_val:
                self.FIELD_VAR_MAP["Request No"].set(req_no_val)
                populated_fields.append("Request No")

            # Handle Equipment Name (from Scraped "Description")
            equip_name_val = extracted_data.get("Description", "") # Key is "Description" from FIELDS_TO_EXTRACT
            if equip_name_val != "Not Found" and equip_name_val:
                 self.FIELD_VAR_MAP["Equipment Name"].set(equip_name_val)
                 populated_fields.append("Equipment Name (from Scraped Description)")

            # Handle Applied Vessel
            vessel_val = extracted_data.get("Applied Vessel", "")
            if vessel_val != "Not Found" and vessel_val:
                 self.FIELD_VAR_MAP["Applied Vessel"].set(vessel_val)
                 populated_fields.append("Applied Vessel")

            # Handle Revision No
            rev_val = extracted_data.get("Revision No", "")
            if rev_val != "Not Found" and rev_val:
                 # Check for 'O' and replace with '0'
                 if rev_val.strip().upper() == 'O':
                     rev_val = '0'
                     print("  Interpreted Revision 'O' as '0'.")
                 self.FIELD_VAR_MAP["Revision No"].set(rev_val)
                 populated_fields.append("Revision No")

            # Handle Drawing No
            dwg_no_val = extracted_data.get("Drawing No", "")
            if dwg_no_val != "Not Found" and dwg_no_val:
                 self.FIELD_VAR_MAP["Drawing No"].set(dwg_no_val)
                 populated_fields.append("Drawing No")

            # --- Link field is now handled AFTER potential move in scrape_data ---
            # Remove this block:
            # link_path_val = extracted_data.get("Primary Downloaded File Path", "")
            # if link_path_val:
            #      self.FIELD_VAR_MAP["Link"].set(link_path_val)
            #      populated_fields.append("Link (Downloaded File Path)")

            # Handle Description (from Scraped "Note")
            note_val = extracted_data.get("Note", "") # Key is "Note"
            desc_widget = self.FIELD_VAR_MAP["Description"]
            if note_val != "Not Found" and note_val and isinstance(desc_widget, tk.Text):
                 desc_widget.delete('1.0', tk.END)
                 desc_widget.insert('1.0', note_val)
                 populated_fields.append("Description (from Scraped Note)")

            # Handle Modeling Staff - Find first name match in known USERS
            staff_val = extracted_data.get("Modeling Staff", "")
            matched_staff = None
            if staff_val and staff_val != "Not Found":
                staff_val_lower = staff_val.lower()
                for known_staff_name in USERS[1:]: # Iterate through valid staff first names
                    if known_staff_name.lower() in staff_val_lower:
                        matched_staff = known_staff_name
                        print(f"  Matched Modeling Staff '{staff_val}' to known staff '{matched_staff}'")
                        break # Found the first match

            if matched_staff:
                 self.FIELD_VAR_MAP["Main Staff"].set(matched_staff)
                 populated_fields.append("Main Staff (from Modeling Staff)")
                 # Optionally set Assigned To as well (if desired):
                 # self.FIELD_VAR_MAP["Assigned To"].set(matched_staff)
                 # if "Assigned To" not in populated_fields: populated_fields.append("Assigned To")
            elif staff_val and staff_val != "Not Found": # If staff name was scraped but no match found
                 print(f"Warning: Scraped Modeling Staff '{staff_val}' did not match any known staff in USERS list.")

            # --- Handle Dates ---
            date_format = "%Y/%m/%d" # Format used in SmartDB example

            # Handle Requested Date
            req_date_str = extracted_data.get("Requested Date", "") # Key is "Requested Date"
            if req_date_str and req_date_str != "Not Found":
                try:
                    req_date_obj = datetime.datetime.strptime(req_date_str, date_format).date()
                    target_widget = self.FIELD_VAR_MAP["Requested Date"]
                    if isinstance(target_widget, DateEntry):
                         target_widget.set_date(req_date_obj)
                         populated_fields.append("Requested Date")
                    else:
                         print(f"Warning: Target widget for Requested Date is not a DateEntry ({type(target_widget)}).")
                except ValueError:
                    print(f"Warning: Could not parse Requested Date '{req_date_str}' with format {date_format}")
                except Exception as date_err:
                     print(f"Error setting Requested Date: {date_err}")

            # Handle Due Date (from Scraped "Deadline")
            due_date_str = extracted_data.get("Deadline", "") # Key is "Deadline"
            if due_date_str and due_date_str != "Not Found":
                 try:
                     # Clean the date string (remove day in parentheses)
                     cleaned_due_date_str = due_date_str.split('(')[0].strip()
                     print(f"  Parsing cleaned Deadline: '{cleaned_due_date_str}'")
                     due_date_obj = datetime.datetime.strptime(cleaned_due_date_str, date_format).date()
                     target_widget = self.FIELD_VAR_MAP["Due Date"]
                     if isinstance(target_widget, DateEntry):
                         target_widget.set_date(due_date_obj)
                         populated_fields.append("Due Date (from Scraped Deadline)")
                     else:
                          print(f"Warning: Target widget for Due Date is not a DateEntry ({type(target_widget)}).")
                 except ValueError:
                     print(f"Warning: Could not parse Due Date (Deadline) '{cleaned_due_date_str}' with format {date_format}")
                 except Exception as date_err:
                     print(f"Error setting Due Date: {date_err}")

            # --- Final Feedback ---
            if populated_fields:
                print(f"Successfully populated fields: {', '.join(populated_fields)}")
                # Message box is now shown after download attempt/move in the main scrape_data flow
            else:
                print("No matching text data found to populate fields.")
                # Message box is now shown after download attempt/move in the main scrape_data flow

        except Exception as populate_e:
             logging.error(f"Error populating fields: {populate_e}", exc_info=True)
             raise # Re-raise to be caught by the main try block

class SchedulerDialog:
    """Dialog for viewing and auto-assigning tasks to balance workload."""
    
    def __init__(self, parent, task_manager, callback=None):
        """Initialize the scheduler dialog.
        
        Args:
            parent: Parent window
            task_manager: TaskManager instance
            callback: Function to call after applying changes
        """
        self.parent = parent
        self.task_manager = task_manager
        self.callback = callback
        self.dialog = tk.Toplevel(parent)
        self.dialog.withdraw()
        
        self.dialog.title("Task Scheduler - Workload Optimizer")
        self.dialog.geometry("1400x800")
        self.dialog.minsize(1200, 700)
        self.dialog.transient(parent)
        
        # Schedule data: {user: {date: [task_ids]}}
        self.schedule = {}
        self.holidays = set()  # Set of dates that are holidays/non-working days
        self.task_hours_per_day = {}  # {user: {date: {task_id: hours}}} - split hours tracking
        self.proposed_changes = False
        self.save_enabled = False  # Will be set to True if database tables exist
        
        # Date range settings
        self.days_to_show = 21  # 3 weeks per view
        self.start_date = datetime.date.today()  # Starting date for the view
        
        # Center window
        self.dialog.update_idletasks()
        x = (self.dialog.winfo_screenwidth() // 2) - (1400 // 2)
        y = (self.dialog.winfo_screenheight() // 2) - (800 // 2)
        self.dialog.geometry(f"1400x800+{x}+{y}")
        
        self._create_ui()
        self._ensure_schedule_tables_exist()
        self._auto_mark_all_sundays()  # Automatically mark Sundays
        self._load_saved_schedule()  # Load saved schedule if exists
        
        self.dialog.deiconify()
        self.dialog.grab_set()
    
    def _create_ui(self):
        """Create the scheduler UI."""
        # Main container
        main_frame = ttk.Frame(self.dialog, padding=15)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Title and info
        header_frame = ttk.Frame(main_frame)
        header_frame.pack(fill=tk.X, pady=(0, 10))
        
        title_label = ttk.Label(
            header_frame, 
            text="📅 Task Scheduler & Workload Optimizer",
            style="Title.TLabel"
        )
        title_label.pack(side=tk.LEFT)
        
        self.info_label = ttk.Label(
            header_frame,
            text="Analyzing workload...",
            style="TLabel"
        )
        self.info_label.pack(side=tk.RIGHT, padx=10)
        
        # Control buttons
        control_frame = ttk.Frame(main_frame)
        control_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Button(
            control_frame,
            text="Generate Schedule",
            command=self._auto_distribute_tasks
        ).pack(side=tk.LEFT, padx=5)
        
        ttk.Button(
            control_frame,
            text="Refresh Analysis",
            command=self._analyze_current_schedule
        ).pack(side=tk.LEFT, padx=5)
        
        # Date navigation controls
        nav_frame = ttk.Frame(control_frame)
        nav_frame.pack(side=tk.LEFT, padx=20)
        
        ttk.Button(
            nav_frame,
            text="◀ Prev Month",
            command=self._prev_month,
            width=12
        ).pack(side=tk.LEFT, padx=2)
        
        self.date_range_label = ttk.Label(
            nav_frame,
            text="",
            font=("Arial", 9, "bold")
        )
        self.date_range_label.pack(side=tk.LEFT, padx=10)
        
        ttk.Button(
            nav_frame,
            text="Next Month ▶",
            command=self._next_month,
            width=12
        ).pack(side=tk.LEFT, padx=2)
        
        ttk.Button(
            nav_frame,
            text="Today",
            command=self._go_to_today,
            width=8
        ).pack(side=tk.LEFT, padx=10)
        
        ttk.Label(
            control_frame,
            text="(Click date headers to toggle working/non-working days)",
            font=("Arial", 8, "italic")
        ).pack(side=tk.LEFT, padx=15)
        
        # Calendar grid container with scrollbars
        grid_container = ttk.Frame(main_frame)
        grid_container.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
        
        # Create canvas with scrollbars
        h_scroll = ttk.Scrollbar(grid_container, orient=tk.HORIZONTAL)
        h_scroll.pack(side=tk.BOTTOM, fill=tk.X)
        
        v_scroll = ttk.Scrollbar(grid_container, orient=tk.VERTICAL)
        v_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.canvas = tk.Canvas(
            grid_container,
            xscrollcommand=h_scroll.set,
            yscrollcommand=v_scroll.set,
            bg="white",
            highlightthickness=0
        )
        self.canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        h_scroll.config(command=self.canvas.xview)
        v_scroll.config(command=self.canvas.yview)
        
        # Frame inside canvas for the grid
        self.grid_frame = tk.Frame(self.canvas, bg="white")
        self.canvas_window = self.canvas.create_window((0, 0), window=self.grid_frame, anchor="nw")
        
        # Bind canvas resize
        self.grid_frame.bind("<Configure>", lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all")))
        
        # Button frame
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X, pady=(10, 0))
        
        ttk.Button(
            button_frame,
            text="Close",
            command=self.dialog.destroy
        ).pack(side=tk.RIGHT, padx=5)
        
        self.save_button = ttk.Button(
            button_frame,
            text="Save Schedule",
            command=self._save_schedule
            # Always enabled - will use file or database
        )
        self.save_button.pack(side=tk.RIGHT, padx=5)
    
    def _analyze_current_schedule(self):
        """Analyze current task assignments and build schedule grid."""
        try:
            # Get all pending tasks (exclude completed and deleted)
            pending_tasks = [t for t in self.task_manager.tasks if not t.get("completed", False) and not t.get("deleted", False)]
            
            # Get all users
            all_users = list(USER_MAPPING.keys())
            
            # Build current schedule
            self.schedule = {user: {} for user in all_users}
            self.task_hours_per_day = {user: {} for user in all_users}
            
            today = datetime.date.today()
            
            for task in pending_tasks:
                user = task.get("assigned_to", "Unassigned")
                if user == "Unassigned" or user not in all_users:
                    continue
                
                target_finish_str = task.get("target_finish", "")
                if not target_finish_str:
                    continue
                
                try:
                    target_finish = datetime.datetime.strptime(target_finish_str, "%Y-%m-%d").date()
                except (ValueError, TypeError) as e:
                    logging.debug(f"Invalid date format for task {task.get('id')}: {target_finish_str}")
                    continue
                
                # For now, assign task to its target finish date
                if target_finish not in self.schedule[user]:
                    self.schedule[user][target_finish] = []
                
                self.schedule[user][target_finish].append(task["id"])
            
            # Update info
            total_pending = len(pending_tasks)
            self.info_label.config(
                text=f"📊 {len(all_users)} users | {total_pending} pending tasks"
            )
            
            # Render the grid
            self._render_grid()
            
        except Exception as e:
            logging.error(f"Error analyzing schedule: {e}", exc_info=True)
            messagebox.showerror("Error", f"Failed to analyze schedule:\n{str(e)}")
    
    def _render_grid(self):
        """Render the calendar grid."""
        # Clear existing grid
        for widget in self.grid_frame.winfo_children():
            widget.destroy()
        
        all_users = list(USER_MAPPING.keys())
        
        # Update date range label
        end_date = self.start_date + datetime.timedelta(days=self.days_to_show - 1)
        self.date_range_label.config(
            text=f"{self.start_date.strftime('%b %d, %Y')} - {end_date.strftime('%b %d, %Y')}"
        )
        
        # Cell dimensions
        cell_width = 60
        cell_height = 30
        header_height = 50
        user_col_width = 80
        
        # Header row - dates
        tk.Label(
            self.grid_frame,
            text="User",
            width=10,
            height=2,
            bg="#2c3e50",
            fg="white",
            font=("Arial", 10, "bold"),
            relief=tk.RAISED,
            borderwidth=1
        ).grid(row=0, column=0, sticky="nsew")
        
        # Date headers
        for day_offset in range(self.days_to_show):
            date = self.start_date + datetime.timedelta(days=day_offset)
            date_str = date.strftime("%m/%d")
            day_name = date.strftime("%a")
            
            # Check if it's today
            is_today = (date == datetime.date.today())
            
            # Check if it's Sunday
            is_sunday = (date.weekday() == 6)
            
            # Color based on day type
            if is_sunday:  # Sunday
                bg_color = "#95a5a6"  # Gray for Sundays (auto non-working)
            elif date in self.holidays:
                bg_color = "#3498db"  # Blue for user-marked holidays
            elif is_today:
                bg_color = "#e74c3c"  # Red for today
            else:
                bg_color = "#2c3e50"  # Dark blue for working days
            
            label = tk.Label(
                self.grid_frame,
                text=f"{day_name}\n{date_str}",
                width=8,
                height=2,
                bg=bg_color,
                fg="white",
                font=("Arial", 8, "bold"),
                relief=tk.RAISED,
                borderwidth=1,
                cursor="hand2" if not is_sunday else "arrow"  # No clickable cursor for Sunday
            )
            label.grid(row=0, column=day_offset + 1, sticky="nsew")
            
            # Make date header clickable to toggle working/non-working day (except Sunday)
            if not is_sunday:
                label.bind("<Button-1>", lambda e, d=date: self._toggle_holiday(d))
        
        # User rows
        for user_idx, user in enumerate(all_users):
            # User name cell
            tk.Label(
                self.grid_frame,
                text=user,
                width=10,
                height=1,
                bg="#34495e",
                fg="white",
                font=("Arial", 10, "bold"),
                relief=tk.RAISED,
                borderwidth=1
            ).grid(row=user_idx + 1, column=0, sticky="nsew")
            
            # Date cells for this user
            for day_offset in range(self.days_to_show):
                date = self.start_date + datetime.timedelta(days=day_offset)
                
                # Calculate hours for this user on this date using split hours
                hours = 0.0
                if hasattr(self, 'task_hours_per_day') and user in self.task_hours_per_day:
                    if date in self.task_hours_per_day[user]:
                        # Use the split hours per task for this specific day
                        hours = sum(self.task_hours_per_day[user][date].values())
                else:
                    # Fallback to old method (full task hours) if not available
                    task_ids = self.schedule.get(user, {}).get(date, [])
                    for task_id in task_ids:
                        task = self.task_manager._find_task_by_id(task_id)
                        if task:
                            hours += float(task.get("qtd_mhr", 0) or 0)
                
                # Determine cell color - holidays first, then workload
                if date in self.holidays:
                    bg_color = "#3498db"  # Blue for holidays/non-working days
                    text_color = "white"
                elif hours == 0:
                    bg_color = "#ecf0f1"
                    text_color = "#95a5a6"
                elif hours <= 8:
                    bg_color = "#2ecc71"
                    text_color = "white"
                elif hours <= 10:
                    bg_color = "#f39c12"
                    text_color = "white"
                else:
                    bg_color = "#e74c3c"
                    text_color = "white"
                
                # Create clickable cell
                cell = tk.Label(
                    self.grid_frame,
                    text=f"{hours:.1f}" if hours > 0 else ("🏖️" if date in self.holidays else "-"),
                    width=8,
                    height=1,
                    bg=bg_color,
                    fg=text_color,
                    font=("Arial", 10, "bold"),
                    relief=tk.RAISED,
                    borderwidth=1,
                    cursor="hand2" if hours > 0 else "arrow"
                )
                cell.grid(row=user_idx + 1, column=day_offset + 1, sticky="nsew")
                
                # Make clickable if there are tasks
                if hours > 0:
                    cell.bind("<Button-1>", lambda e, u=user, d=date: self._show_day_tasks(u, d))
    
    def _show_day_tasks(self, user, date):
        """Show tasks for a specific user on a specific date."""
        task_ids = self.schedule.get(user, {}).get(date, [])
        if not task_ids:
            return
        
        # Create popup window
        popup = tk.Toplevel(self.dialog)
        popup.title(f"Tasks for {user} on {date.strftime('%Y-%m-%d')}")
        popup.transient(self.dialog)
        
        # Center popup relative to scheduler dialog
        popup.update_idletasks()
        popup_width = 800
        popup_height = 400
        x = self.dialog.winfo_x() + (self.dialog.winfo_width() - popup_width) // 2
        y = self.dialog.winfo_y() + (self.dialog.winfo_height() - popup_height) // 2
        popup.geometry(f"{popup_width}x{popup_height}+{x}+{y}")
        
        # Main frame
        main_frame = ttk.Frame(popup, padding=10)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Title
        title = ttk.Label(
            main_frame,
            text=f"📋 Tasks for {user} on {date.strftime('%A, %B %d, %Y')}",
            font=("Arial", 12, "bold")
        )
        title.pack(pady=(0, 10))
        
        # Treeview for tasks
        tree_frame = ttk.Frame(main_frame)
        tree_frame.pack(fill=tk.BOTH, expand=True)
        
        scroll = ttk.Scrollbar(tree_frame)
        scroll.pack(side=tk.RIGHT, fill=tk.Y)
        
        tree = ttk.Treeview(
            tree_frame,
            columns=("title", "hours", "delivery", "status"),
            show="headings",
            yscrollcommand=scroll.set
        )
        scroll.config(command=tree.yview)
        
        tree.heading("title", text="Task")
        tree.heading("hours", text="Hours")
        tree.heading("delivery", text="Delivery Date")
        tree.heading("status", text="Status")
        
        tree.column("title", width=400)
        tree.column("hours", width=80, anchor=tk.CENTER)
        tree.column("delivery", width=120, anchor=tk.CENTER)
        tree.column("status", width=150, anchor=tk.CENTER)
        
        tree.pack(fill=tk.BOTH, expand=True)
        
        # Populate tasks
        total_hours = 0.0
        
        # Get split hours for this specific day if available
        day_task_hours = {}
        if hasattr(self, 'task_hours_per_day') and user in self.task_hours_per_day:
            if date in self.task_hours_per_day[user]:
                day_task_hours = self.task_hours_per_day[user][date]
        
        for task_id in task_ids:
            task = self.task_manager._find_task_by_id(task_id)
            if task:
                # Get hours for THIS specific day
                if task_id in day_task_hours:
                    day_hours = day_task_hours[task_id]
                    total_task_hours = float(task.get("qtd_mhr", 0) or 0)
                    hours_display = f"{day_hours:.1f}h of {total_task_hours:.1f}h"
                else:
                    hours = float(task.get("qtd_mhr", 0) or 0)
                    hours_display = f"{hours:.1f}h"
                    day_hours = hours
                
                total_hours += day_hours
                
                tree.insert("", tk.END, values=(
                    task.get("title", "Untitled"),
                    hours_display,
                    task.get("target_finish", "N/A"),
                    task.get("status", "")
                ))
        
        # Summary
        summary = ttk.Label(
            main_frame,
            text=f"Total: {len(task_ids)} tasks, {total_hours:.1f} hours",
            font=("Arial", 10, "bold")
        )
        summary.pack(pady=(10, 0))
        
        # Close button
        ttk.Button(
            main_frame,
            text="Close",
            command=popup.destroy
        ).pack(pady=(10, 0))
    
    def _auto_distribute_tasks(self):
        """Auto-distribute tasks to balance workload across days."""
        try:
            logging.info("Starting auto-distribute tasks...")
            
            if not messagebox.askyesno(
                "Generate Schedule",
                "This will redistribute pending tasks to balance workload.\n\n"
                "Tasks will be scheduled based on:\n"
                "• Delivery dates (urgent first)\n"
                "• 8-hour daily capacity\n"
                "• Can split tasks across multiple days\n\n"
                "Continue?"
            ):
                logging.info("User cancelled schedule generation")
                return
            
            # Get all pending unassigned or poorly scheduled tasks (exclude completed and deleted)
            pending_tasks = [t for t in self.task_manager.tasks if not t.get("completed", False) and not t.get("deleted", False)]
            logging.info(f"Found {len(pending_tasks)} pending tasks")
            
            all_users = list(USER_MAPPING.keys())
            logging.info(f"Users: {all_users}")
            
            today = datetime.date.today()
            
            # Reset schedule
            self.schedule = {user: {} for user in all_users}
            
            # Sort tasks by delivery date (urgent first)
            # Tasks without dates get a default deadline of 7 days from today
            default_deadline = today + datetime.timedelta(days=7)
            tasks_with_dates = []
            tasks_without_dates = 0
            
            for task in pending_tasks:
                target_finish_str = task.get("target_finish", "")
                if target_finish_str:
                    try:
                        target_finish = datetime.datetime.strptime(target_finish_str, "%Y-%m-%d").date()
                        tasks_with_dates.append((task, target_finish))
                    except Exception as date_error:
                        logging.warning(f"Could not parse date '{target_finish_str}': {date_error}")
                        # Add with default deadline if date parsing fails
                        tasks_with_dates.append((task, default_deadline))
                        tasks_without_dates += 1
                else:
                    # No target_finish set - use default deadline
                    tasks_with_dates.append((task, default_deadline))
                    tasks_without_dates += 1
                    logging.info(f"Task {task.get('id')} ({task.get('title', 'Untitled')}) has no target_finish, using default 7-day deadline")
            
            logging.info(f"Found {len(tasks_with_dates)} tasks total ({tasks_without_dates} without dates, using 7-day default)")
            
            if not tasks_with_dates:
                messagebox.showwarning(
                    "No Tasks to Schedule",
                    "No pending tasks found to schedule."
                )
                return
            
            tasks_with_dates.sort(key=lambda x: x[1])
            
            # Distribute tasks - now with proper task splitting across days and load balancing
            distributed_count = 0
            tasks_assigned = set()  # Track which tasks were assigned
            
            # Create a separate tracking for partial task hours per day
            # {user: {date: {task_id: hours}}}
            task_hours_per_day = {}
            for user in all_users:
                task_hours_per_day[user] = {}
            
            # Track total hours per user for reporting
            user_total_hours = {user: 0.0 for user in all_users}
            
            for task, delivery_date in tasks_with_dates:
                task_hours = float(task.get("qtd_mhr", 0) or 0)
                if task_hours <= 0:
                    logging.debug(f"Skipping task {task.get('id')} - no hours")
                    continue
                
                # Use assigned user, or first user as fallback
                assigned_user = task.get("assigned_to", "")
                if not assigned_user or assigned_user not in all_users:
                    assigned_user = all_users[0]
                    logging.debug(f"Task {task.get('id')} assigned to default user {assigned_user}")
                else:
                    logging.debug(f"Task {task.get('id')} keeping assigned user {assigned_user}")
                
                # If delivery date is in the past, start scheduling from today
                start_date = today
                
                # Find available slots for this user, splitting task across multiple days
                hours_remaining = task_hours
                current_date = start_date
                days_searched = 0
                max_days = 60  # Increased from 30 to ensure all tasks can be scheduled
                task_scheduled = False
                
                while hours_remaining > 0 and days_searched < max_days:
                    # Skip holidays/non-working days
                    if current_date in self.holidays:
                        current_date += datetime.timedelta(days=1)
                        days_searched += 1
                        continue
                    
                    # Initialize day tracking if needed
                    if current_date not in task_hours_per_day[assigned_user]:
                        task_hours_per_day[assigned_user][current_date] = {}
                    
                    # Calculate current workload for this user on this date
                    current_hours = sum(task_hours_per_day[assigned_user][current_date].values())
                    
                    # How much capacity is left? (Max 8 hours per day)
                    capacity_left = 8.0 - current_hours
                    
                    if capacity_left > 0:
                        # Assign what we can to this day (up to 8h total)
                        hours_to_assign = min(capacity_left, hours_remaining)
                        
                        # Track the hours for this task on this day
                        task_hours_per_day[assigned_user][current_date][task["id"]] = hours_to_assign
                        
                        # Add task to this day's schedule
                        if current_date not in self.schedule[assigned_user]:
                            self.schedule[assigned_user][current_date] = []
                        
                        if task["id"] not in self.schedule[assigned_user][current_date]:
                            self.schedule[assigned_user][current_date].append(task["id"])
                            if not task_scheduled:
                                tasks_assigned.add(task["id"])
                                distributed_count += 1
                                task_scheduled = True
                        
                        # Update user's total workload
                        user_total_hours[assigned_user] += hours_to_assign
                        
                        logging.debug(f"Assigned {hours_to_assign:.1f}h of task {task.get('id')} to {assigned_user} on {current_date} (day total: {current_hours + hours_to_assign:.1f}h, user total: {user_total_hours[assigned_user]:.1f}h)")
                        
                        hours_remaining -= hours_to_assign
                    
                    current_date += datetime.timedelta(days=1)
                    days_searched += 1
                
                if hours_remaining > 0:
                    logging.warning(f"Could not fully schedule task {task.get('id')}, {hours_remaining:.1f}h remaining (searched {days_searched} days)")
            
            # Store the detailed hours tracking for display purposes
            self.task_hours_per_day = task_hours_per_day
            
            logging.info(f"Distributed {distributed_count} task assignments")
            
            self.proposed_changes = True
            self._render_grid()
            
            holiday_count = len(self.holidays)
            
            # Build workload summary
            workload_summary = "Workload Distribution:\n"
            for user in all_users:
                total_h = user_total_hours[user]
                workload_summary += f"• {user}: {total_h:.1f}h\n"
            
            date_warning = ""
            if tasks_without_dates > 0:
                date_warning = f"\n⚠️ Note: {tasks_without_dates} tasks had no Target Finish date.\nThey were scheduled with a 7-day default deadline.\nPlease set proper deadlines for better scheduling.\n"
            
            messagebox.showinfo(
                "Schedule Generated",
                f"Successfully distributed {distributed_count} tasks across {len(all_users)} users!\n\n"
                f"{workload_summary}\n"
                f"Non-working days: {holiday_count}"
                f"{date_warning}\n"
                "This is a visualization tool. Make changes in the main task list.\n\n"
                "Legend:\n"
                "- Gray = Sunday (auto non-working)\n"
                "- Blue = User-marked non-working day\n"
                "- Green = Normal (≤8h)\n"
                "- Orange = Slightly over (8-10h)\n"
                "- Red = Overloaded (>10h)"
            )
            
        except Exception as e:
            logging.error(f"Error distributing tasks: {e}", exc_info=True)
            messagebox.showerror("Error", f"Failed to distribute tasks:\n\n{str(e)}\n\nCheck task_manager.log for details.")
    
    def _toggle_holiday(self, date):
        """Toggle a date between working day and holiday."""
        # Prevent toggling Sundays - they're always non-working days
        if date.weekday() == 6:
            return
        
        if date in self.holidays:
            self.holidays.remove(date)
            status = "working day"
        else:
            self.holidays.add(date)
            status = "non-working day"
        
        self._render_grid()
        # Optional: Show brief notification
        # messagebox.showinfo("Day Updated", f"{date.strftime('%Y-%m-%d')} set as {status}")
    
    def _ensure_schedule_tables_exist(self):
        """Check if database tables for saving schedules exist."""
        try:
            cursor = self.task_manager.cursor
            
            # Check if ScheduledTasks table exists
            cursor.execute("""
                SELECT COUNT(*) 
                FROM INFORMATION_SCHEMA.TABLES 
                WHERE TABLE_NAME = 'ScheduledTasks'
            """)
            tasks_table_exists = cursor.fetchone()[0] > 0
            
            # Check if ScheduledHolidays table exists
            cursor.execute("""
                SELECT COUNT(*) 
                FROM INFORMATION_SCHEMA.TABLES 
                WHERE TABLE_NAME = 'ScheduledHolidays'
            """)
            holidays_table_exists = cursor.fetchone()[0] > 0
            
            if tasks_table_exists and holidays_table_exists:
                self.save_enabled = True
                logging.info("Scheduler database save functionality enabled")
            else:
                self.save_enabled = False
                logging.info("Scheduler will use file-based save (database tables not available)")
            
        except Exception as e:
            self.save_enabled = False
            logging.error(f"Error checking schedule tables: {e}", exc_info=True)
    
    def _load_saved_schedule(self):
        """Load saved schedule from file or database."""
        # Try loading from file first (always works)
        if self._load_from_file():
            return
        
        # If no file, try database if available
        if self.save_enabled:
            if self._load_from_database():
                return
        
        # No saved schedule found, analyze current
        logging.info("No saved schedule found, analyzing current tasks")
        self._analyze_current_schedule()
    
    def _load_from_file(self):
        """Load schedule from JSON file."""
        try:
            schedule_file = os.path.join(os.path.dirname(__file__), 'saved_schedule.json')
            
            if not os.path.exists(schedule_file):
                return False
            
            with open(schedule_file, 'r') as f:
                data = json.load(f)
            
            all_users = list(USER_MAPPING.keys())
            self.schedule = {user: {} for user in all_users}
            self.task_hours_per_day = {user: {} for user in all_users}
            
            # Load schedule data
            for user, dates in data.get('schedule', {}).items():
                if user not in all_users:
                    continue
                for date_str, task_ids in dates.items():
                    date = datetime.datetime.strptime(date_str, '%Y-%m-%d').date()
                    self.schedule[user][date] = task_ids
            
            # Load task hours per day
            for user, dates in data.get('task_hours_per_day', {}).items():
                if user not in all_users:
                    continue
                for date_str, tasks in dates.items():
                    date = datetime.datetime.strptime(date_str, '%Y-%m-%d').date()
                    self.task_hours_per_day[user][date] = {int(k): v for k, v in tasks.items()}
            
            # Load holidays
            for date_str in data.get('holidays', []):
                date = datetime.datetime.strptime(date_str, '%Y-%m-%d').date()
                if date.weekday() != 6:  # Don't load Sundays (auto-marked)
                    self.holidays.add(date)
            
            logging.info(f"Loaded schedule from file: {schedule_file}")
            
            # Update info label with task count (exclude completed and deleted)
            pending_tasks = [t for t in self.task_manager.tasks if not t.get("completed", False) and not t.get("deleted", False)]
            all_users = list(USER_MAPPING.keys())
            self.info_label.config(
                text=f"📊 {len(all_users)} users | {len(pending_tasks)} pending tasks"
            )
            
            self._render_grid()
            return True
            
        except Exception as e:
            logging.error(f"Error loading schedule from file: {e}", exc_info=True)
            return False
    
    def _load_from_database(self):
        """Load schedule from database."""
        try:
            cursor = self.task_manager.cursor
            
            # Check if we have any saved schedule
            cursor.execute("SELECT COUNT(*) FROM ScheduledTasks")
            count = cursor.fetchone()[0]
            
            if count == 0:
                # No saved schedule, analyze current
                self._analyze_current_schedule()
                return
            
            # Load saved schedule
            all_users = list(USER_MAPPING.keys())
            self.schedule = {user: {} for user in all_users}
            self.task_hours_per_day = {user: {} for user in all_users}
            
            cursor.execute("""
                SELECT user_name, scheduled_date, task_id, hours_allocated
                FROM ScheduledTasks
                ORDER BY user_name, scheduled_date
            """)
            
            for row in cursor.fetchall():
                user = row[0]
                date = row[1]
                task_id = row[2]
                hours = float(row[3])
                
                if user not in all_users:
                    continue
                
                # Add to schedule
                if date not in self.schedule[user]:
                    self.schedule[user][date] = []
                if task_id not in self.schedule[user][date]:
                    self.schedule[user][date].append(task_id)
                
                # Add to task_hours_per_day
                if date not in self.task_hours_per_day[user]:
                    self.task_hours_per_day[user][date] = {}
                self.task_hours_per_day[user][date][task_id] = hours
            
            # Load saved holidays (not Sundays, those are auto-marked)
            cursor.execute("SELECT holiday_date FROM ScheduledHolidays")
            for row in cursor.fetchall():
                self.holidays.add(row[0])
            
            logging.info(f"Loaded saved schedule from database with {count} entries")
            
            # Update info label with task count (exclude completed and deleted)
            pending_tasks = [t for t in self.task_manager.tasks if not t.get("completed", False) and not t.get("deleted", False)]
            all_users = list(USER_MAPPING.keys())
            self.info_label.config(
                text=f"📊 {len(all_users)} users | {len(pending_tasks)} pending tasks"
            )
            
            self._render_grid()
            return True
            
        except Exception as e:
            logging.error(f"Error loading saved schedule from database: {e}", exc_info=True)
            return False
    
    def _save_schedule(self):
        """Save schedule to database if available, otherwise to file."""
        # Disable button to prevent multiple clicks
        self.save_button.config(state=tk.DISABLED)
        
        try:
            if self.save_enabled:
                # Try database first
                if self._save_to_database():
                    return
            
            # Fall back to file save
            self._save_to_file()
        finally:
            # Re-enable button after save completes (if dialog still exists)
            try:
                if self.save_button.winfo_exists():
                    self.save_button.config(state=tk.NORMAL)
            except (tk.TclError, AttributeError) as e:
                logging.debug(f"Button re-enable error (dialog closed): {e}")
    
    def _save_to_file(self):
        """Save schedule to JSON file."""
        try:
            schedule_file = os.path.join(os.path.dirname(__file__), 'saved_schedule.json')
            
            # Convert schedule data to serializable format
            data = {
                'schedule': {},
                'task_hours_per_day': {},
                'holidays': [],
                'saved_at': datetime.datetime.now().isoformat()
            }
            
            # Save schedule
            for user, dates in self.schedule.items():
                data['schedule'][user] = {}
                for date, task_ids in dates.items():
                    data['schedule'][user][date.strftime('%Y-%m-%d')] = task_ids
            
            # Save task hours per day
            for user, dates in self.task_hours_per_day.items():
                data['task_hours_per_day'][user] = {}
                for date, tasks in dates.items():
                    data['task_hours_per_day'][user][date.strftime('%Y-%m-%d')] = tasks
            
            # Save holidays (excluding Sundays)
            for holiday_date in self.holidays:
                if holiday_date.weekday() != 6:
                    data['holidays'].append(holiday_date.strftime('%Y-%m-%d'))
            
            # Write to file
            with open(schedule_file, 'w') as f:
                json.dump(data, f, indent=2)
            
            # Count entries
            save_count = sum(len(tasks) for dates in self.task_hours_per_day.values() for tasks in dates.values())
            holiday_count = len(data['holidays'])
            
            logging.info(f"Saved schedule to file: {schedule_file}")
            
            messagebox.showinfo(
                "Schedule Saved",
                f"Schedule saved successfully to file!\n\n"
                f"• {save_count} task assignments\n"
                f"• {holiday_count} marked holidays\n\n"
                f"File: saved_schedule.json\n"
                "This schedule will load automatically next time."
            )
            
        except Exception as e:
            logging.error(f"Error saving schedule to file: {e}", exc_info=True)
            messagebox.showerror("Save Error", f"Failed to save schedule:\n{str(e)}")
    
    def _save_to_database(self):
        """Save schedule to database. Returns True if successful."""
        try:
            cursor = self.task_manager.cursor
            
            # Clear existing saved schedule
            cursor.execute("DELETE FROM ScheduledTasks")
            cursor.execute("DELETE FROM ScheduledHolidays")
            
            # Save schedule
            save_count = 0
            for user, dates in self.task_hours_per_day.items():
                for date, tasks in dates.items():
                    for task_id, hours in tasks.items():
                        cursor.execute("""
                            INSERT INTO ScheduledTasks (user_name, scheduled_date, task_id, hours_allocated)
                            VALUES (%s, %s, %s, %s)
                        """, (user, date, task_id, hours))
                        save_count += 1
            
            # Save holidays (excluding Sundays since they're auto-marked)
            holiday_count = 0
            for holiday_date in self.holidays:
                if holiday_date.weekday() != 6:  # Don't save Sundays
                    cursor.execute("""
                        INSERT INTO ScheduledHolidays (holiday_date)
                        VALUES (%s)
                    """, (holiday_date,))
                    holiday_count += 1
            
            logging.info(f"Saved schedule: {save_count} task entries, {holiday_count} holidays")
            
            messagebox.showinfo(
                "Schedule Saved",
                f"Schedule saved successfully to database!\n\n"
                f"• {save_count} task assignments\n"
                f"• {holiday_count} marked holidays\n\n"
                "This schedule will load automatically next time."
            )
            
            return True
            
        except Exception as e:
            logging.error(f"Error saving schedule to database: {e}", exc_info=True)
            return False
    
    def _auto_mark_all_sundays(self):
        """Automatically mark all Sundays as non-working days (silent, no popup)."""
        today = datetime.date.today()
        
        # Mark Sundays for the entire year
        for day_offset in range(365):  # Full year
            date = today + datetime.timedelta(days=day_offset)
            if date.weekday() == 6:  # Sunday
                self.holidays.add(date)
    
    def _prev_month(self):
        """Navigate to previous month."""
        # Move back by 21 days (3 weeks)
        self.start_date = self.start_date - datetime.timedelta(days=21)
        self._render_grid()
    
    def _next_month(self):
        """Navigate to next month."""
        # Move forward by 21 days (3 weeks)
        self.start_date = self.start_date + datetime.timedelta(days=21)
        self._render_grid()
    
    def _go_to_today(self):
        """Navigate back to today."""
        self.start_date = datetime.date.today()
        self._render_grid()
    
    def _apply_schedule(self):
        """Apply the proposed schedule to tasks."""
        if not self.proposed_changes:
            messagebox.showinfo("No Changes", "No schedule changes to apply.")
            return
        
        try:
            if not messagebox.askyesno(
                "Confirm Changes",
                "Apply the proposed schedule?\n\n"
                "This will update task assignments and dates in the database."
            ):
                return
            
            # Note: Currently we only reassign to dates, not split tasks
            # A more advanced version would need to track partial task assignments
            
            changes_made = 0
            for user, dates in self.schedule.items():
                for date, task_ids in dates.items():
                    for task_id in task_ids:
                        task = self.task_manager._find_task_by_id(task_id)
                        if task:
                            # Update assigned user and target dates if changed
                            if task.get("assigned_to") != user:
                                self.task_manager.update_task(task_id, assigned_to=user)
                                changes_made += 1
            
            messagebox.showinfo(
                "Success",
                f"Applied schedule changes!\n\n"
                f"{changes_made} task assignments updated."
            )
            
            if self.callback:
                self.callback()
            
            self.dialog.destroy()
            
        except Exception as e:
            logging.error(f"Error applying schedule: {e}", exc_info=True)
            messagebox.showerror("Error", f"Failed to apply schedule:\n{str(e)}")


class TaskManager:
    """Class to manage daily tasks."""
    
    def __init__(self):
        """Initialize the task manager."""
        # Lazy import: only load pymssql when TaskManager is instantiated
        global pymssql
        import pymssql
        
        # Database configuration loaded from config.ini
        # server and database will be set in _connect() from app_config
        self.server = None
        self.database = None
        self.user = None  # Will be set by _get_user_credentials
        self.password = None  # Will be set by _get_user_credentials
        
        # Initialize internal state
        self._tasks = []  # List of tasks loaded from the database
        self._deleted_tasks = []  # Track recently deleted tasks for undo (used by older undo logic)
        # self._transaction_active = False # REMOVED - Not needed with autocommit=True
        
        # Task loading configuration - Load all tasks (instant display makes limit unnecessary)
        self.all_tasks_loaded = True  # Always load all tasks
        self.total_task_count = 0  # Total tasks in database
        
        # Deleted tasks stack for undo functionality
        self.deleted_tasks_stack = []  # Stack of deleted task batches
        
        # Create standardized connection references
        self.conn = None  # Main connection object
        self.cursor = None  # Main cursor object
        
        # Initialize everything
        self._get_user_credentials()  # Set up credentials (sets self.user and self.password)
        self._connect()  # Connect to database (sets self.server and self.database from config)
        if self.conn: # Only load if connection succeeded
            self.total_task_count = self._get_total_task_count()  # Get total count first
            self._tasks = self._load_tasks(limit=None)  # Load all tasks
            self._ensure_deleted_field_exists()
            self._ensure_users_table_exists()  # Ensure Users table exists
            self._ensure_sessions_table_exists()  # Ensure UserSessions table exists
            self._register_user_session()  # Register this user session
            self._ensure_sessions_table_exists()  # Ensure UserSessions table exists
            self._register_user_session()  # Register this user session
            
            # Initialize enhanced feature managers
            if ENHANCED_FEATURES_AVAILABLE:
                try:
                    self.time_tracking = TimeTrackingManager(self)
                    self.subtasks = SubtaskManager(self)
                    self.templates = TemplateManager(self)
                    self.search = AdvancedSearchManager(self)
                    self.dashboard = DashboardManager(self)
                    logging.info("Enhanced features initialized successfully")
                except Exception as e:
                    logging.error(f"Error initializing enhanced features: {e}", exc_info=True)
                    # Set to None so we can check if they're available
                    self.time_tracking = None
                    self.subtasks = None
                    self.templates = None
                    self.search = None
                    self.dashboard = None
            else:
                self.time_tracking = None
                self.subtasks = None
                self.templates = None
                self.search = None
                self.dashboard = None
        else:
            # Handle case where initial connection failed
             self._tasks = []
             logging.error("Initial database connection failed. Task list will be empty.")
             # Consider raising an error or exiting if connection is critical
    
    def _ensure_deleted_field_exists(self):
        """Ensure the 'deleted' field exists in the Tasks table."""
        if not self.conn or not self.cursor: return # Cannot proceed without connection
        try:
            # Check if the deleted column exists
            self.cursor.execute("""
                SELECT COUNT(*)
                FROM INFORMATION_SCHEMA.COLUMNS
                WHERE TABLE_NAME = 'Tasks'
                AND COLUMN_NAME = 'deleted'
            """)
            column_exists = self.cursor.fetchone()[0] > 0

            if not column_exists:
                # Add the column if it doesn't exist
                logging.info("Attempting to add 'deleted' column...")
                self.cursor.execute("""
                    ALTER TABLE Tasks
                    ADD deleted BIT NOT NULL DEFAULT 0
                """)
                # No explicit commit needed due to autocommit=True
                logging.info("Added 'deleted' column to Tasks table (autocommitted).")
        except pymssql.Error as e:
            logging.error("Error checking/adding deleted column", exc_info=True)
            # Rollback might not be strictly necessary with autocommit but good practice if error occurs
            try: self.conn.rollback()
            except Exception: pass
            messagebox.showerror("Schema Update Error", f"Could not ensure 'deleted' column exists: {e}")
    
    def _ensure_users_table_exists(self):
        """Ensure the Users table exists for storing user mappings."""
        if not self.conn or not self.cursor: return
        try:
            # Check if Users table exists
            self.cursor.execute("""
                SELECT COUNT(*)
                FROM INFORMATION_SCHEMA.TABLES
                WHERE TABLE_NAME = 'Users'
            """)
            table_exists = self.cursor.fetchone()[0] > 0
            
            if not table_exists:
                logging.info("Creating Users table...")
                self.cursor.execute("""
                    CREATE TABLE Users (
                        id INT IDENTITY(1,1) PRIMARY KEY,
                        display_name NVARCHAR(100) NOT NULL UNIQUE,
                        tcn NVARCHAR(20) NOT NULL UNIQUE,
                        sql_username NVARCHAR(50) NOT NULL UNIQUE,
                        created_date DATETIME NOT NULL DEFAULT GETDATE(),
                        active BIT NOT NULL DEFAULT 1
                    )
                """)
                logging.info("Created Users table successfully.")
                
                # Insert default users from USER_MAPPING
                inserted_count = 0
                for display_name, data in USER_MAPPING.items():
                    tcn = data.get('tcn', '')
                    sql_user = data.get('sql_user', '')
                    if tcn and sql_user:
                        try:
                            self.cursor.execute("""
                                INSERT INTO Users (display_name, tcn, sql_username)
                                VALUES (%s, %s, %s)
                            """, (display_name, tcn, sql_user))
                            inserted_count += 1
                        except pymssql.Error as insert_err:
                            # Log but don't fail - might be permission issue or duplicate
                            logging.warning(f"Could not insert default user {display_name}: {insert_err}")
                
                if inserted_count > 0:
                    logging.info(f"Inserted {inserted_count} default users into Users table.")
                else:
                    logging.warning("Could not insert default users - may need DBA to grant INSERT permission on Users table.")
        except pymssql.Error as e:
            logging.error("Error creating Users table", exc_info=True)
    
    def _ensure_sessions_table_exists(self):
        """Ensure the UserSessions table exists for tracking active sessions."""
        if not self.conn or not self.cursor: return
        
        self.sessions_table_available = False  # Track if table is available
        
        try:
            # Check if UserSessions table exists
            self.cursor.execute("""
                SELECT COUNT(*)
                FROM INFORMATION_SCHEMA.TABLES
                WHERE TABLE_NAME = 'UserSessions'
            """)
            table_exists = self.cursor.fetchone()[0] > 0
            
            if table_exists:
                self.sessions_table_available = True
                logging.info("UserSessions table found and ready.")
                return
            
            # Try to create table (will fail if no permissions)
            logging.info("Creating UserSessions table...")
            self.cursor.execute("""
                CREATE TABLE UserSessions (
                    session_id INT IDENTITY(1,1) PRIMARY KEY,
                    username NVARCHAR(100) NOT NULL,
                    login_time DATETIME NOT NULL DEFAULT GETDATE(),
                    last_heartbeat DATETIME NOT NULL DEFAULT GETDATE(),
                    is_active BIT NOT NULL DEFAULT 1,
                    machine_name NVARCHAR(100),
                    INDEX idx_username_active (username, is_active, last_heartbeat)
                )
            """)
            self.sessions_table_available = True
            logging.info("UserSessions table created successfully.")
            
        except pymssql.Error as e:
            # Permission denied or other error
            error_msg = str(e)
            if "permission denied" in error_msg.lower():
                logging.warning("UserSessions table does not exist and cannot be created due to permissions.")
                logging.warning("Session tracking (Online/Offline status) will not be available.")
                logging.warning("Please ask your DBA to run 'create_sessions_table.sql' to enable this feature.")
            else:
                logging.error(f"Error with UserSessions table: {e}", exc_info=True)
            self.sessions_table_available = False
    
    def _register_user_session(self):
        """Register the current user session in the database."""
        if not self.conn or not self.cursor: return
        
        # Check if sessions table is available
        if not hasattr(self, 'sessions_table_available') or not self.sessions_table_available:
            logging.debug("Session tracking not available - UserSessions table does not exist.")
            return
            
        try:
            import socket
            machine_name = socket.gethostname()
            
            # Get current Windows username
            current_user = getpass.getuser()
            
            # Insert new session
            self.cursor.execute("""
                INSERT INTO UserSessions (username, machine_name, login_time, last_heartbeat, is_active)
                VALUES (%s, %s, GETDATE(), GETDATE(), 1)
            """, (current_user, machine_name))
            
            # Store session ID for later cleanup
            self.cursor.execute("SELECT @@IDENTITY")
            self.session_id = self.cursor.fetchone()[0]
            
            logging.info(f"User session registered: {current_user} on {machine_name} (Session ID: {self.session_id})")
            
            # Start heartbeat thread to keep session alive
            self._start_heartbeat_thread()
            
        except Exception as e:
            logging.warning(f"Could not register user session: {e}")
            self.sessions_table_available = False
    
    def _start_heartbeat_thread(self):
        """Start a background thread to update session heartbeat."""
        if not hasattr(self, 'sessions_table_available') or not self.sessions_table_available:
            return
            
        import time
        
        def heartbeat():
            while True:
                try:
                    time.sleep(60)  # Update every minute
                    if self.conn and self.cursor and hasattr(self, 'session_id'):
                        self.cursor.execute("""
                            UPDATE UserSessions
                            SET last_heartbeat = GETDATE()
                            WHERE session_id = %s AND is_active = 1
                        """, (self.session_id,))
                except Exception as e:
                    logging.debug(f"Heartbeat stopped: {e}")
                    break
        
        heartbeat_thread = threading.Thread(target=heartbeat, daemon=True)
        heartbeat_thread.start()
    
    def close_user_session(self):
        """Mark the current user session as inactive."""
        if not self.conn or not self.cursor or not hasattr(self, 'session_id'): return
        if not hasattr(self, 'sessions_table_available') or not self.sessions_table_available: return
        
        try:
            self.cursor.execute("""
                UPDATE UserSessions
                SET is_active = 0
                WHERE session_id = %s
            """, (self.session_id,))
            self.conn.commit()  # Commit the transaction
            logging.info(f"User session closed (Session ID: {self.session_id})")
        except Exception as e:
            logging.debug(f"Could not close user session: {e}")
    
    def _get_user_credentials(self):
        """Retrieves SQL Server credentials based on the current user.
        
        SECURITY WARNING: Passwords are currently hardcoded.
        For production, use Windows Credential Manager:
          1. Store password: cmdkey /generic:TaskManager_DB_<username> /user:<sql_user> /pass:<password>
          2. Enable in config.ini: use_windows_credential_manager = true
        """
        username = getpass.getuser()  # Get the current logged-in username
        
        # Use shared database credentials for all application users
        # All users connect with the same database account
        self.user = "cdx_clam"
        logging.info(f"Application user '{username}' will use shared database user: {self.user}")
        
        # Check if using Windows Credential Manager (not recommended with shared credentials)
        use_credential_manager = app_config.getboolean('Credentials', 'use_windows_credential_manager', fallback=False)
        
        # Get password from Windows Credential Manager or use configured password
        if use_credential_manager:
            try:
                import keyring
                password = keyring.get_password("TaskManager_DB", self.user)
                if password:
                    self.password = password
                    logging.info(f"Retrieved password from Windows Credential Manager for {self.user}")
                else:
                    logging.error(f"Password not found in Credential Manager for {self.user}. Falling back to configured password.")
                    self.password = "Clam@access"
            except ImportError:
                logging.error("keyring library not installed. Install with: pip install keyring")
                self.password = "Clam@access"
            except Exception as e:
                logging.error(f"Error retrieving password from Credential Manager: {e}")
                self.password = "Clam@access"
        else:
            # Using shared database password for all users
            self.password = "Clam@access"
            logging.info("Using configured shared database password.")
    
    def _connect(self):
        """Connect to the SQL Server database with autocommit=True and retry logic."""
        # Get configuration
        self.server = app_config.get('Database', 'server', fallback='10.195.96.58:1433')
        self.database = app_config.get('Database', 'database', fallback='TaskManager1')
        timeout = app_config.getint('Database', 'timeout', fallback=10)
        enable_retry = app_config.getboolean('Database', 'enable_retry', fallback=True)
        max_attempts = app_config.getint('Database', 'max_retry_attempts', fallback=3)
        retry_delay = app_config.getint('Database', 'retry_delay_seconds', fallback=2)
        
        # Close existing connection if any
        if self.conn:
            try:
                self.conn.close()
            except Exception as close_e:
                logging.warning(f"Ignoring error while closing previous connection: {close_e}")
            self.conn = None
            self.cursor = None
        
        # Attempt connection with retry logic
        last_error = None
        for attempt in range(1, max_attempts + 1 if enable_retry else 2):
            try:
                logging.info(f"Database connection attempt {attempt}/{max_attempts if enable_retry else 1}...")
                logging.info(f"Connecting to server: {self.server}, database: {self.database}, user: {self.user}")
                
                # Connect to the database with autocommit enabled
                self.conn = pymssql.connect(
                    server=self.server,
                    database=self.database,
                    user=self.user,
                    password=self.password,
                    timeout=timeout,
                    autocommit=True  # SET AUTOCOMMIT TO TRUE
                )

                # Create a cursor
                self.cursor = self.conn.cursor()

                logging.info(f"Successfully connected to SQL Server (autocommit=True) on attempt {attempt}")
                return  # Success - exit function

            except pymssql.Error as e:
                last_error = e
                error_code = e.args[0] if e.args else None
                error_msg = str(e)
                
                # Check for specific error codes
                if error_code == 18456:
                    logging.error(f"Authentication failed for user '{self.user}' (Error 18456)")
                    logging.error(f"Possible causes:")
                    logging.error(f"  1. Username or password is incorrect")
                    logging.error(f"  2. User '{self.user}' doesn't exist on server {self.server}")
                    logging.error(f"  3. User doesn't have permission to access database '{self.database}'")
                else:
                    logging.error(f"Database connection attempt {attempt} failed: {error_msg}")
                
                if attempt < max_attempts and enable_retry and error_code != 18456:
                    # Don't retry authentication errors - they won't succeed
                    # Exponential backoff: 2s, 4s, 8s, etc.
                    delay = retry_delay * (2 ** (attempt - 1))
                    logging.warning(f"Retrying in {delay} seconds...")
                    time.sleep(delay)
                else:
                    # Final attempt failed or authentication error
                    self.conn = None
                    self.cursor = None
                    
                    if error_code == 18456:
                        logging.error(f"Authentication failed: Please verify credentials for user '{self.user}'")
                        raise ConnectionError(
                            f"Authentication failed for user '{self.user}' on {self.server}.\n\n"
                            f"Error 18456: Login failed.\n\n"
                            f"Please verify:\n"
                            f"1. User '{self.user}' exists on SQL Server {self.server}\n"
                            f"2. Password is correct\n"
                            f"3. User has permission to access database '{self.database}'\n"
                            f"4. SQL Server allows SQL Server authentication (not just Windows Auth)"
                        ) from e
                    else:
                        logging.error(f"All {max_attempts} connection attempts failed")
                        raise ConnectionError(f"Database connection failed after {max_attempts} attempts: {e}") from e
                    
            except Exception as e:
                last_error = e
                logging.error(f"Unexpected error during connection attempt {attempt}: {e}", exc_info=True)
                self.conn = None
                self.cursor = None
                if attempt >= max_attempts or not enable_retry:
                    raise ConnectionError(f"Unexpected connection error: {e}") from e
                else:
                    delay = retry_delay * (2 ** (attempt - 1))
                    logging.warning(f"Retrying in {delay} seconds...")
                    time.sleep(delay)
    
    def _get_total_task_count(self) -> int:
        """Get total count of non-deleted tasks in database."""
        if not self.conn or not self.cursor:
            return 0
        try:
            self.cursor.execute("SELECT COUNT(*) FROM [dbo].[Tasks] WHERE deleted = 0")
            count = self.cursor.fetchone()[0]
            logging.info(f"Total non-deleted tasks in database: {count}")
            return count
        except Exception as e:
            logging.error(f"Error getting task count: {e}")
            return 0
    
    def _load_tasks(self, limit: int = None) -> List[Dict[str, Any]]:
        """Load tasks from the SQL Server database.
        
        Args:
            limit: Maximum number of tasks to load (most recent first). None = load all.
        """
        if not self.conn or not self.cursor:
             logging.error("Cannot load tasks: No database connection.")
             return []
        try:
            # Build query with optional limit
            # Note: We load ALL tasks (including deleted ones) so that deleted tasks can be viewed/recovered
            # Filtering of deleted tasks is handled by get_filtered_tasks() based on show_deleted parameter
            if limit:
                # Load most recent tasks first (by created_date or id)
                query = f"SELECT TOP {limit} * FROM [dbo].[Tasks] ORDER BY created_date DESC, id DESC"
                self.all_tasks_loaded = False
                logging.info(f"Loading {limit} most recent tasks (including deleted)...")
            else:
                # Load all tasks
                query = "SELECT * FROM [dbo].[Tasks] ORDER BY created_date DESC, id DESC"
                self.all_tasks_loaded = True
                logging.info("Loading all tasks (including deleted)...")
            
            self.cursor.execute(query)
            rows = self.cursor.fetchall()
            column_names = [desc[0] for desc in self.cursor.description]

            # Convert rows to a list of dictionaries using column names
            tasks = []
            for row in rows:
                task = dict(zip(column_names, row))

                # --- Data Type Conversions and Formatting ---
                # Convert dates to string YYYY-MM-DD or None
                for date_key in ['created_date', 'due_date', 'requested_date']:
                    if date_key in task and isinstance(task[date_key], datetime.datetime):
                        task[date_key] = task[date_key].strftime("%Y-%m-%d")
                    elif date_key in task and isinstance(task[date_key], datetime.date):
                         task[date_key] = task[date_key].strftime("%Y-%m-%d")
                    elif date_key in task and not isinstance(task[date_key], str):
                         task[date_key] = None # Handle potential non-date, non-string values

                # Handle target dates separately to ensure empty strings for None
                for date_key in ['target_start', 'target_finish']:
                    if date_key in task and isinstance(task[date_key], (datetime.datetime, datetime.date)):
                        task[date_key] = task[date_key].strftime("%Y-%m-%d")
                    else:
                        task[date_key] = "" # Ensure it's an empty string if None or invalid

                # Format last_modified
                if 'last_modified' in task and isinstance(task['last_modified'], datetime.datetime):
                    task['last_modified'] = task['last_modified'].strftime("%Y-%m-%d %H:%M:%S")
                elif 'last_modified' in task:
                     task['last_modified'] = None # Or keep as is if needed

                # Ensure boolean fields are bool
                if 'completed' in task:
                    task['completed'] = bool(task['completed'])
                if 'deleted' in task:
                    task['deleted'] = bool(task['deleted'])
                else: # Handle case where 'deleted' column might be missing initially
                    task['deleted'] = False

                # Ensure numeric fields are numbers (or None/0 if appropriate)
                for num_key in ['qtd_mhr', 'actual_mhr']:
                     if num_key in task and task[num_key] is not None:
                         try:
                             task[num_key] = int(task[num_key])
                         except (ValueError, TypeError):
                              task[num_key] = 0 # Or None, depending on desired default
                     else:
                          task[num_key] = 0 # Default if null

                tasks.append(task)

            logging.info(f"Loaded {len(tasks)} tasks from the database.")
            return tasks
        except pymssql.Error as e:
            logging.error("Error loading tasks from database", exc_info=True)
            
            # Check if it's a permission error
            error_msg = str(e)
            if "permission was denied" in error_msg.lower() or "229" in str(e):
                # Permission error - provide helpful message with option to copy script
                result = messagebox.askyesno(
                    "Database Permission Required",
                    "Your database user doesn't have permission to access the Tasks table.\n\n"
                    "A DBA needs to run the permission script to grant access.\n\n"
                    "The script is in: grant_permissions.sql\n\n"
                    "Would you like to copy the SQL script to clipboard?",
                    icon=messagebox.WARNING
                )
                if result:
                    # Copy script to clipboard
                    script = self.generate_permission_script()
                    try:
                        import tkinter as tk
                        temp_root = tk.Tk()
                        temp_root.withdraw()
                        temp_root.clipboard_clear()
                        temp_root.clipboard_append(script)
                        temp_root.update()
                        temp_root.destroy()
                        messagebox.showinfo("Script Copied", 
                            "SQL permission script copied to clipboard!\n\n"
                            "Send this to your DBA to run in SQL Server Management Studio.")
                    except Exception:
                        messagebox.showerror("Error", "Could not copy to clipboard. Check grant_permissions.sql file instead.")
            else:
                # Other database error
                messagebox.showerror("Database Error", f"Error loading tasks: {str(e)}")
            
            return []
        except Exception as e:
            logging.error(f"Unexpected error loading tasks: {str(e)}", exc_info=True)
            messagebox.showerror("Application Error", f"Unexpected error loading tasks: {str(e)}")
            return []
    
    def _get_current_user(self):
        """Get the current user's username."""
        try:
            return os.getenv('USERNAME') or os.getenv('USER') or 'Unknown User'
        except (OSError, KeyError) as e:
            logging.warning(f"Error getting current user: {e}")
            return 'Unknown User'
    
    def add_task(self, title: str, description: str = "", due_date: Optional[str] = None, 
                priority: str = "medium", category: str = "general", 
                main_staff: Optional[str] = None, assigned_to: Optional[str] = None,
                applied_vessel: str = "", rev: str = "", drawing_no: str = "",
                link: str = "", sdb_link: str = "", request_no: str = "",
                requested_date: Optional[str] = None,
                target_start: Optional[str] = None, target_finish: Optional[str] = None,
                qtd_mhr: int = 0, actual_mhr: int = 0) -> Optional[int]: # Return new ID or None on failure
        """
        Add a new task to the SQL Server database (autocommitted).
        Returns the new task ID on success, None on failure.
        """
        if not self.conn or not self.cursor:
             logging.error("Cannot add task: No database connection.")
             messagebox.showerror("Database Error", "No database connection available.")
             return None

        try:
            # Get current date and time for metadata columns
            # Ensure datetime objects are used for timestamp columns if DB expects them
            current_datetime = datetime.datetime.now()
            # Use datetime object directly if column type is DATETIME/DATETIME2
            created_date_dt = current_datetime
            last_modified_dt = current_datetime
            # Or format if column type is VARCHAR/NVARCHAR
            # created_date_str = current_datetime.strftime("%Y-%m-%d %H:%M:%S")
            # last_modified_str = current_datetime.strftime("%Y-%m-%d %H:%M:%S")

            created_by = self._get_current_user()
            modified_by = created_by

            # --- Input Validation and Formatting ---
            # Validate/Format dates (ensure they are 'YYYY-MM-DD' strings or None)
            def format_date(date_input):
                if isinstance(date_input, datetime.date):
                    return date_input.strftime("%Y-%m-%d")
                elif isinstance(date_input, str) and date_input:
                    try:
                        # Validate format YYYY-MM-DD
                        datetime.datetime.strptime(date_input, '%Y-%m-%d')
                        return date_input
                    except ValueError:
                        logging.warning(f"Invalid date format provided: {date_input}. Setting to None.")
                        return None
                return None # Return None for empty strings, None, or other invalid types

            due_date = format_date(due_date)
            requested_date = format_date(requested_date)
            target_start = format_date(target_start)
            target_finish = format_date(target_finish)

            # Ensure MHR values are integers
            try:
                qtd_mhr = int(qtd_mhr) if qtd_mhr is not None else 0
            except (ValueError, TypeError):
                 logging.warning(f"Invalid qtd_mhr value: {qtd_mhr}. Setting to 0.")
                 qtd_mhr = 0
            try:
                actual_mhr = int(actual_mhr) if actual_mhr is not None else 0
            except (ValueError, TypeError):
                 logging.warning(f"Invalid actual_mhr value: {actual_mhr}. Setting to 0.")
                 actual_mhr = 0


            # --- Database Interaction (autocommitted) ---
            sql = """
                INSERT INTO Tasks (
                    title, description, created_date, created_by, due_date, priority, category,
                    main_staff, assigned_to, completed, applied_vessel, rev, drawing_no,
                    link, sdb_link, request_no, requested_date,
                    target_start, target_finish,
                    qtd_mhr, actual_mhr, last_modified, modified_by, deleted
                ) OUTPUT INSERTED.id
                VALUES (
                    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s
                );
            """
            # Prepare parameters tuple - ensure order matches INSERT list
            params = (
                title, description, created_date_dt, created_by, due_date, priority.lower(), category,
                main_staff, assigned_to, False, applied_vessel, rev, drawing_no, # completed=False
                link, sdb_link, request_no, requested_date,
                target_start, target_finish,
                qtd_mhr, actual_mhr, last_modified_dt, modified_by, False # deleted=False
            )

            logging.debug(f"Executing INSERT with params: {params}")
            self.cursor.execute(sql, params)
            result = self.cursor.fetchone()

            if result is None:
                logging.error("INSERT statement did not return an ID via OUTPUT clause (autocommit might have failed implicitly).")
                # No explicit rollback needed for autocommit errors usually, DB handles it
                return None

            new_id = result[0]
            # No explicit commit needed here

            # --- Update Local Cache ---
            created_date_str = created_date_dt.strftime("%Y-%m-%d")
            last_modified_str = last_modified_dt.strftime("%Y-%m-%d %H:%M:%S")
            new_task = {
                "id": new_id,
                "title": title, "description": description, "created_date": created_date_str,
                "created_by": created_by, "due_date": due_date, "priority": priority, "category": category,
                "main_staff": main_staff, "assigned_to": assigned_to, "completed": False,
                "applied_vessel": applied_vessel, "rev": rev, "drawing_no": drawing_no,
                "link": link, "sdb_link": sdb_link, "request_no": request_no,
                "requested_date": requested_date,
                "target_start": target_start, "target_finish": target_finish,
                "qtd_mhr": qtd_mhr, "actual_mhr": actual_mhr,
                "last_modified": last_modified_str, "modified_by": modified_by, "deleted": False
            }
            self._tasks.append(new_task) # Add to local list

            logging.info(f"Task {new_id} added successfully by {created_by} (autocommitted).")
            return new_id # Return the new ID on success

        except pymssql.Error as e:
            logging.error(f"Database error adding task: {str(e)}", exc_info=True)
            # No explicit rollback needed
            return None # Indicate failure

        except Exception as e:
            logging.error(f"Unexpected error adding task: {str(e)}", exc_info=True)
            # No explicit rollback needed
            return None # Indicate failure
    
    def get_filtered_tasks(self, show_completed: bool = False, category: Optional[str] = None, 
                          main_staff: Optional[str] = None, assigned_to: Optional[str] = None,
                          show_deleted: bool = False) -> List[Dict[str, Any]]:
        """Get tasks filtered by various criteria."""
        filtered_tasks = []
        
        for task in self._tasks:
            # Skip deleted tasks unless explicitly requested
            if task.get("deleted", False) and not show_deleted:
                continue
                
            # Apply other filters as before
            if not show_completed and task.get("completed", False):
                continue
            
            task_category = task.get("category", "")
            if category and category != "All":
                if not task_category or task_category.lower() != category.lower():
                    continue
                
            task_main_staff = task.get("main_staff", "")
            if main_staff and main_staff != "All":
                if not task_main_staff or task_main_staff.lower() != main_staff.lower():
                    continue
                
            task_assigned_to = task.get("assigned_to", "")
            if assigned_to and assigned_to != "All":
                if not task_assigned_to or task_assigned_to.lower() != assigned_to.lower():
                    continue
                
            filtered_tasks.append(task)
        
        return filtered_tasks
    
    def update_task(self, task_id: int, **kwargs) -> bool: # Return True/False
        """Update a task in the SQL Server database (autocommitted). Returns True on success, False on failure."""
        if not self.conn or not self.cursor:
             logging.error(f"Cannot update task {task_id}: No database connection.")
             return False
        try:
            # --- Input Validation and Formatting ---
            def validate_date(date_str):
                if date_str:
                    try:
                        # Ensure the date is in the correct format YYYY-MM-DD
                        datetime.datetime.strptime(date_str, "%Y-%m-%d")
                        return date_str
                    except ValueError:
                         logging.warning(f"Invalid date format during update: {date_str}. Setting to None.")
                         return None
                return None

            # Process date fields if present in kwargs
            if 'due_date' in kwargs:
                kwargs['due_date'] = validate_date(kwargs['due_date'])
            if 'requested_date' in kwargs:
                kwargs['requested_date'] = validate_date(kwargs['requested_date'])
            if 'target_start' in kwargs:
                kwargs['target_start'] = validate_date(kwargs['target_start'])
            if 'target_finish' in kwargs:
                kwargs['target_finish'] = validate_date(kwargs['target_finish'])

            # Convert any remaining date objects to strings (shouldn't happen if TaskDialog passes strings)
            for key, value in kwargs.items():
                if isinstance(value, datetime.date):
                    kwargs[key] = value.strftime("%Y-%m-%d")

            # Ensure MHR values are integers if present
            if 'qtd_mhr' in kwargs:
                 try:
                     kwargs['qtd_mhr'] = int(kwargs['qtd_mhr']) if kwargs['qtd_mhr'] is not None else None
                 except (ValueError, TypeError):
                     logging.warning(f"Invalid qtd_mhr value during update: {kwargs['qtd_mhr']}. Skipping update for this field.")
                     del kwargs['qtd_mhr'] # Remove invalid value
            if 'actual_mhr' in kwargs:
                 try:
                     kwargs['actual_mhr'] = int(kwargs['actual_mhr']) if kwargs['actual_mhr'] is not None else None
                 except (ValueError, TypeError):
                     logging.warning(f"Invalid actual_mhr value during update: {kwargs['actual_mhr']}. Skipping update for this field.")
                     del kwargs['actual_mhr'] # Remove invalid value


            # --- Metadata and Query Building ---
            now_dt = datetime.datetime.now() # Use datetime object for DB if appropriate
            # now_str = now_dt.strftime("%Y-%m-%d %H:%M:%S")
            current_user = self._get_current_user()

            # Add last_modified and modified_by to kwargs for update
            kwargs['last_modified'] = now_dt
            kwargs['modified_by'] = current_user

            # Build the SQL SET clause dynamically
            set_clauses = []
            params = []
            for key, value in kwargs.items():
                # Ensure we don't try to update the ID itself
                if key != "id":
                    # Use correct placeholder syntax for pymssql
                    set_clauses.append(f"[{key}] = %s") # Bracket column names for safety
                    params.append(value)

            # Check if there's anything to update
            if not set_clauses:
                logging.warning(f"No valid fields provided for updating task {task_id}.")
                return True # No changes needed, consider it a success

            # Finalize the SQL query
            sql = f"UPDATE [dbo].[Tasks] SET {', '.join(set_clauses)} WHERE id = %s"
            params.append(task_id) # Add the task_id for the WHERE clause

            # --- Database Execution (autocommitted) ---
            self.cursor.execute(sql, params)
            rows_affected = self.cursor.rowcount # Check if the row was actually updated

            if rows_affected == 0:
                # This might mean the task_id doesn't exist or the values were the same
                logging.warning(f"Update task {task_id}: Task not found or no changes made in DB (Rows affected: 0).")
                # Check if task exists locally to differentiate
                if not self._find_task_by_id_local(task_id):
                     messagebox.showerror("Error", f"Task with ID {task_id} not found for update.")
                     return False
                 # If task exists but no rows affected, maybe values were identical? Treat as success.

            # --- UPDATE LOCAL MEMORY CACHE ---
            # Find and update the task in the local _tasks list
            local_task = self._find_task_by_id_local(task_id)
            if local_task:
                # Update the local task dictionary with new values
                for key, value in kwargs.items():
                    local_task[key] = value
                logging.debug(f"Updated task {task_id} in local memory cache")
            else:
                logging.warning(f"Task {task_id} not found in local cache after DB update")

            logging.info(f"Task {task_id} updated successfully by {current_user} (autocommitted).")
            return True # Indicate success

        except pymssql.Error as e:
            logging.error(f"Database error updating task {task_id}: {str(e)}", exc_info=True)
            # No explicit rollback needed
            messagebox.showerror("Database Error", f"Error updating task: {str(e)}")
            return False # Indicate failure

        except Exception as e:
            logging.error(f"Unexpected error updating task {task_id}: {str(e)}", exc_info=True)
            # No explicit rollback needed
            messagebox.showerror("Application Error", f"An unexpected error occurred during update: {str(e)}")
            return False # Indicate failure


    def _find_task_by_id(self, task_id: int) -> Optional[Dict[str, Any]]:
        """Find a task by ID directly from the SQL Server database."""
        if not self.conn or not self.cursor:
             logging.error(f"Cannot find task {task_id}: No database connection.")
             return None
        try:
            # SQL query to find a task by ID
            sql = "SELECT * FROM Tasks WHERE id = %s"
            self.cursor.execute(sql, (task_id,))
            row = self.cursor.fetchone()

            if row:
                # Convert row to a dictionary using column descriptions
                column_names = [desc[0] for desc in self.cursor.description]
                task = dict(zip(column_names, row))

                # Apply the same data type conversions as in _load_tasks
                for date_key in ['created_date', 'due_date', 'requested_date', 'target_start', 'target_finish']:
                    if date_key in task and isinstance(task[date_key], (datetime.datetime, datetime.date)):
                        task[date_key] = task[date_key].strftime("%Y-%m-%d")
                    elif date_key in task and not isinstance(task[date_key], str):
                        task[date_key] = None
                if 'last_modified' in task and isinstance(task['last_modified'], datetime.datetime):
                    task['last_modified'] = task['last_modified'].strftime("%Y-%m-%d %H:%M:%S")
                elif 'last_modified' in task:
                    task['last_modified'] = None
                if 'completed' in task: task['completed'] = bool(task['completed'])
                if 'deleted' in task: task['deleted'] = bool(task['deleted'])
                else: task['deleted'] = False
                for num_key in ['qtd_mhr', 'actual_mhr']:
                    if num_key in task and task[num_key] is not None:
                        try: task[num_key] = int(task[num_key])
                        except (ValueError, TypeError): task[num_key] = 0
                    else: task[num_key] = 0

                return task
            else:
                logging.warning(f"Task with ID {task_id} not found in database.")
                return None
        except pymssql.Error as e:
            logging.error(f"Database error finding task by ID {task_id}: {str(e)}", exc_info=True)
            # Avoid showing messagebox here, return None
            return None
        except Exception as e:
            logging.error(f"Unexpected error finding task {task_id}: {str(e)}", exc_info=True)
            return None

    def _find_task_by_id_local(self, task_id: int) -> Optional[Dict[str, Any]]:
        """Find a task by ID in the local cache (_tasks list)."""
        for task in self._tasks:
            if task.get('id') == task_id:
                return task
        return None


    def reload_tasks(self, load_all=False):
        """Reload tasks from the database into the local cache.
        
        Args:
            load_all: Kept for compatibility but ignored (always loads all)
        """
        logging.info("Reloading all tasks from database...")
        # Re-establish connection if needed (optional, depends on connection stability)
        # self._connect()
        if self.conn:
            self.total_task_count = self._get_total_task_count()
            self._tasks = self._load_tasks(limit=None)  # Always load all tasks
            self.all_tasks_loaded = True
            logging.info(f"Task reload complete. {len(self._tasks)} tasks loaded.")
            return self._tasks
        else:
            logging.error("Cannot reload tasks, no database connection.")
            self._tasks = []
            # Show error to user?
            messagebox.showerror("Connection Error", "Cannot reload tasks, database connection lost.")
            return []

    def begin_transaction(self):
        """Begin an explicit database transaction."""
        try:
            # Check if connection is valid before executing
            if self.conn and self.cursor:
                 # pymssql doesn't have an explicit BEGIN TRANSACTION command like standard SQL sometimes.
                 # Turning autocommit off implicitly starts a transaction.
                 # However, since we connect with autocommit=True, we need to manage this carefully.
                 # For pymssql, explicit transaction control might involve SET IMPLICIT_TRANSACTIONS OFF/ON
                 # or simply relying on the commit/rollback methods to define boundaries.
                 # Let's assume calling commit/rollback manages the scope when autocommit is True globally.
                 # A safer approach might be to temporarily disable autocommit if pymssql supports it per-session.
                 # For simplicity here, we'll just log the intention. The commit/rollback call IS the boundary.
                 logging.debug("Starting explicit transaction block (commit/rollback will define scope).")
            else:
                 logging.error("Cannot begin transaction: No database connection.")
                 raise pymssql.Error("No database connection to begin transaction.")
        except pymssql.Error as e:
            logging.error(f"Error attempting to manage transaction state: {e}", exc_info=True)
            raise

    def commit_transaction(self):
        """Commit the current explicit database transaction."""
        try:
             if self.conn:
                 self.conn.commit()
                 logging.debug("Explicit transaction committed.")
             else:
                 logging.error("Cannot commit transaction: No database connection.")
                 # Raise error? Or just log? Let's raise.
                 raise pymssql.Error("No database connection to commit transaction.")
        except pymssql.Error as e:
            logging.error(f"Error committing explicit transaction: {e}", exc_info=True)
            raise # Re-raise so the caller knows commit failed

    def rollback_transaction(self):
        """Rollback the current explicit database transaction."""
        try:
             if self.conn:
                 self.conn.rollback()
                 logging.debug("Explicit transaction rolled back.")
             else:
                 logging.error("Cannot rollback transaction: No database connection.")
                 # Don't raise here, as rollback is usually called during error handling
        except pymssql.Error as e:
            logging.error(f"Error rolling back explicit transaction: {e}", exc_info=True)
            # Don't re-raise during error handling rollback

    def get_sql_username(self):
        """Get the SQL username used for the database connection."""
        return self.user
    
    def generate_permission_script(self):
        """
        Generate SQL script that a DBA can run to grant necessary permissions.
        
        Returns:
            str: SQL script to grant permissions
        """
        script = f"""-- TaskManager Database Permissions Script
-- Run this script as a DBA to grant necessary permissions to TaskManager users

USE TaskManagerDB;
GO

-- Grant permissions to all TaskManager users
"""
        
        # Get all unique SQL usernames from USER_MAPPING
        sql_users = set()
        for data in USER_MAPPING.values():
            sql_user = data.get('sql_user', '')
            if sql_user:
                sql_users.add(sql_user)
        
        for sql_user in sorted(sql_users):
            script += f"""
-- Permissions for {sql_user}
GRANT SELECT, INSERT, UPDATE, DELETE ON dbo.Tasks TO [{sql_user}];
GRANT SELECT, INSERT, UPDATE, DELETE ON dbo.Users TO [{sql_user}];
"""
        
        script += "\nGO\nPRINT 'Permissions granted successfully!';"
        return script
    
    def save_user_to_db(self, display_name, tcn, sql_username):
        """
        Save a new user to the Users table.
        
        Args:
            display_name: The display name (e.g., "Mike")
            tcn: The TCN number (e.g., "a0012345")
            sql_username: The SQL username (e.g., "TaskUser8")
        
        Returns:
            bool: True if successful, False otherwise
        """
        if not self.conn or not self.cursor:
            logging.error(f"Cannot save user {display_name}: No database connection.")
            return False
        
        try:
            # Check if user already exists
            self.cursor.execute("""
                SELECT COUNT(*) FROM Users 
                WHERE display_name = %s OR tcn = %s OR sql_username = %s
            """, (display_name, tcn, sql_username))
            
            if self.cursor.fetchone()[0] > 0:
                logging.warning(f"User {display_name} already exists in database.")
                return False
            
            # Insert new user
            self.cursor.execute("""
                INSERT INTO Users (display_name, tcn, sql_username)
                VALUES (%s, %s, %s)
            """, (display_name, tcn, sql_username))
            
            logging.info(f"Saved user {display_name} to database successfully.")
            return True
            
        except pymssql.Error as e:
            logging.error(f"Error saving user {display_name}: {str(e)}", exc_info=True)
            return False
        except Exception as e:
            logging.error(f"Unexpected error saving user {display_name}: {str(e)}", exc_info=True)
            return False
    
    def update_user_in_db(self, old_display_name, new_display_name, new_tcn):
        """
        Update an existing user in the Users table.
        
        Args:
            old_display_name: The current display name
            new_display_name: The new display name
            new_tcn: The new TCN number
        
        Returns:
            bool: True if successful, False otherwise
        """
        if not self.conn or not self.cursor:
            logging.error(f"Cannot update user {old_display_name}: No database connection.")
            return False
        
        try:
            self.cursor.execute("""
                UPDATE Users
                SET display_name = %s, tcn = %s
                WHERE display_name = %s
            """, (new_display_name, new_tcn, old_display_name))
            
            logging.info(f"Updated user {old_display_name} to {new_display_name} in database.")
            return True
            
        except pymssql.Error as e:
            logging.error(f"Error updating user {old_display_name}: {str(e)}", exc_info=True)
            return False
    
    def delete_user_from_db(self, display_name):
        """
        Delete a user from the Users table (soft delete by setting active=0).
        
        Args:
            display_name: The display name to delete
        
        Returns:
            bool: True if successful, False otherwise
        """
        if not self.conn or not self.cursor:
            logging.error(f"Cannot delete user {display_name}: No database connection.")
            return False
        
        try:
            self.cursor.execute("""
                UPDATE Users
                SET active = 0
                WHERE display_name = %s
            """, (display_name,))
            
            logging.info(f"Deleted user {display_name} from database.")
            return True
            
        except pymssql.Error as e:
            logging.error(f"Error deleting user {display_name}: {str(e)}", exc_info=True)
            return False
    
    def load_users_from_db(self):
        """
        Load all active users from the Users table and merge with default USER_MAPPING.
        
        Returns:
            dict: Dictionary of user mappings {display_name: {tcn, sql_username}}
        """
        # Start with the default USER_MAPPING (the 7 hardcoded users)
        users = USER_MAPPING.copy()
        
        if not self.conn or not self.cursor:
            logging.error("Cannot load users: No database connection.")
            return users
        
        try:
            self.cursor.execute("""
                SELECT display_name, tcn, sql_username
                FROM Users
                WHERE active = 1
                ORDER BY display_name
            """)
            
            # Merge database users with defaults (database takes precedence for matching names)
            db_count = 0
            for row in self.cursor.fetchall():
                display_name, tcn, sql_username = row
                users[display_name] = {
                    "tcn": tcn,
                    "sql_user": sql_username
                }
                db_count += 1
            
            logging.info(f"Loaded {db_count} users from database, merged with {len(USER_MAPPING)} default users. Total: {len(users)} users.")
            return users
            
        except pymssql.Error as e:
            logging.error(f"Error loading users: {str(e)}", exc_info=True)
            # Return default users if database fails
            logging.info("Falling back to default USER_MAPPING due to database error.")
            return users

    def recover_task(self, task_id: int) -> bool:
        """Recover a soft-deleted task (autocommitted)."""
        if not self.conn or not self.cursor:
             logging.error(f"Cannot recover task {task_id}: No database connection.")
             messagebox.showerror("Database Error", "No database connection available.")
             return False

        # Find task locally or in DB (logic remains same as previous version)
        task = self._find_task_by_id_local(task_id)
        task_from_db = None
        if not task or not task.get("deleted", False):
             task_from_db = self._find_task_by_id(task_id)
             if not task_from_db: logging.warning(f"Recover task {task_id}: Task not found."); return False
             if not task_from_db.get("deleted", False): logging.warning(f"Recover task {task_id}: Task is not deleted."); return False
             task = task_from_db # Use DB version

        try:
            now_dt = datetime.datetime.now()
            current_user = self._get_current_user()
            sql = "UPDATE [dbo].[Tasks] SET deleted = 0, last_modified = %s, modified_by = %s WHERE id = %s AND deleted = 1"
            params = (now_dt, current_user, task_id)

            self.cursor.execute(sql, params)
            rows_affected = self.cursor.rowcount
            # No explicit commit needed

            if rows_affected > 0:
                 logging.info(f"Task {task_id} recovered in DB by {current_user} (autocommitted).")
                 # Update local cache (logic remains same)
                 task["deleted"] = False
                 task["last_modified"] = now_dt.strftime("%Y-%m-%d %H:%M:%S")
                 task["modified_by"] = current_user
                 if not self._find_task_by_id_local(task_id): self._tasks.append(task) # Add back if missing locally
                 return True
            else:
                 logging.warning(f"Recover task {task_id} affected 0 rows (autocommit). Task might not exist or was already recovered.")
                 # (Sync logic remains same)
                 db_task = self._find_task_by_id(task_id)
                 if db_task and not db_task.get('deleted'):
                      if task: # Update local if it exists
                         task['deleted'] = False
                         task['last_modified'] = db_task.get('last_modified')
                         task['modified_by'] = db_task.get('modified_by')
                      return True # State matches request
                 return False

        except pymssql.Error as e:
            logging.error(f"Database error recovering task {task_id}: {str(e)}", exc_info=True)
            # No explicit rollback needed
            messagebox.showerror("Database Error", f"Error recovering task: {str(e)}")
            return False
        except Exception as e:
             logging.error(f"Unexpected error recovering task {task_id}: {str(e)}", exc_info=True)
             # No explicit rollback needed
             messagebox.showerror("Application Error", f"An unexpected error occurred during recovery: {str(e)}")
             return False

    def permanently_delete_task(self, task_id: int) -> bool:
        """Permanently remove a task from the system (autocommitted)."""
        if not self.conn or not self.cursor:
             logging.error(f"Cannot permanently delete task {task_id}: No database connection.")
             messagebox.showerror("Database Error", "No database connection available.")
             return False
        try:
            # Delete from the database
            sql = "DELETE FROM [dbo].[Tasks] WHERE id = %s"
            self.cursor.execute(sql, (task_id,))
            rows_affected = self.cursor.rowcount
            # No explicit commit needed

            if rows_affected > 0:
                 logging.info(f"Task {task_id} permanently deleted from DB (autocommitted).")
                 # Remove from the local list
                 original_len = len(self._tasks)
                 self._tasks = [task for task in self._tasks if task["id"] != task_id]
                 if len(self._tasks) == original_len:
                      logging.warning(f"Task {task_id} permanently deleted from DB but was not found in local cache.")

                 # Remove from older deleted_tasks list if present
                 self._deleted_tasks = [t for t in self._deleted_tasks if t["id"] != task_id]
                 # Also consider removing from the main deleted_tasks_stack if it was there?
                 # This might interfere with undo if the permanent delete is meant to be undoable itself.
                 # Let's assume permanent delete is final and doesn't affect undo stack.

                 return True
            else:
                 logging.warning(f"Permanent delete for task {task_id} affected 0 rows. Task may not have existed.")
                 # Check if it existed locally
                 if self._find_task_by_id_local(task_id):
                      # It existed locally but not in DB? Sync issue. Remove locally.
                      self._tasks = [task for task in self._tasks if task["id"] != task_id]
                 return False # Didn't delete anything from DB

        except pymssql.Error as e:
            logging.error(f"Database error permanently deleting task {task_id}: {str(e)}", exc_info=True)
            # No explicit rollback needed
            messagebox.showerror("Database Error", f"Error permanently deleting task: {str(e)}")
            return False
        except Exception as e:
             logging.error(f"Unexpected error permanently deleting task {task_id}: {str(e)}", exc_info=True)
             # No explicit rollback needed
             messagebox.showerror("Application Error", f"An unexpected error occurred during permanent deletion: {str(e)}")
             return False


    def permanently_delete_all_deleted_tasks(self) -> int:
        """Permanently remove all soft-deleted tasks (autocommitted)."""
        if not self.conn or not self.cursor:
             logging.error("Cannot permanently delete all tasks: No database connection.")
             messagebox.showerror("Database Error", "No database connection available.")
             return 0
        try:
            # Delete all deleted tasks from the database
            sql = "DELETE FROM [dbo].[Tasks] WHERE deleted = 1"
            self.cursor.execute(sql)
            deleted_count = self.cursor.rowcount
            # No explicit commit needed

            if deleted_count > 0:
                 logging.info(f"Permanently deleted {deleted_count} tasks marked as deleted from DB (autocommitted).")
            else:
                 logging.info("No tasks marked as deleted found to permanently delete.")

            # Update local list regardless of DB count, to ensure consistency
            original_count = len(self._tasks)
            self._tasks = [task for task in self._tasks if not task.get("deleted", False)]
            local_removed_count = original_count - len(self._tasks)
            if local_removed_count != deleted_count:
                 logging.warning(f"DB deleted count ({deleted_count}) differs from local cache removed count ({local_removed_count}). Cache might have been out of sync.")

            # Clear the older deleted_tasks list
            self._deleted_tasks = []
            # Clear the main undo stack? This is debatable. If permanent delete is final, maybe clear it.
            # self.deleted_tasks_stack = [] # Uncomment if permanent delete should clear undo history

            return deleted_count
        except pymssql.Error as e:
            logging.error(f"Database error permanently deleting all tasks: {str(e)}", exc_info=True)
            # No explicit rollback needed
            messagebox.showerror("Database Error", f"Error permanently deleting tasks: {str(e)}")
            return 0
        except Exception as e:
             logging.error(f"Unexpected error permanently deleting all tasks: {str(e)}", exc_info=True)
             # No explicit rollback needed
             messagebox.showerror("Application Error", f"An unexpected error occurred during permanent deletion: {str(e)}")
             return 0

    @property
    def tasks(self):
        """Get the list of tasks. Use property for consistent access."""
        return self._tasks

    def recover_batch(self, task_ids: List[int]) -> List[int]:
        """Recover multiple tasks at once (each recovery is autocommitted)."""
        # This method calls self.recover_task which now uses autocommit.
        # Therefore, the overall batch recovery is NOT atomic.
        # If atomicity is required for batch recovery, this needs explicit transaction handling like batch_delete.
        # For now, we keep it simple: recover one by one.
        recovered_ids = []
        for task_id in task_ids:
            if self.recover_task(task_id): # recover_task now uses autocommit
                recovered_ids.append(task_id)
        return recovered_ids

    def delete_task(self, task_id: int) -> bool:
        """
        Soft delete a task by marking it as deleted (autocommitted).
        Updates local cache and undo stack. Returns True/False.
        """
        if not self.conn or not self.cursor:
            logging.error(f"Cannot delete task {task_id}: No database connection.")
            messagebox.showerror("Database Error", "No database connection available.")
            return False

        task_to_delete = self._find_task_by_id_local(task_id)
        task_copy = None
        if task_to_delete and not task_to_delete.get("deleted", False):
             task_copy = task_to_delete.copy()
             # Keep older _deleted_tasks list update if needed for compatibility
             self._deleted_tasks.append(task_copy)

        try:
            now_dt = datetime.datetime.now()
            current_user = self._get_current_user()

            sql = "UPDATE [dbo].[Tasks] SET deleted = 1, last_modified = %s, modified_by = %s WHERE id = %s AND deleted = 0"
            params = (now_dt, current_user, task_id)

            self.cursor.execute(sql, params)
            rows_affected = self.cursor.rowcount
            # No explicit commit needed

            if rows_affected > 0:
                logging.info(f"Task {task_id} marked as deleted in DB by {current_user} (autocommitted).")
                if task_to_delete: # Update local cache if it exists
                    task_to_delete["deleted"] = True
                    task_to_delete["last_modified"] = now_dt.strftime("%Y-%m-%d %H:%M:%S")
                    task_to_delete["modified_by"] = current_user
                self.deleted_tasks_stack.append([task_id]) # Add to primary undo stack
                return True
            elif task_to_delete and task_to_delete.get("deleted"):
                 logging.warning(f"Task {task_id} was already marked as deleted.")
                 return True # State matches request
            else:
                 logging.warning(f"Soft delete for task {task_id} affected 0 rows (autocommit). Task might not exist or was already deleted.")
                 if task_to_delete and not task_to_delete.get("deleted"):
                      messagebox.showwarning("Sync Issue?", f"Could not mark task {task_id} as deleted in the database. It might have been deleted by another user. Refreshing data is recommended.")
                 return False # DB state not changed as expected

        except pymssql.Error as e:
            logging.error(f"Database error soft-deleting task {task_id}: {str(e)}", exc_info=True)
            # No explicit rollback needed
            messagebox.showerror("Database Error", f"Error deleting task: {str(e)}")
            if task_copy and task_copy in self._deleted_tasks: self._deleted_tasks.remove(task_copy)
            return False
        except Exception as e:
             logging.error(f"Unexpected error deleting task {task_id}: {str(e)}", exc_info=True)
             # No explicit rollback needed
             messagebox.showerror("Application Error", f"An unexpected error occurred during deletion: {str(e)}")
             if task_copy and task_copy in self._deleted_tasks: self._deleted_tasks.remove(task_copy)
             return False

    def batch_delete_tasks(self, task_ids: List[int]) -> Tuple[List[int], List[int]]:
        """
        Soft delete multiple tasks using an explicit transaction.
        Updates local cache and undo stack. Returns (succeeded_ids, failed_ids).
        """
        if not task_ids: return [], []
        if not self.conn or not self.cursor:
            logging.error("Cannot batch delete tasks: No database connection.")
            messagebox.showerror("Database Error", "No database connection available.")
            return [], list(task_ids) # Return all as failed

        success_ids = []
        failed_ids = [] # Start with empty failed
        tasks_to_undo = []

        # Prepare data for undo stack before starting transaction
        eligible_task_ids_for_delete = []
        for task_id in task_ids:
             task = self._find_task_by_id_local(task_id)
             if task and not task.get("deleted", False):
                  tasks_to_undo.append(task.copy())
                  eligible_task_ids_for_delete.append(task_id)
             else:
                 # Task doesn't exist locally or is already deleted
                 failed_ids.append(task_id)

        if not eligible_task_ids_for_delete:
             logging.warning("Batch delete: No eligible tasks found to delete.")
             return [], list(task_ids) # All originally passed IDs failed/skipped

        try:
            # Begin explicit transaction (overrides autocommit for this block)
            self.begin_transaction()

            now_dt = datetime.datetime.now()
            current_user = self._get_current_user()
            placeholders = ', '.join(['%s'] * len(eligible_task_ids_for_delete))

            sql = f"""
                UPDATE [dbo].[Tasks]
                SET deleted = 1, last_modified = %s, modified_by = %s
                WHERE id IN ({placeholders}) AND deleted = 0
            """
            params = [now_dt, current_user] + eligible_task_ids_for_delete

            self.cursor.execute(sql, params)
            rows_affected = self.cursor.rowcount

            # Commit the explicit transaction
            self.commit_transaction()

            logging.info(f"Batch delete transaction committed for {len(eligible_task_ids_for_delete)} eligible tasks. Rows affected in DB: {rows_affected}.")

            # If rows_affected doesn't match len(eligible_task_ids_for_delete), some might have been deleted between local check and DB update.
            # We'll consider the operation successful for those IDs passed to the DB if no error occurred.

            # Update local cache for successfully processed IDs
            now_str = now_dt.strftime("%Y-%m-%d %H:%M:%S")
            for task_id in eligible_task_ids_for_delete:
                 task = self._find_task_by_id_local(task_id)
                 if task: # Should exist as we checked before
                     task['deleted'] = True
                     task['last_modified'] = now_str
                     task['modified_by'] = current_user
                 success_ids.append(task_id) # Add to success list

            # Add successful deletions to the undo stack
            if success_ids:
                self.deleted_tasks_stack.append(success_ids)
                self._deleted_tasks.extend(tasks_to_undo)

            logging.info(f"Batch delete result: {len(success_ids)} succeeded, {len(failed_ids)} failed/skipped.")

        except pymssql.Error as e:
            logging.error(f"Database error during batch delete transaction: {str(e)}", exc_info=True)
            try:
                self.rollback_transaction() # Rollback explicit transaction
            except pymssql.Error as rb_e:
                logging.error(f"Error during rollback after batch delete error: {rb_e}", exc_info=True)
            messagebox.showerror("Database Error", f"Error deleting tasks: {str(e)}")
            success_ids = [] # Reset success on error
            failed_ids = list(task_ids) # Mark all originally passed IDs as failed

        except Exception as e:
            logging.error(f"Unexpected error during batch delete transaction: {str(e)}", exc_info=True)
            try:
                 if self.conn: self.rollback_transaction()
            except Exception as rb_e:
                 logging.error(f"Error during rollback after unexpected batch delete error: {rb_e}", exc_info=True)
            messagebox.showerror("Application Error", f"An unexpected error occurred during batch deletion: {str(e)}")
            success_ids = []
            failed_ids = list(task_ids)

        return list(set(success_ids)), list(set(failed_ids))

    def recover_all_deleted_tasks(self) -> int:
        """
        Recover all soft-deleted tasks by setting their 'deleted' flag to 0 (autocommitted).
        Returns the number of tasks recovered in the database.
        """
        if not self.conn or not self.cursor:
             logging.error("Cannot recover all deleted tasks: No database connection.")
             messagebox.showerror("Database Error", "No database connection available.")
             return 0

        recovered_count_db = 0
        try:
            now_dt = datetime.datetime.now()
            current_user = self._get_current_user()

            # Update all deleted tasks in the database in one go
            sql = "UPDATE [dbo].[Tasks] SET deleted = 0, last_modified = %s, modified_by = %s WHERE deleted = 1"
            params = (now_dt, current_user)

            self.cursor.execute(sql, params)
            recovered_count_db = self.cursor.rowcount
            # No explicit commit needed due to autocommit=True

            if recovered_count_db > 0:
                 logging.info(f"Recovered {recovered_count_db} tasks marked as deleted in DB (autocommitted).")

                 # Update local cache
                 now_str = now_dt.strftime("%Y-%m-%d %H:%M:%S")
                 recovered_in_cache = 0
                 for task in self._tasks:
                     if task.get("deleted", False):
                         task["deleted"] = False
                         task["last_modified"] = now_str
                         task["modified_by"] = current_user
                         recovered_in_cache += 1
                 logging.info(f"Updated {recovered_in_cache} tasks in local cache during recover all.")
                 # Note: DB count and cache count might differ if cache was out of sync

            else:
                 logging.info("No tasks marked as deleted found to recover.")

            # Clear the older _deleted_tasks list (if still used)
            self._deleted_tasks = []
            # Clear the main undo stack? Recover all might be considered a bulk action that clears undo.
            # self.deleted_tasks_stack = [] # Uncomment if Recover All should clear undo history

            return recovered_count_db

        except pymssql.Error as e:
            logging.error(f"Database error recovering all deleted tasks: {str(e)}", exc_info=True)
            # No explicit rollback needed
            messagebox.showerror("Database Error", f"Error recovering all tasks: {str(e)}")
            return 0
        except Exception as e:
             logging.error(f"Unexpected error recovering all deleted tasks: {str(e)}", exc_info=True)
             # No explicit rollback needed
             messagebox.showerror("Application Error", f"An unexpected error occurred during recover all: {str(e)}")
             return 0


class MainMenuApp:
    """Main menu launcher for Task Manager application."""
    
    def __init__(self, root):
        """Initialize the main menu."""
        self.root = root
        
        # Hide window during setup to prevent flash
        self.root.withdraw()
        
        self.root.title("Task Manager - Main Menu")
        self.root.geometry("900x600")
        self.root.minsize(800, 500)
        
        # Set application icon
        try:
            if getattr(sys, 'frozen', False):
                base_path = os.path.dirname(sys.executable)
            else:
                base_path = os.path.dirname(os.path.abspath(__file__))
            
            icon_path = os.path.join(base_path, 'TaskManager_main.png')
            if os.path.exists(icon_path):
                self.main_icon_image = tk.PhotoImage(file=icon_path)
                self.root.iconphoto(True, self.main_icon_image)
            else:
                icon_path = os.path.join(base_path, 'TaskManager_main.ico')
                if os.path.exists(icon_path):
                    self.root.iconbitmap(icon_path)
        except Exception as e:
            logging.warning(f"Could not set main menu icon: {e}")
        
        # Initialize variables
        self.current_view = None
        self.task_manager_app = None
        self.saved_position = None  # To save window position
        
        # Initialize TaskManager with database connection
        try:
            self.task_manager = TaskManager()
        except Exception as e:
            logging.error(f"Failed to initialize TaskManager: {str(e)}", exc_info=True)
            self.task_manager = None
            messagebox.showwarning("Database Connection", 
                f"Could not connect to the database.\n\n"
                f"Some features (like adding SQL users) may not be available.\n\n"
                f"Error: {str(e)}", 
                parent=self.root)
        
        # Show main menu
        self.show_main_menu()
        
        # Center the window
        self.center_window()
        
        # Show window after everything is set up
        self.root.deiconify()
        
        # Close protocol
        self.root.protocol("WM_DELETE_WINDOW", self.on_close)
    
    def center_window(self):
        """Center the window on the screen."""
        self.root.update_idletasks()
        width = self.root.winfo_width()
        height = self.root.winfo_height()
        x = (self.root.winfo_screenwidth() // 2) - (width // 2)
        y = (self.root.winfo_screenheight() // 2) - (height // 2)
        self.root.geometry(f'{width}x{height}+{x}+{y}')
    
    def clear_window(self):
        """Clear all widgets from the root window."""
        for widget in self.root.winfo_children():
            widget.destroy()
    
    def show_main_menu(self):
        """Display the main menu."""
        self.clear_window()
        self.current_view = "menu"
        self.root.title("Task Manager - Main Menu")
        
        # Force main menu to fixed size - reduced height to eliminate white space
        # Reset min/max size constraints first
        self.root.minsize(350, 280)
        self.root.maxsize(350, 280)
        self.root.resizable(False, False)
        
        # Restore saved position or center if first time
        if self.saved_position:
            x, y = self.saved_position
            # Set size AND position explicitly
            self.root.geometry(f"350x280+{x}+{y}")
        else:
            # Set size first, then center
            self.root.geometry("350x280")
            self.center_window()
        
        # Force update to apply geometry changes
        self.root.update_idletasks()
        
        # Create main container with gradient-like background
        main_frame = tk.Frame(self.root, bg='white')
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Header section with color
        header_frame = tk.Frame(main_frame, bg='#2c3e50', height=50)
        header_frame.pack(fill=tk.X)
        header_frame.pack_propagate(False)
        
        # Title label in header
        title_label = tk.Label(
            header_frame,
            text="Task Manager",
            font=('Segoe UI', 14, 'bold'),
            bg='#2c3e50',
            fg='white'
        )
        title_label.pack(expand=True)
        
        # Content frame - don't expand vertically to avoid empty space
        content_frame = tk.Frame(main_frame, bg='white')
        content_frame.pack(fill=tk.X, padx=25, pady=(10, 10))
        
        # Button container
        button_frame = tk.Frame(content_frame, bg='white')
        button_frame.pack(fill=tk.X)
        
        # Modern button style configuration with explicit height
        button_config = {
            'width': 22,
            'height': 1,
            'font': ('Segoe UI', 10),
            'cursor': 'hand2',
            'relief': tk.FLAT,
            'bd': 0,
            'padx': 20,
            'pady': 10
        }
        
        # Task Manager button
        task_manager_btn = tk.Button(
            button_frame,
            text="📋 Task Manager",
            command=self.open_task_manager,
            bg='#3498db',
            fg='white',
            activebackground='#2980b9',
            activeforeground='white',
            **button_config
        )
        task_manager_btn.pack(pady=4, fill=tk.X)
        
        # User Management button
        user_mgmt_btn = tk.Button(
            button_frame,
            text="👥 User Management",
            command=self.open_add_user,
            bg='#9b59b6',
            fg='white',
            activebackground='#8e44ad',
            activeforeground='white',
            **button_config
        )
        user_mgmt_btn.pack(pady=4, fill=tk.X)
        
        # Export/Import button
        export_import_btn = tk.Button(
            button_frame,
            text="⇄ Export/Import",
            command=self.open_export_import,
            bg='#e67e22',
            fg='white',
            activebackground='#d35400',
            activeforeground='white',
            **button_config
        )
        export_import_btn.pack(pady=4, fill=tk.X)
        
        # Dashboard button (if enhanced features available)
        if self.task_manager and hasattr(self.task_manager, 'dashboard') and self.task_manager.dashboard:
            dashboard_btn = tk.Button(
                button_frame,
                text="📊 Dashboard",
                command=self.open_dashboard,
                bg='#2ecc71',
                fg='white',
                activebackground='#27ae60',
                activeforeground='white',
                **button_config
            )
            dashboard_btn.pack(pady=4, fill=tk.X)
        
        # Templates button (if enhanced features available)
        if self.task_manager and hasattr(self.task_manager, 'templates') and self.task_manager.templates:
            templates_btn = tk.Button(
                button_frame,
                text="📋 Templates",
                command=self.open_templates,
                bg='#16a085',
                fg='white',
                activebackground='#138d75',
                activeforeground='white',
                **button_config
            )
            templates_btn.pack(pady=4, fill=tk.X)
    
    def open_task_manager(self):
        """Open the Task Manager in the same window."""
        logging.info("Opening Task Manager from main menu...")
        
        try:
            # Save current window position before switching
            self.root.update_idletasks()
            self.saved_position = (self.root.winfo_x(), self.root.winfo_y())
            
            # Initialize TaskManager with database connection if not already done
            if not self.task_manager:
                self.task_manager = TaskManager()
            
            # Get main menu's current position and size BEFORE clearing/resizing
            self.root.update_idletasks()
            main_x = self.root.winfo_x()
            main_y = self.root.winfo_y()
            main_width = self.root.winfo_width()
            main_height = self.root.winfo_height()
            
            # Clear window and resize for task manager
            self.clear_window()
            self.current_view = "taskmanager"
            self.root.title("Task Manager")
            
            # Reset window constraints to allow resizing
            # Get screen dimensions for maxsize
            screen_width = self.root.winfo_screenwidth()
            screen_height = self.root.winfo_screenheight()
            
            # Calculate centered position relative to main menu
            window_width = 900
            window_height = 600
            x = main_x + (main_width // 2) - (window_width // 2)
            y = main_y + (main_height // 2) - (window_height // 2)
            
            # Set geometry and constraints for Task Manager
            self.root.geometry(f"{window_width}x{window_height}+{x}+{y}")
            self.root.minsize(800, 500)
            self.root.maxsize(screen_width, screen_height)  # Reset maxsize to screen size
            self.root.resizable(True, True)
            
            # Create a container frame for the TaskManagerApp content
            content_frame = tk.Frame(self.root)
            content_frame.pack(fill=tk.BOTH, expand=True)
            
            # Initialize TaskManagerApp in the content frame with back button callback
            self.task_manager_app = TaskManagerApp(content_frame, task_manager=self.task_manager, is_embedded=True, back_callback=self.back_to_menu)
            
        except ConnectionError as ce:
            logging.error(f"Database connection failed: {ce}")
            messagebox.showerror("Connection Error", f"Could not connect to database:\n{ce}")
            self.show_main_menu()
        except Exception as e:
            logging.error(f"Error opening Task Manager: {e}", exc_info=True)
            messagebox.showerror("Error", f"Could not open Task Manager:\n{e}")
            self.show_main_menu()
    
    def back_to_menu(self):
        """Return to main menu from Task Manager."""
        logging.info("Returning to main menu...")
        
        # Clean up task manager app if needed
        if self.task_manager_app:
            try:
                # Don't close database connection, just cleanup the app
                if hasattr(self.task_manager_app, 'cleanup'):
                    self.task_manager_app.cleanup()
            except Exception as e:
                logging.warning(f"Error during task manager cleanup: {e}")
            self.task_manager_app = None
        
        # Show main menu
        self.show_main_menu()
    
    def open_add_user(self):
        """Open User Management interface."""
        logging.info("Opening User Management from main menu...")
        
        try:
            # Save current window position before switching
            self.root.update_idletasks()
            self.saved_position = (self.root.winfo_x(), self.root.winfo_y())
            
            # Clear window and resize for user management - compact size
            self.clear_window()
            self.current_view = "usermanagement"
            self.root.title("User Management")
            self.root.geometry("400x500")
            self.root.minsize(400, 500)
            self.root.resizable(False, False)
            
            # Restore saved position or center
            if self.saved_position:
                x, y = self.saved_position
                self.root.geometry(f"400x500+{x}+{y}")
            else:
                self.center_window()
            
            # Create User Management interface
            self.user_mgmt = UserManagementApp(self.root, self, self.task_manager)
            self.user_mgmt.show()
            
        except Exception as e:
            logging.error(f"Error opening User Management: {str(e)}", exc_info=True)
            messagebox.showerror("Error", f"Failed to open User Management: {str(e)}", parent=self.root)
            self.show_main_menu()
    
    def open_export_import(self):
        """Show dialog to choose between Export and Import."""
        # Create dialog
        choice_dialog = tk.Toplevel(self.root)
        choice_dialog.title("Export/Import")
        choice_dialog.geometry("420x300")
        choice_dialog.minsize(420, 300)
        choice_dialog.configure(bg='white')
        choice_dialog.transient(self.root)
        choice_dialog.grab_set()
        
        # Center the dialog relative to main menu
        choice_dialog.update_idletasks()
        dialog_width = 420
        dialog_height = 300
        x = self.root.winfo_x() + (self.root.winfo_width() // 2) - (dialog_width // 2)
        y = self.root.winfo_y() + (self.root.winfo_height() // 2) - (dialog_height // 2)
        choice_dialog.geometry(f"+{x}+{y}")
        
        # Header section with gradient effect
        header_frame = tk.Frame(choice_dialog, bg='#2c3e50', height=50)
        header_frame.pack(fill=tk.X)
        header_frame.pack_propagate(False)
        
        title_label = tk.Label(
            header_frame,
            text="Export/Import Data",
            font=("Segoe UI", 14, "bold"),
            bg="#2c3e50",
            fg="white"
        )
        title_label.pack(expand=True)
        
        # Main container
        container = tk.Frame(choice_dialog, bg='white')
        container.pack(fill=tk.BOTH, expand=True, padx=30, pady=15)
        
        # Instruction label
        instruction_label = tk.Label(
            container,
            text="Choose an operation:",
            font=("Segoe UI", 10),
            bg="white",
            fg="#7f8c8d"
        )
        instruction_label.pack(pady=(0, 10))
        
        # Export button with modern flat design
        export_btn = tk.Button(
            container,
            text="  📤  Export Tasks to Excel",
            command=lambda: [choice_dialog.destroy(), self.open_export()],
            bg='#3498db',
            fg='white',
            activebackground='#2980b9',
            activeforeground='white',
            font=("Segoe UI", 11, "bold"),
            height=2,
            cursor="hand2",
            relief=tk.FLAT,
            bd=0
        )
        export_btn.pack(pady=6, fill=tk.X)
        
        # Hover effects for export button
        def on_export_enter(e):
            export_btn.config(bg='#2980b9')
        def on_export_leave(e):
            export_btn.config(bg='#3498db')
        export_btn.bind("<Enter>", on_export_enter)
        export_btn.bind("<Leave>", on_export_leave)
        
        # Import button with modern flat design
        import_btn = tk.Button(
            container,
            text="  📥  Import Tasks from Excel",
            command=lambda: [choice_dialog.destroy(), self.open_import()],
            bg='#27ae60',
            fg='white',
            activebackground='#229954',
            activeforeground='white',
            font=("Segoe UI", 11, "bold"),
            height=2,
            cursor="hand2",
            relief=tk.FLAT,
            bd=0
        )
        import_btn.pack(pady=6, fill=tk.X)
        
        # Hover effects for import button
        def on_import_enter(e):
            import_btn.config(bg='#229954')
        def on_import_leave(e):
            import_btn.config(bg='#27ae60')
        import_btn.bind("<Enter>", on_import_enter)
        import_btn.bind("<Leave>", on_import_leave)
        
        # Cancel button with subtle design
        cancel_btn = tk.Button(
            container,
            text="Cancel",
            command=choice_dialog.destroy,
            bg='#ecf0f1',
            fg='#34495e',
            activebackground='#bdc3c7',
            activeforeground='#2c3e50',
            font=("Segoe UI", 10, "bold"),
            width=25,
            height=2,
            cursor="hand2",
            relief=tk.FLAT,
            bd=0
        )
        cancel_btn.pack(pady=(15, 0))
        
        # Hover effects for cancel button
        def on_cancel_enter(e):
            cancel_btn.config(bg='#bdc3c7')
        def on_cancel_leave(e):
            cancel_btn.config(bg='#ecf0f1')
        cancel_btn.bind("<Enter>", on_cancel_enter)
        cancel_btn.bind("<Leave>", on_cancel_leave)
    
    def open_export(self):
        """Export tasks to Excel file."""
        try:
            # Import openpyxl when needed
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            messagebox.showerror(
                "Missing Dependency",
                "The 'openpyxl' library is required for Excel export.\n\n"
                "Please install it using:\npip install openpyxl",
                parent=self.root
            )
            return
        
        # Show modern export dialog with task selection
        export_dialog = tk.Toplevel(self.root)
        export_dialog.title("Export Tasks to Excel")
        export_dialog.geometry("700x650")
        export_dialog.minsize(700, 400)  # Set minimum size to show at least one item
        export_dialog.transient(self.root)
        export_dialog.grab_set()
        
        # Center the dialog
        export_dialog.update_idletasks()
        x = self.root.winfo_x() + (self.root.winfo_width() - 700) // 2
        y = self.root.winfo_y() + (self.root.winfo_height() - 650) // 2
        export_dialog.geometry(f"+{x}+{y}")
        
        # Main container
        container = ttk.Frame(export_dialog)
        container.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Header
        header_frame = tk.Frame(container, bg='#2c3e50', height=60)
        header_frame.pack(fill=tk.X, pady=(0, 10))
        header_frame.pack_propagate(False)
        
        ttk.Label(
            header_frame,
            text="📊 Select Tasks to Export",
            font=('Segoe UI', 14, 'bold'),
            background='#2c3e50',
            foreground='white'
        ).pack(pady=15)
        
        # Filter frame
        filter_frame = ttk.LabelFrame(container, text="Filter Tasks", padding=10)
        filter_frame.pack(fill=tk.X, pady=(0, 10))
        
        filter_var = tk.StringVar(value="all")
        
        ttk.Radiobutton(filter_frame, text="All Tasks", variable=filter_var, value="all").pack(side=tk.LEFT, padx=10)
        ttk.Radiobutton(filter_frame, text="Active Only", variable=filter_var, value="active").pack(side=tk.LEFT, padx=10)
        ttk.Radiobutton(filter_frame, text="Completed Only", variable=filter_var, value="completed").pack(side=tk.LEFT, padx=10)
        
        # Store task data and selection state (define early for use in functions)
        task_data = {}
        selected_tasks = set()
        result = {'proceed': False, 'tasks': []}
        
        def populate_tree(filter_type="all"):
            """Populate tree based on filter."""
            tree.delete(*tree.get_children())
            task_data.clear()
            selected_tasks.clear()
            
            # Get tasks based on filter
            if filter_type == "active":
                tasks = self.task_manager.get_filtered_tasks(show_completed=False, show_deleted=False)
            elif filter_type == "completed":
                all_tasks = self.task_manager.get_filtered_tasks(show_completed=True, show_deleted=False)
                tasks = [task for task in all_tasks if task.get('completed')]
            else:  # all
                tasks = self.task_manager.get_filtered_tasks(show_completed=True, show_deleted=False)
            
            # Sort by ID
            tasks.sort(key=lambda t: t.get('id', 0))
            
            # Populate tree
            for task in tasks:
                task_id = task.get('id')
                status = "Completed" if task.get('completed') else "Active"
                
                item_id = tree.insert("", tk.END, values=(
                    "☐",
                    task_id,
                    task.get('title', ''),
                    task.get('applied_vessel', ''),
                    status
                ))
                
                task_data[item_id] = task
                
            count_label.config(text=f"Total: {len(tasks)} tasks")
        
        def toggle_selection(event):
            """Toggle checkbox on click."""
            region = tree.identify_region(event.x, event.y)
            if region == "cell":
                item = tree.identify_row(event.y)
                if item:
                    if item in selected_tasks:
                        selected_tasks.remove(item)
                        tree.set(item, "selected", "☐")
                    else:
                        selected_tasks.add(item)
                        tree.set(item, "selected", "☑")
                    update_selection_count()
        
        def select_all():
            """Select all tasks."""
            selected_tasks.clear()
            for item in tree.get_children():
                selected_tasks.add(item)
                tree.set(item, "selected", "☑")
            update_selection_count()
        
        def deselect_all():
            """Deselect all tasks."""
            selected_tasks.clear()
            for item in tree.get_children():
                tree.set(item, "selected", "☐")
            update_selection_count()
        
        def update_selection_count():
            """Update the selection count label."""
            selection_label.config(text=f"Selected: {len(selected_tasks)} tasks")
        
        def on_filter_change(*args):
            """Handle filter change."""
            populate_tree(filter_var.get())
        
        # Bottom buttons - pack first with side=BOTTOM to anchor to bottom
        button_frame = ttk.Frame(container)
        button_frame.pack(side=tk.BOTTOM, fill=tk.X, pady=(10, 0))
        
        # Info labels - pack from bottom
        info_frame = ttk.Frame(container)
        info_frame.pack(side=tk.BOTTOM, fill=tk.X, pady=(0, 5))
        
        count_label = ttk.Label(info_frame, text="Total: 0 tasks", font=('Segoe UI', 9))
        count_label.pack(side=tk.LEFT, padx=10)
        
        selection_label = ttk.Label(info_frame, text="Selected: 0 tasks", font=('Segoe UI', 9, 'bold'))
        selection_label.pack(side=tk.LEFT, padx=10)
        
        # Selection buttons frame - pack from bottom
        selection_btn_frame = ttk.Frame(container)
        selection_btn_frame.pack(side=tk.BOTTOM, fill=tk.X, pady=(0, 5))
        
        ttk.Button(selection_btn_frame, text="Select All", command=select_all).pack(side=tk.LEFT, padx=5)
        ttk.Button(selection_btn_frame, text="Deselect All", command=deselect_all).pack(side=tk.LEFT, padx=5)
        
        # Task list frame - pack LAST so it fills remaining space
        list_frame = ttk.LabelFrame(container, text="Select Tasks", padding=10)
        list_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
        
        # Create treeview with checkboxes
        columns = ("selected", "id", "title", "vessel", "status")
        tree = ttk.Treeview(list_frame, columns=columns, show='headings', height=15)
        
        tree.heading("selected", text="☑")
        tree.heading("id", text="ID")
        tree.heading("title", text="Equipment Name")
        tree.heading("vessel", text="Applied Vessel")
        tree.heading("status", text="Status")
        
        tree.column("selected", width=40, anchor=tk.CENTER)
        tree.column("id", width=50, anchor=tk.CENTER)
        tree.column("title", width=300)
        tree.column("vessel", width=150)
        tree.column("status", width=100, anchor=tk.CENTER)
        
        # Scrollbar
        scrollbar = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Bind events
        tree.bind("<Button-1>", toggle_selection)
        filter_var.trace('w', on_filter_change)
        
        def on_export():
            if not selected_tasks:
                messagebox.showwarning("No Selection", "Please select at least one task to export.", parent=export_dialog)
                return
            
            result['proceed'] = True
            result['tasks'] = [task_data[item] for item in selected_tasks]
            export_dialog.destroy()
        
        def on_cancel():
            result['proceed'] = False
            export_dialog.destroy()
        
        ttk.Button(button_frame, text="Export Selected", command=on_export, width=15).pack(side=tk.RIGHT, padx=5)
        ttk.Button(button_frame, text="Cancel", command=on_cancel, width=15).pack(side=tk.RIGHT, padx=5)
        
        # Initial population
        populate_tree()
        
        # Wait for dialog to close
        export_dialog.wait_window()
        
        if not result['proceed']:
            return  # User cancelled
        
        tasks = result['tasks']
        
        if not tasks:
            messagebox.showwarning("No Tasks", "No tasks selected for export.", parent=self.root)
            return
        
        # Ask user where to save the file
        file_path = filedialog.asksaveasfilename(
            title="Save Excel File",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            parent=self.root
        )
        
        if not file_path:
            return  # User cancelled
        
        try:
            # Create a new workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Tasks"
            
            # Row 4: Group headers (TARGET above START and FINISH)
            ws.merge_cells('L4:M4')  # Merge for TARGET (START is col 12, FINISH is col 13)
            ws['L4'] = 'TARGET'
            ws['L4'].font = Font(bold=True)
            ws['L4'].alignment = Alignment(horizontal='center')
            
            # Row 5: Column headers
            headers = [
                'REQUEST NO.',
                'EQUIPMENT NAME',
                'REV',
                'APPLIED VESSEL',
                'DRAWING NO.',
                'REQUESTED DATE',
                'DELIVERY DATE',
                'LINK',
                '',  # Hidden column for actual LINK value
                'SDB Link',
                '',  # Hidden column for actual SDB Link value
                'START',  # TARGET START
                'FINISH',  # TARGET FINISH
                'Name',
                'TO',
                'QTD',
                'ACT',
                'STATUS'
            ]
            
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=5, column=col_idx)
                cell.value = header
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            
            # Get all non-deleted tasks
            # Tasks are already filtered based on user selection from dialog
            # Sort tasks by ID or creation date
            tasks.sort(key=lambda t: t.get('id', 0))
            
            # Export data starting from row 6
            row_idx = 6
            for task in tasks:
                # REQUEST NO.
                ws.cell(row=row_idx, column=1).value = task.get('request_no', '')
                
                # EQUIPMENT NAME
                ws.cell(row=row_idx, column=2).value = task.get('title', '')
                
                # REV
                rev_value = task.get('rev', '')
                ws.cell(row=row_idx, column=3).value = rev_value if rev_value != '' else 0
                
                # APPLIED VESSEL
                ws.cell(row=row_idx, column=4).value = task.get('applied_vessel', '')
                
                # DRAWING NO.
                ws.cell(row=row_idx, column=5).value = task.get('drawing_no', '')
                
                # REQUESTED DATE
                requested_date = task.get('requested_date')
                if requested_date:
                    ws.cell(row=row_idx, column=6).value = requested_date
                
                # DELIVERY DATE (DUE DATE)
                due_date = task.get('due_date')
                if due_date:
                    ws.cell(row=row_idx, column=7).value = due_date
                
                # LINK - Display "LINK" text, actual value in next column
                link_value = task.get('link', '')
                if link_value:
                    link_cell = ws.cell(row=row_idx, column=8)
                    link_cell.value = 'LINK'
                    link_cell.hyperlink = link_value
                    link_cell.font = Font(color="0563C1", underline="single")  # Blue hyperlink style
                ws.cell(row=row_idx, column=9).value = link_value  # Actual value in hidden column
                
                # SDB Link - Display text, actual value in next column
                sdb_link_value = task.get('sdb_link', '')
                if sdb_link_value:
                    sdb_link_cell = ws.cell(row=row_idx, column=10)
                    sdb_link_cell.value = 'SDB Link'
                    sdb_link_cell.hyperlink = sdb_link_value
                    sdb_link_cell.font = Font(color="0563C1", underline="single")  # Blue hyperlink style
                ws.cell(row=row_idx, column=11).value = sdb_link_value  # Actual value in hidden column
                
                # TARGET START
                target_start = task.get('target_start')
                if target_start:
                    ws.cell(row=row_idx, column=12).value = target_start
                
                # TARGET FINISH
                target_finish = task.get('target_finish')
                if target_finish:
                    ws.cell(row=row_idx, column=13).value = target_finish
                
                # NAME (MAIN STAFF)
                ws.cell(row=row_idx, column=14).value = task.get('main_staff', '')
                
                # TO (ASSIGNED TO)
                ws.cell(row=row_idx, column=15).value = task.get('assigned_to', '')
                
                # QTD (QTD MHR)
                qtd_mhr = task.get('qtd_mhr', 0)
                ws.cell(row=row_idx, column=16).value = qtd_mhr if qtd_mhr is not None else 0
                
                # ACT (ACTUAL MHR)
                actual_mhr = task.get('actual_mhr', 0)
                ws.cell(row=row_idx, column=17).value = actual_mhr if actual_mhr is not None else 0
                
                # STATUS
                status = 'COMPLETED' if task.get('completed') else 'ACTIVE'
                ws.cell(row=row_idx, column=18).value = status
                
                row_idx += 1
            
            # Auto-size columns
            for col_idx in range(1, 19):
                column_letter = get_column_letter(col_idx)
                max_length = 0
                for cell in ws[column_letter]:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except (TypeError, AttributeError) as e:
                        logging.debug(f"Cell value error in column {column_letter}: {e}")
                adjusted_width = min(max_length + 2, 50)  # Cap at 50
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Hide the actual link value columns (9 and 11)
            ws.column_dimensions[get_column_letter(9)].hidden = True
            ws.column_dimensions[get_column_letter(11)].hidden = True
            
            # Save the workbook
            wb.save(file_path)
            
            messagebox.showinfo(
                "Export Complete",
                f"Successfully exported {len(tasks)} tasks to:\n{file_path}",
                parent=self.root
            )
            
        except Exception as e:
            logging.error(f"Error during Excel export: {str(e)}", exc_info=True)
            messagebox.showerror(
                "Export Error",
                f"An error occurred during export:\n\n{str(e)}",
                parent=self.root
            )
    
    def open_import(self):
        """Import tasks from Excel file."""
        try:
            # Import openpyxl when needed
            import openpyxl
            from openpyxl.cell.cell import TYPE_FORMULA, TYPE_STRING
        except ImportError:
            messagebox.showerror(
                "Missing Dependency",
                "The 'openpyxl' library is required for Excel import.\n\n"
                "Please install it using:\npip install openpyxl",
                parent=self.root
            )
            return
        
        # Open file dialog to select Excel file
        file_path = filedialog.askopenfilename(
            title="Select Excel File to Import",
            filetypes=[("Excel files", "*.xlsx *.xls *.xlsm *.xlsb"), ("All files", "*.*")],
            parent=self.root
        )
        
        if not file_path:
            return  # User cancelled
        
        # Validate file extension
        if not file_path.lower().endswith(('.xlsx', '.xls', '.xlsm', '.xlsb')):
            messagebox.showerror(
                "Invalid File",
                "Please select an Excel file (.xlsx, .xls, .xlsm, or .xlsb)",
                parent=self.root
            )
            return
        
        try:
            # Load workbook with data_only=True to get calculated values instead of formulas
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb.active
            
            # Headers are in row 5 (1-indexed in Excel, so row 5)
            header_row = 5
            
            # Read headers from row 5, looking at row 4 for context (merged cells)
            headers = {}
            current_group_context = "" # Tracks context from Row 4 (e.g., "TARGET", "ACTUAL")
            
            logging.info("=== EXCEL IMPORT: Reading Headers ===")

            for col_idx, cell in enumerate(ws[header_row], start=1):
                # --- Row 4 Context Check ---
                r4_cell = ws.cell(row=4, column=col_idx)
                if r4_cell.value:
                    # Found a new group label like "TARGET" or "ACTUAL"
                    current_group_context = str(r4_cell.value).strip().upper()
                    logging.info(f"Col {col_idx}: Found Row 4 context = '{current_group_context}'")
                
                # --- Row 5 Header ---
                if cell.value:
                    h_val = str(cell.value).strip()
                    
                    # 1. Primary Fix: Store exact match ONLY if not already present.
                    # This effectively ignores the second "START" (Actuals) if the first "START" (Targets) was already found.
                    if h_val not in headers:
                        headers[h_val] = col_idx
                        logging.info(f"Col {col_idx}: Mapped header '{h_val}' -> column {col_idx}")
                    else:
                        logging.info(f"Col {col_idx}: Skipped duplicate header '{h_val}' (already mapped to column {headers[h_val]})")
                    
                    # 2. Contextual Fix: Store "TARGET START" or "ACTUAL START" pointing to specific columns
                    # This allows precise mapping if we want it.
                    if current_group_context:
                        compound_header = f"{current_group_context} {h_val}" # e.g., "TARGET START"
                        if compound_header not in headers:
                             headers[compound_header] = col_idx
                             logging.info(f"Col {col_idx}: Created compound header '{compound_header}' -> column {col_idx}")
            
            logging.info(f"Final headers dictionary: {headers}")

            # Define possible column headers for each field (list of aliases)
            # This allows flexibility in Excel header names
            field_definitions = {
                'request_no': ['REQUEST NO.', 'REQUEST NO', 'Request No', 'REQ NO'],
                'title': ['EQUIPMENT NAME', 'TITLE', 'Equipment Name', 'DESCRIPTION'],
                'applied_vessel': ['APPLIED VESSEL', 'Applied Vessel', 'VESSEL'],
                'rev': ['REV', 'Rev', 'REVISION'],
                'drawing_no': ['DRAWING NO.', 'DRAWING NO', 'Drawing No', 'DWG NO'],
                'requested_date': ['REQUESTED DATE', 'Requested Date', 'REQ DATE'],
                'due_date': ['DELIVERY DATE', 'Delivery Date', 'DUE DATE', 'DEADLINE'],
                'target_start': ['TARGET START', 'START', 'Start', 'Target Start'], # Prioritize compound key
                'target_finish': ['TARGET FINISH', 'FINISH', 'Finish', 'Target Finish'], # Prioritize compound key
                'link': ['LINK', 'Link', 'FOLDER PATH'],
                'sdb_link': ['SDB Link', 'SDB URL', 'NOTE', 'Note'],
                'main_staff': ['Name', 'NAME', 'Main Staff', 'MAIN STAFF'],
                'assigned_to': ['TO', 'To', 'ASSIGNED TO'],
                'qtd_mhr': ['QTD', 'Qtd', 'QTD MHR', 'BUDGET MHR'],
                'actual_mhr': ['ACT', 'Act', 'ACTUAL MHR', 'USED MHR'],
                'status': ['STATUS', 'Status', 'COMPLETION STATUS']
            }
            
            # Build active mapping based on headers found in the file
            column_mapping = {}
            for db_field, aliases in field_definitions.items():
                found = False
                for alias in aliases:
                    if alias in headers:
                        column_mapping[alias] = db_field
                        logging.info(f"Mapped Excel column '{alias}' (col {headers[alias]}) -> DB field '{db_field}'")
                        found = True
                        break # Only map the first matching alias found
                
                if not found:
                    logging.warning(f"No matching Excel column found for DB field '{db_field}'")
            
            logging.info(f"=== Final column_mapping: {column_mapping} ===")
            
            # Check for critical missing columns only (optional)
            # Only warn if 'title' (Equipment Name) is missing as it's essential
            if 'title' not in column_mapping.values():
                 messagebox.showwarning(
                    "Missing Equipment Name",
                    "Could not find an 'EQUIPMENT NAME' or 'TITLE' column.\n"
                    "Some tasks may be skipped.",
                    parent=self.root
                )

            # Previously checked for ALL missing columns, which caused false alarms with aliases.
            # We now rely on whatever we found.

            
            # Import data starting from row 6 (row after headers)
            imported_count = 0
            failed_count = 0
            errors = []
            
            # Create progress dialog
            progress_dialog = tk.Toplevel(self.root)
            progress_dialog.title("Importing Tasks")
            progress_dialog.geometry("400x120")
            progress_dialog.transient(self.root)
            progress_dialog.grab_set()
            
            # Center the dialog
            progress_dialog.update_idletasks()
            x = self.root.winfo_x() + (self.root.winfo_width() // 2) - (progress_dialog.winfo_width() // 2)
            y = self.root.winfo_y() + (self.root.winfo_height() // 2) - (progress_dialog.winfo_height() // 2)
            progress_dialog.geometry(f"+{x}+{y}")
            
            # Progress label
            progress_label = tk.Label(progress_dialog, text="Preparing import...", font=('Arial', 10))
            progress_label.pack(pady=(20, 10))
            
            # Progress bar with green color
            style = ttk.Style()
            style.theme_use('default')
            style.configure("green.Horizontal.TProgressbar", 
                          troughcolor='white',
                          background='#4CAF50',  # Green color
                          bordercolor='#d0d0d0',
                          lightcolor='#4CAF50',
                          darkcolor='#4CAF50')
            
            progress_bar = ttk.Progressbar(
                progress_dialog, 
                length=350, 
                mode='determinate',
                style="green.Horizontal.TProgressbar"
            )
            progress_bar.pack(pady=10)
            
            # Calculate total rows to process
            total_rows = ws.max_row - header_row
            progress_bar['maximum'] = total_rows
            
            progress_dialog.update()
            
            for row_idx in range(header_row + 1, ws.max_row + 1):
                # Update progress
                current_row = row_idx - header_row
                percentage = int((current_row / total_rows) * 100)
                progress_bar['value'] = current_row
                progress_label['text'] = f"Imported {imported_count} tasks (Processing row {current_row} of {total_rows} - {percentage}%)"
                progress_dialog.update()
                
                # Skip the header row itself if somehow included
                if row_idx == header_row:
                    continue
                    
                try:
                    # Extract data from row
                    task_data = {}
                    
                    logging.info(f"--- Processing Row {row_idx} ---")
                    
                    for excel_col, db_field in column_mapping.items():
                        if excel_col in headers:
                            col_idx = headers[excel_col]
                            cell = ws.cell(row=row_idx, column=col_idx)
                            
                            # Handle LINK and SDB Link - read from next column (actual hidden column)
                            if db_field in ['link', 'sdb_link']:
                                # The actual link value is in the next column (col_idx + 1)
                                next_col_idx = col_idx + 1
                                actual_cell = ws.cell(row=row_idx, column=next_col_idx)
                                task_data[db_field] = str(actual_cell.value) if actual_cell.value is not None else ""
                                logging.info(f"  {db_field}: Read from next column {next_col_idx} -> '{task_data[db_field]}'")
                            # Handle date cells
                            elif db_field in ['requested_date', 'due_date', 'target_start', 'target_finish']:
                                if cell.value:
                                    logging.info(f"  {db_field}: Raw cell value = '{cell.value}' (type: {type(cell.value).__name__})")
                                    # Handle datetime/date objects directly
                                    if isinstance(cell.value, datetime.datetime):
                                        task_data[db_field] = cell.value.strftime('%Y-%m-%d')
                                        logging.info(f"  {db_field}: Converted datetime -> '{task_data[db_field]}'")
                                    elif isinstance(cell.value, datetime.date):
                                        task_data[db_field] = cell.value.strftime('%Y-%m-%d')
                                        logging.info(f"  {db_field}: Converted date -> '{task_data[db_field]}'")
                                    else:
                                        # Parse string dates - try multiple formats
                                        date_str = str(cell.value).strip()
                                        
                                        # Remove time portion if present (e.g., "1/15/2026 00:00:00")
                                        if ' ' in date_str:
                                            date_str = date_str.split(' ')[0]
                                        
                                        parsed_date = None
                                        # Try mm/dd/yyyy format first (your Excel format)
                                        for date_format in ['%m/%d/%Y', '%m/%d/%y', '%Y-%m-%d', '%d/%m/%Y', '%Y/%m/%d', '%d-%b-%y', '%d-%b-%Y', '%Y.%m.%d']:
                                            try:
                                                parsed_date = datetime.datetime.strptime(date_str, date_format)
                                                break
                                            except (ValueError, TypeError):
                                                continue
                                        
                                        if parsed_date:
                                            task_data[db_field] = parsed_date.strftime('%Y-%m-%d')
                                            logging.info(f"  {db_field}: Parsed '{cell.value}' -> '{task_data[db_field]}'")
                                        else:
                                            task_data[db_field] = None
                                            logging.warning(f"  {db_field}: Could not parse date '{cell.value}' in row {row_idx}, column {excel_col}")
                                else:
                                    task_data[db_field] = None
                                    logging.info(f"  {db_field}: Empty cell, set to None")
                            # Handle numeric fields
                            elif db_field in ['qtd_mhr', 'actual_mhr']:
                                if cell.value is not None:
                                    try:
                                        task_data[db_field] = int(cell.value)
                                    except (ValueError, TypeError) as e:
                                        logging.debug(f"Invalid number for {db_field}: {cell.value}")
                                        task_data[db_field] = 0
                                else:
                                    task_data[db_field] = 0
                            else:
                                # Ensure we preserve '0' or other falsy values that are not None
                                task_data[db_field] = str(cell.value) if cell.value is not None else ""
                                # Log status field for debugging
                                if db_field == 'status':
                                    logging.info(f"  {db_field}: Read value = '{cell.value}' -> '{task_data[db_field]}'")
                    
                    # Skip empty rows (no title)
                    if not task_data.get('title') or task_data.get('title') == 'None':
                        continue
                    
                    # Skip header row if it appears in data (case-insensitive check)
                    title_upper = str(task_data.get('title', '')).strip().upper()
                    if title_upper in ['EQUIPMENT NAME', 'TITLE', '']:
                        continue
                    
                    # Set default values for missing fields
                    task_data.setdefault('description', '')
                    task_data.setdefault('priority', 'medium')
                    task_data.setdefault('category', 'general')
                    
                    # Check if task is completed based on STATUS column
                    is_completed = False
                    status_value = task_data.get('status', '').strip().upper()
                    logging.info(f"  Row {row_idx}: Checking STATUS field - Raw: '{task_data.get('status', '')}', Cleaned: '{status_value}'")
                    if status_value == 'COMPLETED':
                        is_completed = True
                        logging.info(f"  Row {row_idx}: ✓ Task will be marked as COMPLETED")
                    
                    logging.info(f"  Final task_data for row {row_idx}:")
                    logging.info(f"    title: {task_data.get('title')}")
                    logging.info(f"    target_start: {task_data.get('target_start')}")
                    logging.info(f"    target_finish: {task_data.get('target_finish')}")
                    logging.info(f"    rev: {task_data.get('rev')}")
                    logging.info(f"    completed: {is_completed}")
                    
                    # Add task to database
                    if self.task_manager:
                        result = self.task_manager.add_task(
                            title=task_data.get('title', ''),
                            description=task_data.get('description', ''),
                            due_date=task_data.get('due_date'),
                            priority=task_data.get('priority', 'medium'),
                            category=task_data.get('category', 'general'),
                            main_staff=task_data.get('main_staff'),
                            assigned_to=task_data.get('assigned_to'),
                            applied_vessel=task_data.get('applied_vessel', ''),
                            rev=task_data.get('rev', ''),
                            drawing_no=task_data.get('drawing_no', ''),
                            link=task_data.get('link', ''),
                            sdb_link=task_data.get('sdb_link', ''),
                            request_no=task_data.get('request_no', ''),
                            requested_date=task_data.get('requested_date'),
                            target_start=task_data.get('target_start'),
                            target_finish=task_data.get('target_finish'),
                            qtd_mhr=task_data.get('qtd_mhr', 0),
                            actual_mhr=task_data.get('actual_mhr', 0)
                        )
                        
                        # If task was added successfully and should be marked completed, update it
                        if result:
                            imported_count += 1
                            if is_completed:
                                logging.info(f"  Updating task {result} to mark as completed...")
                                update_success = self.task_manager.update_task(result, completed=True)
                                if update_success:
                                    logging.info(f"  ✓ Task {result} successfully marked as completed")
                                else:
                                    logging.error(f"  ✗ Failed to mark task {result} as completed")
                        else:
                            failed_count += 1
                            errors.append(f"Row {row_idx}: Failed to add task '{task_data.get('title', 'Unknown')}'")
                    
                except Exception as e:
                    failed_count += 1
                    errors.append(f"Row {row_idx}: {str(e)}")
                    logging.error(f"Error importing row {row_idx}: {str(e)}", exc_info=True)
            
            # Close workbook
            wb.close()
            
            # Close progress dialog
            progress_dialog.destroy()
            
            # Show modern summary dialog
            summary_dialog = tk.Toplevel(self.root)
            summary_dialog.title("Import Complete")
            summary_dialog.withdraw()  # Hide initially to prevent flash
            summary_dialog.geometry("450x300" if failed_count > 0 and errors else "450x200")
            summary_dialog.transient(self.root)
            summary_dialog.configure(bg='white')
            
            # Center the dialog
            summary_dialog.update_idletasks()
            x = self.root.winfo_x() + (self.root.winfo_width() // 2) - (summary_dialog.winfo_width() // 2)
            y = self.root.winfo_y() + (self.root.winfo_height() // 2) - (summary_dialog.winfo_height() // 2)
            summary_dialog.geometry(f"+{x}+{y}")
            
            # Header with icon
            header_frame = tk.Frame(summary_dialog, bg='#4CAF50', height=60)
            header_frame.pack(fill=tk.X)
            header_frame.pack_propagate(False)
            
            header_label = tk.Label(
                header_frame, 
                text="✓ Import Complete",
                font=('Arial', 16, 'bold'),
                bg='#4CAF50',
                fg='white'
            )
            header_label.pack(expand=True)
            
            # Content frame
            content_frame = tk.Frame(summary_dialog, bg='white')
            content_frame.pack(fill=tk.BOTH, expand=True, padx=30, pady=20)
            
            # Success count
            success_frame = tk.Frame(content_frame, bg='white')
            success_frame.pack(fill=tk.X, pady=5)
            
            tk.Label(
                success_frame,
                text="Successfully Imported:",
                font=('Arial', 11),
                bg='white',
                fg='#333'
            ).pack(side=tk.LEFT)
            
            tk.Label(
                success_frame,
                text=f"{imported_count} tasks",
                font=('Arial', 11, 'bold'),
                bg='white',
                fg='#4CAF50'
            ).pack(side=tk.RIGHT)
            
            # Failed count (if any)
            if failed_count > 0:
                fail_frame = tk.Frame(content_frame, bg='white')
                fail_frame.pack(fill=tk.X, pady=5)
                
                tk.Label(
                    fail_frame,
                    text="Failed:",
                    font=('Arial', 11),
                    bg='white',
                    fg='#333'
                ).pack(side=tk.LEFT)
                
                tk.Label(
                    fail_frame,
                    text=f"{failed_count} tasks",
                    font=('Arial', 11, 'bold'),
                    bg='white',
                    fg='#f44336'
                ).pack(side=tk.RIGHT)
                
                # Error details (if any)
                if errors:
                    tk.Label(
                        content_frame,
                        text="Error Details:",
                        font=('Arial', 10, 'bold'),
                        bg='white',
                        fg='#333'
                    ).pack(anchor=tk.W, pady=(10, 5))
                    
                    error_text = tk.Text(
                        content_frame,
                        height=5,
                        font=('Courier', 9),
                        bg='#f5f5f5',
                        relief=tk.FLAT,
                        wrap=tk.WORD
                    )
                    error_text.pack(fill=tk.BOTH, expand=True)
                    error_msg = "\n".join(errors[:10])
                    if len(errors) > 10:
                        error_msg += f"\n... and {len(errors) - 10} more errors"
                    error_text.insert('1.0', error_msg)
                    error_text.config(state=tk.DISABLED)
            
            # Button frame
            button_frame = tk.Frame(summary_dialog, bg='white')
            button_frame.pack(fill=tk.X, padx=30, pady=(0, 20))
            
            ok_button = tk.Button(
                button_frame,
                text="OK",
                command=summary_dialog.destroy,
                font=('Arial', 10, 'bold'),
                bg='#4CAF50',
                fg='white',
                activebackground='#45a049',
                activeforeground='white',
                relief=tk.FLAT,
                cursor='hand2',
                width=15,
                height=1
            )
            ok_button.pack(side=tk.RIGHT)
            
            # Bind Enter key to close
            summary_dialog.bind('<Return>', lambda e: summary_dialog.destroy())
            summary_dialog.bind('<Escape>', lambda e: summary_dialog.destroy())
            
            # Show dialog after everything is configured
            summary_dialog.deiconify()
            summary_dialog.grab_set()
            
            # Reload tasks from database to ensure cache is updated
            if self.task_manager:
                self.task_manager.reload_tasks()
            
            # Refresh task manager if it's open
            if self.task_manager_app:
                self.task_manager_app.refresh_task_list()
            
        except Exception as e:
            logging.error(f"Error during Excel import: {str(e)}", exc_info=True)
            messagebox.showerror(
                "Import Error",
                f"An error occurred during import:\n\n{str(e)}",
                parent=self.root
            )
    
    def on_close(self):
        """Handle window close event."""
        logging.info("Application closing...")
        
        # Cleanup task manager
        if self.task_manager_app:
            try:
                self.task_manager_app.on_close()
            except Exception as e:
                logging.warning(f"Error during app cleanup: {e}")
        
        if self.task_manager:
            try:
                # Close user session first
                self.task_manager.close_user_session()
                
                if self.task_manager.conn:
                    self.task_manager.conn.close()
                    logging.info("Database connection closed properly")
            except Exception as e:
                logging.warning(f"Error closing database: {e}")
        
        logging.info("Flushing log handlers before exit...")
        for handler in logging.getLogger().handlers:
            try:
                handler.flush()
                handler.close()
            except Exception as e:
                logging.warning(f"Error closing handler: {e}")
        
        self.root.quit()
        self.root.destroy()
    
    def run(self):
        """Start the application event loop."""
        self.root.mainloop()


def parse_version(version_str):
    """Parse version string into tuple of integers for comparison."""
    try:
        # Remove 'v' prefix if present and split by '.'
        version_str = version_str.lstrip('v').strip()
        return tuple(map(int, version_str.split('.')))
    except (ValueError, AttributeError):
        return (0, 0, 0)

def check_for_updates():
    """Check for updates and auto-update if newer version is available."""
    try:
        # Current version - loaded from config
        CURRENT_VERSION = app_config.get('Application', 'version', fallback='0.23.0')
        
        # Update server configuration - loaded from config
        UPDATE_SERVER = app_config.get('Update', 'server', fallback=r'\\10.195.103.198\shared\TaskManager\updates')
        VERSION_FILE = app_config.get('Update', 'version_file', fallback='version.json')
        
        logging.info(f"Checking for updates... Current version: {CURRENT_VERSION}")
        logging.info(f"Update server: {UPDATE_SERVER}")
        
        # Determine if using network path or HTTP
        is_network = UPDATE_SERVER.startswith("\\\\") or UPDATE_SERVER.startswith("//")
        is_http = UPDATE_SERVER.startswith("http")
        
        if is_network:
            # Network share approach
            version_file_path = os.path.join(UPDATE_SERVER, VERSION_FILE)
            
            # Check if update server is accessible
            if not os.path.exists(UPDATE_SERVER):
                logging.info(f"Update server not accessible: {UPDATE_SERVER}")
                return False
            
            # Read version file
            if not os.path.exists(version_file_path):
                logging.info("Version file not found on update server")
                return False
            
            with open(version_file_path, 'r') as f:
                version_info = json.load(f)
        
        elif is_http:
            # HTTP server approach
            import urllib.request
            version_url = f"{UPDATE_SERVER}/{VERSION_FILE}"
            
            try:
                with urllib.request.urlopen(version_url, timeout=5) as response:
                    version_info = json.loads(response.read().decode())
            except Exception as e:
                logging.info(f"Could not reach update server: {e}")
                return False
        else:
            logging.warning("Invalid UPDATE_SERVER configuration")
            return False
        
        # Compare versions
        latest_version = version_info.get("version", "0.0.0")
        update_url = version_info.get("download_url", "")
        release_notes = version_info.get("release_notes", "Bug fixes and improvements")
        
        logging.info(f"Latest version: {latest_version}")
        
        # Proper semantic version comparison
        current_ver_tuple = parse_version(CURRENT_VERSION)
        latest_ver_tuple = parse_version(latest_version)
        
        logging.info(f"Version comparison: Current {current_ver_tuple} vs Latest {latest_ver_tuple}")
        
        if latest_ver_tuple <= current_ver_tuple:
            logging.info("Already running the latest version")
            return False
        
        # Show update dialog
        root = tk.Tk()
        root.withdraw()
        
        update_msg = (
            f"A new update is available!\n\n"
            f"Current Version: {CURRENT_VERSION}\n"
            f"Latest Version: {latest_version}\n\n"
            f"What's New:\n{release_notes}\n\n"
            f"Download and install update now?"
        )
        
        if not messagebox.askyesno("Update Available", update_msg, icon='info'):
            logging.info("User declined update")
            root.destroy()
            return False
        
        # Download and install update
        try:
            import tempfile
            
            # Get current executable path and folder
            if getattr(sys, 'frozen', False):
                # CRITICAL FIX for Nuitka --onefile:
                # sys.executable returns the TEMP extraction path (e.g., C:\Users\...\Temp\onefile_XXX\...)
                # sys.argv[0] preserves the ORIGINAL launch path (e.g., C:\Program Files\TaskManager\TaskManager.exe)
                # This ensures the update replaces the actual exe, not the temporary extracted copy
                
                original_path = os.path.abspath(sys.argv[0])
                temp_path = sys.executable
                
                # Use original path if it's a valid exe file, otherwise fallback to sys.executable
                if original_path.lower().endswith('.exe') and os.path.isfile(original_path):
                    current_exe = original_path
                    logging.info(f"Using original exe path: {current_exe}")
                else:
                    current_exe = temp_path
                    logging.info(f"Fallback to sys.executable: {current_exe}")
                
                # Log both paths for diagnostics
                logging.info(f"Original launch path (sys.argv[0]): {original_path}")
                logging.info(f"Runtime executable path (sys.executable): {temp_path}")
                
                app_dir = os.path.dirname(current_exe)  # The folder containing the exe
            else:
                logging.info("Not running as executable, skipping update")
                root.destroy()
                return False
            
            # Download new version
            if is_network:
                # For single-file exe (Nuitka --onefile)
                new_exe_path = os.path.join(UPDATE_SERVER, update_url)
                if not os.path.exists(new_exe_path):
                    raise FileNotFoundError(f"Update file not found: {new_exe_path}")
            else:
                # HTTP download
                new_exe_url = f"{UPDATE_SERVER}/{update_url}"
                new_exe_path = os.path.join(tempfile.gettempdir(), "TaskManager_new.exe")
                
                import urllib.request
                urllib.request.urlretrieve(new_exe_url, new_exe_path)
            
            # Create update script
            update_script = os.path.join(tempfile.gettempdir(), "update_taskmanager.bat")
            
            # Get just the filename of the current exe
            current_exe_name = os.path.basename(current_exe)
            
            # Escape backslashes for batch script paths
            new_exe_path_escaped = new_exe_path.replace('\\', '\\\\')
            current_exe_escaped = current_exe.replace('\\', '\\\\')
            
            with open(update_script, 'w') as f:
                # Enhanced update script with retry logic and better error handling
                f.write(f"""@echo off
echo ============================================
echo TaskManager Auto-Update
echo ============================================
echo.
echo Waiting for application to close...
timeout /t 3 /nobreak >nul

REM Kill any running instances - retry up to 3 times
set RETRY=0
:KILL_LOOP
set /a RETRY+=1
taskkill /F /IM "{current_exe_name}" >nul 2>&1
if %RETRY% LSS 3 (
    timeout /t 1 /nobreak >nul
    tasklist /FI "IMAGENAME eq {current_exe_name}" 2>NUL | find /I /N "{current_exe_name}">NUL
    if %ERRORLEVEL% EQU 0 goto KILL_LOOP
)

echo Waiting for file handles to release...
timeout /t 2 /nobreak >nul

REM Verify source file exists
if not exist "{new_exe_path}" (
    echo ERROR: Source file not found!
    echo Path: {new_exe_path}
    pause
    del "%~f0"
    exit /b 1
)

REM Backup old version just in case
echo Creating backup...
if exist "{current_exe}" (
    copy /Y "{current_exe}" "{current_exe}.backup" >nul 2>&1
)

REM Copy the new version with retry
echo Copying new version...
set COPY_RETRY=0
:COPY_LOOP
set /a COPY_RETRY+=1
copy /Y "{new_exe_path}" "{current_exe}" >nul 2>&1
if %ERRORLEVEL% EQU 0 goto COPY_SUCCESS
if %COPY_RETRY% LSS 5 (
    echo Retry %COPY_RETRY%/5...
    timeout /t 2 /nobreak >nul
    goto COPY_LOOP
)

REM Copy failed
echo.
echo ============================================
echo ERROR: Update failed after 5 attempts!
echo ============================================
echo Error code: %ERRORLEVEL%
echo Source: {new_exe_path}
echo Destination: {current_exe}
echo.
echo Possible causes:
echo - File is locked by another process
echo - Insufficient permissions (try running as administrator)
echo - Antivirus blocking the update
echo.
echo Restoring backup...
if exist "{current_exe}.backup" (
    copy /Y "{current_exe}.backup" "{current_exe}" >nul 2>&1
    del "{current_exe}.backup" >nul 2>&1
)
pause
del "%~f0"
exit /b 1

:COPY_SUCCESS
echo.
echo ============================================
echo Update successful!
echo ============================================
echo Cleaning up...
if exist "{current_exe}.backup" del "{current_exe}.backup" >nul 2>&1
echo Restarting application...
timeout /t 1 /nobreak >nul
start "" "{current_exe}"
del "%~f0"
exit /b 0
""")
            
            logging.info(f"Update script created: {update_script}")
            logging.info(f"Will copy from: {new_exe_path}")
            logging.info(f"Will copy to: {current_exe}")
            
            messagebox.showinfo(
                "Installing Update",
                "The application will now close and restart to complete the update.\n\nThis may take a few seconds."
            )
            
            # Close any open windows properly
            root.destroy()
            
            # Launch update script with visible window for debugging
            # Using shell=True and start command to ensure it runs independently
            subprocess.Popen(f'cmd /c start "TaskManager Update" /wait "{update_script}"', shell=True)
            
            # Small delay to ensure subprocess starts
            import time
            time.sleep(0.5)
            
            # Force exit
            os._exit(0)
            
        except Exception as e:
            logging.error(f"Update failed: {e}", exc_info=True)
            messagebox.showerror(
                "Update Failed",
                f"Failed to install update:\n{str(e)}\n\nPlease contact IT support."
            )
            root.destroy()
            return False
    
    except Exception as e:
        logging.error(f"Error checking for updates: {e}", exc_info=True)
        return False


# ============================================
# ENHANCED FEATURES UI METHODS
# ============================================

# Add these methods to TaskManagerApp class (they reference self.task_manager)
def start_timer(self, task_id):
    """Start timer for a task."""
    if not hasattr(self.task_manager, 'time_tracking') or not self.task_manager.time_tracking:
        messagebox.showinfo("Feature Not Available", "Time tracking feature is not enabled.\n\nRun 'sql/create_enhanced_features.sql' first!")
        return
    
    current_user = self.task_manager._get_current_user()
    if self.task_manager.time_tracking.start_timer(task_id, current_user):
        messagebox.showinfo("Timer Started", f"Timer started for task {task_id}.\n\nTime will be logged when you stop the timer.")
        self.refresh_task_list()
    else:
        messagebox.showerror("Error", "Could not start timer. Check if timer is already running.")

def stop_timer(self, task_id):
    """Stop timer for a task."""
    if not hasattr(self.task_manager, 'time_tracking') or not self.task_manager.time_tracking:
        return
    
    duration_minutes = self.task_manager.time_tracking.stop_timer(task_id)
    if duration_minutes is not None:
        hours = duration_minutes // 60
        minutes = duration_minutes % 60
        messagebox.showinfo("Timer Stopped", f"Timer stopped!\n\nTime logged: {hours}h {minutes}m ({duration_minutes} minutes)")
        self.refresh_task_list()
    else:
        messagebox.showerror("Error", "Could not stop timer.")

def view_time_log(self, task_id):
    """View time log for a task."""
    if not hasattr(self.task_manager, 'time_tracking') or not self.task_manager.time_tracking:
        return
    
    entries = self.task_manager.time_tracking.get_time_entries(task_id)
    total_minutes = self.task_manager.time_tracking.get_total_time_logged(task_id)
    
    # Create dialog
    dialog = tk.Toplevel(self.root)
    dialog.title(f"Time Log - Task {task_id}")
    dialog.geometry("600x400")
    dialog.transient(self.root)
    
    # Header
    header_frame = tk.Frame(dialog, bg='#2c3e50', height=50)
    header_frame.pack(fill=tk.X)
    header_frame.pack_propagate(False)
    
    tk.Label(header_frame, text=f"Total Time: {total_minutes // 60}h {total_minutes % 60}m", 
             font=('Arial', 12, 'bold'), bg='#2c3e50', fg='white').pack(expand=True)
    
    # Time entries list
    list_frame = ttk.Frame(dialog)
    list_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
    
    tree = ttk.Treeview(list_frame, columns=("date", "user", "duration", "type"), show='headings')
    tree.heading("date", text="Date")
    tree.heading("user", text="User")
    tree.heading("duration", text="Duration")
    tree.heading("type", text="Type")
    
    tree.column("date", width=150)
    tree.column("user", width=100)
    tree.column("duration", width=100)
    tree.column("type", width=80)
    
    for entry in entries:
        date_str = entry['start_time'].strftime('%Y-%m-%d %H:%M') if entry['start_time'] else "N/A"
        duration_str = f"{entry['duration_minutes'] // 60}h {entry['duration_minutes'] % 60}m"
        tree.insert("", tk.END, values=(date_str, entry['user_name'], duration_str, entry['entry_type']))
    
    tree.pack(fill=tk.BOTH, expand=True)
    
    ttk.Button(dialog, text="Close", command=dialog.destroy).pack(pady=10)

def manage_subtasks(self, task_id):
    """Open subtask management dialog."""
    if not hasattr(self.task_manager, 'subtasks') or not self.task_manager.subtasks:
        messagebox.showinfo("Feature Not Available", "Subtasks feature is not enabled.\n\nRun 'sql/create_enhanced_features.sql' first!")
        return
    
    # Create dialog
    dialog = tk.Toplevel(self.root)
    dialog.title(f"Subtasks - Task {task_id}")
    dialog.geometry("700x500")
    dialog.transient(self.root)
    
    # Header
    task = self.task_manager._find_task_by_id(task_id)
    header_frame = tk.Frame(dialog, bg='#2c3e50', height=60)
    header_frame.pack(fill=tk.X)
    header_frame.pack_propagate(False)
    
    tk.Label(header_frame, text=task['title'] if task else f"Task {task_id}", 
             font=('Arial', 12, 'bold'), bg='#2c3e50', fg='white').pack(pady=5)
    
    completion_pct = self.task_manager.subtasks.get_completion_percentage(task_id)
    tk.Label(header_frame, text=f"Completion: {completion_pct}%", 
             font=('Arial', 10), bg='#2c3e50', fg='#3498db').pack()
    
    # Subtasks list
    list_frame = ttk.Frame(dialog)
    list_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
    
    tree = ttk.Treeview(list_frame, columns=("status", "title", "assigned"), show='headings')
    tree.heading("status", text="✓")
    tree.heading("title", text="Subtask")
    tree.heading("assigned", text="Assigned To")
    
    tree.column("status", width=40)
    tree.column("title", width=400)
    tree.column("assigned", width=120)
    
    def refresh_subtasks():
        tree.delete(*tree.get_children())
        subtasks = self.task_manager.subtasks.get_subtasks(task_id)
        for subtask in subtasks:
            status = "☑" if subtask['completed'] else "☐"
            tree.insert("", tk.END, values=(status, subtask['title'], subtask['assigned_to'] or ""), 
                       tags=(subtask['subtask_id'],))
    
    def toggle_subtask(event):
        item = tree.selection()
        if item:
            subtask_id = tree.item(item[0])['tags'][0]
            self.task_manager.subtasks.toggle_subtask(subtask_id)
            refresh_subtasks()
            # Update parent task display
            self.refresh_task_list()
    
    def add_subtask():
        title = simpledialog.askstring("Add Subtask", "Enter subtask title:", parent=dialog)
        if title:
            self.task_manager.subtasks.add_subtask(task_id, title)
            refresh_subtasks()
            self.refresh_task_list()
    
    def delete_subtask():
        item = tree.selection()
        if item:
            subtask_id = tree.item(item[0])['tags'][0]
            if messagebox.askyesno("Confirm", "Delete this subtask?", parent=dialog):
                self.task_manager.subtasks.delete_subtask(subtask_id)
                refresh_subtasks()
                self.refresh_task_list()
    
    tree.bind("<Double-1>", toggle_subtask)
    tree.pack(fill=tk.BOTH, expand=True)
    
    # Buttons
    btn_frame = ttk.Frame(dialog)
    btn_frame.pack(fill=tk.X, padx=10, pady=10)
    
    ttk.Button(btn_frame, text="Add Subtask", command=add_subtask).pack(side=tk.LEFT, padx=5)
    ttk.Button(btn_frame, text="Delete", command=delete_subtask).pack(side=tk.LEFT, padx=5)
    ttk.Button(btn_frame, text="Close", command=dialog.destroy).pack(side=tk.RIGHT, padx=5)
    
    refresh_subtasks()

def open_advanced_search(self):
    """Open advanced search dialog."""
    if not hasattr(self.task_manager, 'search') or not self.task_manager.search:
        messagebox.showinfo("Feature Not Available", "Advanced search feature is not enabled.\n\nRun 'sql/create_enhanced_features.sql' first!")
        return
    
    messagebox.showinfo("Advanced Search", 
                       "Advanced Search feature coming soon!\n\n" +
                       "For now, the search box supports:\n" +
                       "- Full-text search across all fields\n" +
                       "- Case-insensitive matching\n\n" +
                       "Try typing keywords to search!")

# Bind methods to TaskManagerApp class
TaskManagerApp.start_timer = start_timer
TaskManagerApp.stop_timer = stop_timer
TaskManagerApp.view_time_log = view_time_log
TaskManagerApp.manage_subtasks = manage_subtasks
TaskManagerApp.open_advanced_search = open_advanced_search

# Add methods to MainMenuApp class
def open_dashboard(self):
    """Open dashboard window with real-time metrics and modern UI."""
    if not self.task_manager or not hasattr(self.task_manager, 'dashboard') or not self.task_manager.dashboard:
        messagebox.showinfo("Feature Not Available", 
                           "Dashboard feature is not enabled.\n\n" +
                           "Run 'sql/create_enhanced_features.sql' first!")
        return
    
    try:
        # Create dashboard window
        dashboard_win = tk.Toplevel(self.root)
        dashboard_win.title("Task Manager Dashboard")
        dashboard_win.geometry("1200x800")
        dashboard_win.transient(self.root)
        dashboard_win.minsize(1000, 600)
        
        # Store window reference for refresh
        dashboard_win.dashboard_app = self
        
        # Modern header with gradient effect (simulated with frame)
        header_frame = tk.Frame(dashboard_win, bg='#1a252f', height=70)
        header_frame.pack(fill=tk.X)
        header_frame.pack_propagate(False)
        
        # Title with icon
        title_frame = tk.Frame(header_frame, bg='#1a252f')
        title_frame.pack(side=tk.LEFT, padx=25, pady=15)
        
        tk.Label(title_frame, text="📊", 
                font=('Segoe UI Emoji', 24), bg='#1a252f', fg='#3498db').pack(side=tk.LEFT, padx=(0, 10))
        tk.Label(title_frame, text="Task Manager Dashboard", 
                font=('Segoe UI', 18, 'bold'), bg='#1a252f', fg='white').pack(side=tk.LEFT)
        
        # Header buttons
        btn_frame = tk.Frame(header_frame, bg='#1a252f')
        btn_frame.pack(side=tk.RIGHT, padx=25, pady=15)
        
        refresh_btn = tk.Button(btn_frame, text="🔄 Refresh", 
                               command=lambda: self._refresh_dashboard(dashboard_win),
                               bg='#3498db', fg='white', font=('Segoe UI', 10),
                               relief=tk.FLAT, padx=15, pady=8, cursor='hand2',
                               activebackground='#2980b9', activeforeground='white')
        refresh_btn.pack(side=tk.LEFT, padx=5)
        
        close_btn = tk.Button(btn_frame, text="✕ Close", 
                             command=dashboard_win.destroy,
                             bg='#34495e', fg='white', font=('Segoe UI', 10),
                             relief=tk.FLAT, padx=15, pady=8, cursor='hand2',
                             activebackground='#2c3e50', activeforeground='white')
        close_btn.pack(side=tk.LEFT, padx=5)
        
        # Main content frame with scrollbar
        main_frame = tk.Frame(dashboard_win, bg='#f5f6fa')
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        canvas = tk.Canvas(main_frame, bg='#f5f6fa', highlightthickness=0)
        scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg='#f5f6fa')
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        window_id = canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Ensure scrollable frame fills width
        def on_canvas_configure(event):
            canvas.itemconfig(window_id, width=event.width)
        canvas.bind('<Configure>', on_canvas_configure)
        
        # Fetch dashboard data
        summary = self.task_manager.dashboard.get_summary_metrics()
        user_workload = self.task_manager.dashboard.get_user_workload()
        time_tracking = self.task_manager.dashboard.get_time_tracking_summary()
        
        # Store reference for refresh
        dashboard_win.scrollable_frame = scrollable_frame
        dashboard_win.canvas = canvas
        
        # === SECTION 1: Modern Metric Cards ===
        cards_container = tk.Frame(scrollable_frame, bg='#f5f6fa')
        cards_container.pack(fill=tk.X, padx=20, pady=20)
        
        def create_modern_card(parent, title, value, icon, color, subtitle=""):
            card = tk.Frame(parent, bg='white', relief=tk.FLAT, bd=0)
            card.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=8)
            
            # Add subtle shadow effect with border
            card.configure(highlightbackground='#dfe6e9', highlightthickness=1)
            
            # Colored top bar
            top_bar = tk.Frame(card, bg=color, height=4)
            top_bar.pack(fill=tk.X)
            
            # Content
            content = tk.Frame(card, bg='white')
            content.pack(fill=tk.BOTH, expand=True, padx=20, pady=15)
            
            # Icon and title row
            header = tk.Frame(content, bg='white')
            header.pack(fill=tk.X, pady=(0, 10))
            
            tk.Label(header, text=icon, font=('Segoe UI Emoji', 20), 
                    bg='white', fg=color).pack(side=tk.LEFT, padx=(0, 10))
            tk.Label(header, text=title, font=('Segoe UI', 10), 
                    bg='white', fg='#636e72').pack(side=tk.LEFT, anchor='w')
            
            # Value
            tk.Label(content, text=str(value), font=('Segoe UI', 32, 'bold'), 
                    bg='white', fg='#2d3436').pack(anchor='w')
            
            # Subtitle
            if subtitle:
                tk.Label(content, text=subtitle, font=('Segoe UI', 9), 
                        bg='white', fg='#b2bec3').pack(anchor='w', pady=(5, 0))
            
            return card
        
        total_tasks = summary.get('total_tasks', 0)
        completed = summary.get('completed_tasks', 0)
        completion_rate = f"{(completed / total_tasks * 100):.0f}%" if total_tasks > 0 else "0%"
        
        create_modern_card(cards_container, "TOTAL TASKS", total_tasks, "📋", "#3498db", "All tasks in system")
        create_modern_card(cards_container, "ACTIVE", summary.get('active_tasks', 0), "🔵", "#2ecc71", "In progress")
        create_modern_card(cards_container, "COMPLETED", completed, "✅", "#27ae60", f"{completion_rate} success rate")
        create_modern_card(cards_container, "OVERDUE", summary.get('overdue_tasks', 0), "⚠️", "#e74c3c", "Need attention")
        
        # === SECTION 2: Charts Row ===
        charts_row = tk.Frame(scrollable_frame, bg='#f5f6fa')
        charts_row.pack(fill=tk.X, padx=20, pady=(0, 20))
        
        # Left: Priority Distribution (Pie Chart)
        priority_card = tk.Frame(charts_row, bg='white', highlightbackground='#dfe6e9', highlightthickness=1)
        priority_card.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 10))
        
        tk.Label(priority_card, text="📊 Priority Distribution", 
                font=('Segoe UI', 12, 'bold'), bg='white', fg='#2d3436').pack(anchor='w', padx=20, pady=(15, 10))
        
        # Simple ASCII pie chart visualization
        chart_frame = tk.Frame(priority_card, bg='white')
        chart_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=(0, 15))
        
        high_count = summary.get('high_priority_active', 0)
        medium_count = summary.get('medium_priority_active', 0)
        low_count = summary.get('low_priority_active', 0)
        total_active = high_count + medium_count + low_count
        
        if total_active > 0:
            # Create visual bars
            max_width = 300
            
            def create_bar(parent, label, value, total, color):
                row = tk.Frame(parent, bg='white')
                row.pack(fill=tk.X, pady=8)
                
                label_text = tk.Label(row, text=label, font=('Segoe UI', 10), 
                                     bg='white', fg='#636e72', width=15, anchor='w')
                label_text.pack(side=tk.LEFT)
                
                bar_container = tk.Frame(row, bg='#ecf0f1', height=25, width=max_width)
                bar_container.pack(side=tk.LEFT, padx=10)
                bar_container.pack_propagate(False)
                
                if value > 0:
                    bar_width = int((value / total) * max_width)
                    bar = tk.Frame(bar_container, bg=color, height=25, width=bar_width)
                    bar.pack(side=tk.LEFT)
                    bar.pack_propagate(False)
                
                percent = (value / total * 100) if total > 0 else 0
                count_label = tk.Label(row, text=f"{value} ({percent:.0f}%)", 
                                      font=('Segoe UI', 10, 'bold'), 
                                      bg='white', fg=color)
                count_label.pack(side=tk.LEFT, padx=10)
            
            create_bar(chart_frame, "🔴 High Priority", high_count, total_active, '#e74c3c')
            create_bar(chart_frame, "🟠 Medium Priority", medium_count, total_active, '#f39c12')
            create_bar(chart_frame, "🟢 Low Priority", low_count, total_active, '#2ecc71')
        else:
            tk.Label(chart_frame, text="No active tasks", 
                    font=('Segoe UI', 10), bg='white', fg='#b2bec3').pack(pady=40)
        
        # Right: Monthly Schedule Progress (Replaces "Overall Progress")
        progress_card = tk.Frame(charts_row, bg='white', highlightbackground='#dfe6e9', highlightthickness=1)
        progress_card.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(10, 0))
        
        tk.Label(progress_card, text="📅 Monthly Schedule", 
                font=('Segoe UI', 12, 'bold'), bg='white', fg='#2d3436').pack(anchor='w', padx=20, pady=(15, 10))
        
        progress_content = tk.Frame(progress_card, bg='white')
        progress_content.pack(fill=tk.BOTH, expand=True, padx=20, pady=(0, 15))
        
        # Circular progress indicator
        canvas_size = 200
        progress_canvas = tk.Canvas(progress_content, width=canvas_size, height=canvas_size, 
                                    bg='white', highlightthickness=0)
        progress_canvas.pack(pady=20)
        
        # Draw progress circle
        center = canvas_size / 2
        radius = 70
        width = 15
        
        # Background circle
        progress_canvas.create_oval(center-radius, center-radius, center+radius, center+radius,
                                    outline='#ecf0f1', width=width)
        
        # Monthly Metrics
        monthly_total = summary.get('monthly_due_total', 0)
        monthly_completed = summary.get('monthly_due_completed', 0)
        monthly_percent = summary.get('monthly_progress', 0)
        
        if monthly_total > 0:
            extent = (monthly_percent / 100) * 360
            
            # Determine color based on completion
            if monthly_percent >= 75:
                arc_color = '#2ecc71'
            elif monthly_percent >= 50:
                arc_color = '#f39c12'
            else:
                arc_color = '#3498db'
            
            progress_canvas.create_arc(center-radius, center-radius, center+radius, center+radius,
                                      start=90, extent=-extent, outline=arc_color, width=width, style=tk.ARC)
            
            # Center text
            progress_canvas.create_text(center, center-10, text=f"{monthly_percent:.0f}%",
                                       font=('Segoe UI', 28, 'bold'), fill='#2d3436')
            progress_canvas.create_text(center, center+20, text=f"{monthly_completed}/{monthly_total}",
                                       font=('Segoe UI', 12), fill='#636e72')
            progress_canvas.create_text(center, center+40, text="Completed",
                                       font=('Segoe UI', 10), fill='#b2bec3')
        else:
            progress_canvas.create_text(center, center, text="No Tasks\nDue Month",
                                       font=('Segoe UI', 12), fill='#b2bec3', justify=tk.CENTER)
        
        # === SECTION 3: Bottom Row (Workload & Upcoming) ===
        bottom_row = tk.Frame(scrollable_frame, bg='#f5f6fa')
        bottom_row.pack(fill=tk.X, padx=20, pady=(0, 20))
        
        # Left: User Workload
        workload_card = tk.Frame(bottom_row, bg='white', highlightbackground='#dfe6e9', highlightthickness=1)
        workload_card.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 10))
        
        tk.Label(workload_card, text="👥 User Workload", 
                font=('Segoe UI', 12, 'bold'), bg='white', fg='#2d3436').pack(anchor='w', padx=20, pady=(15, 10))
        
        tree_frame = tk.Frame(workload_card, bg='white')
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=(0, 15))
        
        style = ttk.Style()
        style.configure("Dashboard.Treeview", 
                       background="white", foreground="#2d3436", fieldbackground="white",
                       font=('Segoe UI', 10))
        style.configure("Dashboard.Treeview.Heading",
                       font=('Segoe UI', 10, 'bold'), foreground='#636e72')
        
        workload_tree = ttk.Treeview(tree_frame, 
                                     columns=("user", "active", "overdue"), 
                                     show='headings', height=8, style="Dashboard.Treeview")
        
        workload_tree.heading("user", text="User")
        workload_tree.heading("active", text="Active Tasks")
        workload_tree.heading("overdue", text="Overdue")
        
        workload_tree.column("user", width=150)
        workload_tree.column("active", width=100, anchor='center')
        workload_tree.column("overdue", width=100, anchor='center')
        
        for user_data in user_workload:
            workload_tree.insert("", tk.END, values=(
                user_data.get('user_name', 'Unknown'),
                user_data.get('active_tasks', 0),
                user_data.get('overdue_tasks', 0)
            ))
        
        tree_scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=workload_tree.yview)
        workload_tree.configure(yscrollcommand=tree_scrollbar.set)
        
        workload_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        tree_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # Right: Upcoming Deadlines (New Feature)
        upcoming_card = tk.Frame(bottom_row, bg='white', highlightbackground='#dfe6e9', highlightthickness=1)
        upcoming_card.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(10, 0))
        
        tk.Label(upcoming_card, text="📅 Upcoming Deadlines (Next 7 Days)", 
                font=('Segoe UI', 12, 'bold'), bg='white', fg='#2d3436').pack(anchor='w', padx=20, pady=(15, 10))
        
        upcoming_frame = tk.Frame(upcoming_card, bg='white')
        upcoming_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=(0, 15))

        upcoming_tree = ttk.Treeview(upcoming_frame, 
                                     columns=("date", "priority", "title", "assigned"), 
                                     show='headings', height=8, style="Dashboard.Treeview")
        
        upcoming_tree.heading("date", text="Due Date")
        upcoming_tree.heading("priority", text="Pri")
        upcoming_tree.heading("title", text="Task")
        upcoming_tree.heading("assigned", text="Assigned To")
        
        upcoming_tree.column("date", width=90)
        upcoming_tree.column("priority", width=50, anchor='center')
        upcoming_tree.column("title", width=250)
        upcoming_tree.column("assigned", width=100)
        
        # Configure tags for priority coloring
        upcoming_tree.tag_configure('High', foreground='#e74c3c')    # Red
        upcoming_tree.tag_configure('Medium', foreground='#f39c12')  # Orange
        upcoming_tree.tag_configure('Low', foreground='#2ecc71')     # Green
        upcoming_tree.tag_configure('Default', foreground='#2d3436') # Dark Grey

        # Fetch upcoming tasks
        try:
            upcoming_tasks = self.task_manager.dashboard.get_upcoming_deadlines()
            if upcoming_tasks:
                for task in upcoming_tasks:
                    priority_val = task.get('priority', 'Low')
                    priority_icon = "🔴" if priority_val == 'High' else "🟠" if priority_val == 'Medium' else "🟢"
                    
                    # Insert with priority tag
                    upcoming_tree.insert("", tk.END, values=(
                        task.get('due_date'),
                        priority_icon,
                        task.get('title'),
                        task.get('assigned_to')
                    ), tags=(priority_val,))
            else:
                # Add placeholder if empty
                upcoming_tree.insert("", tk.END, values=("", "", "No upcoming deadlines found", ""), tags=('Default',))
        except Exception:
            pass # Handle case where metod might not exist yet if hot-reloading issues occur
            
        upcoming_scrollbar = ttk.Scrollbar(upcoming_frame, orient="vertical", command=upcoming_tree.yview)
        upcoming_tree.configure(yscrollcommand=upcoming_scrollbar.set)
        
        upcoming_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        upcoming_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # === SECTION 4: Time Tracking Summary ===
        if time_tracking:
            time_card = tk.Frame(scrollable_frame, bg='white', highlightbackground='#dfe6e9', highlightthickness=1)
            time_card.pack(fill=tk.BOTH, expand=True, padx=20, pady=(0, 20))
            
            tk.Label(time_card, text="⏱️ Time Tracking Summary (Top 10)", 
                    font=('Segoe UI', 12, 'bold'), bg='white', fg='#2d3436').pack(anchor='w', padx=20, pady=(15, 10))
            
            time_tree_frame = tk.Frame(time_card, bg='white')
            time_tree_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=(0, 15))
            
            time_tree = ttk.Treeview(time_tree_frame, 
                                    columns=("task_id", "title", "user", "hours"), 
                                    show='headings', height=10, style="Dashboard.Treeview")
            
            time_tree.heading("task_id", text="Task ID")
            time_tree.heading("title", text="Task Title")
            time_tree.heading("user", text="Assigned To")
            time_tree.heading("hours", text="Hours Logged")
            
            time_tree.column("task_id", width=80, anchor='center')
            time_tree.column("title", width=400)
            time_tree.column("user", width=150)
            time_tree.column("hours", width=120, anchor='center')
            
            for task_data in time_tracking[:10]:  # Top 10 tasks
                total_hours = task_data.get('total_hours_logged', 0)
                time_tree.insert("", tk.END, values=(
                    task_data.get('task_id', ''),
                    task_data.get('title', '')[:60],
                    task_data.get('assigned_to', 'Unassigned'),
                    f"{total_hours:.1f}h"
                ))
            
            time_scrollbar = ttk.Scrollbar(time_tree_frame, orient="vertical", command=time_tree.yview)
            time_tree.configure(yscrollcommand=time_scrollbar.set)
            
            time_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            time_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # === SECTION 5: Performance Metrics ===
        metrics_card = tk.Frame(scrollable_frame, bg='white', highlightbackground='#dfe6e9', highlightthickness=1)
        metrics_card.pack(fill=tk.X, padx=20, pady=(0, 20))
        
        tk.Label(metrics_card, text="📊 Performance Metrics", 
                font=('Segoe UI', 12, 'bold'), bg='white', fg='#2d3436').pack(anchor='w', padx=20, pady=(15, 10))
        
        metrics_content = tk.Frame(metrics_card, bg='white')
        metrics_content.pack(fill=tk.X, padx=20, pady=(0, 15))
        
        avg_completion = summary.get('avg_completion_days', 0)
        total_estimated = summary.get('total_estimated_mhr', 0)
        total_actual = summary.get('total_actual_mhr', 0)
        
        metrics_data = [
            ("⏳ Average Completion Time", f"{avg_completion:.1f} days" if avg_completion else "N/A", '#3498db'),
            ("📋 Total Estimated MHR", f"{total_estimated:,.0f}", '#9b59b6'),
            ("✅ Total Actual MHR", f"{total_actual:,.0f}", '#2ecc71'),
            ("🎯 Completion Rate", f"{completion_rate}", '#27ae60')
        ]
        
        metrics_grid = tk.Frame(metrics_content, bg='white')
        metrics_grid.pack(fill=tk.X)
        
        for i, (label, value, color) in enumerate(metrics_data):
            metric_box = tk.Frame(metrics_grid, bg='white')
            if i < 2:
                metric_box.grid(row=0, column=i, sticky='ew', padx=10, pady=10)
            else:
                metric_box.grid(row=1, column=i-2, sticky='ew', padx=10, pady=10)
            
            metrics_grid.columnconfigure(i % 2, weight=1)
            
            tk.Label(metric_box, text=label, font=('Segoe UI', 10), 
                    bg='white', fg='#636e72').pack(anchor='w')
            tk.Label(metric_box, text=value, font=('Segoe UI', 18, 'bold'), 
                    bg='white', fg=color).pack(anchor='w', pady=(5, 0))
        
        # Pack canvas and scrollbar
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Bind mousewheel for scrolling
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        
        canvas.bind_all("<MouseWheel>", _on_mousewheel)
        dashboard_win.protocol("WM_DELETE_WINDOW", lambda: [canvas.unbind_all("<MouseWheel>"), dashboard_win.destroy()])
        
    except Exception as e:
        logging.error(f"Error opening dashboard: {e}", exc_info=True)
        messagebox.showerror("Dashboard Error", 
                           f"Failed to load dashboard:\n{str(e)}\n\n" +
                           "Check task_manager.log for details.")

def _refresh_dashboard(self, dashboard_win):
    """Refresh dashboard data without changing window position."""
    try:
        # Store current window geometry and scroll position
        geometry = dashboard_win.geometry()
        scroll_pos = dashboard_win.canvas.yview()[0]
        
        # Clear existing scrollable content
        for widget in dashboard_win.scrollable_frame.winfo_children():
            widget.destroy()
        
        # Re-fetch data
        summary = self.task_manager.dashboard.get_summary_metrics()
        user_workload = self.task_manager.dashboard.get_user_workload()
        time_tracking = self.task_manager.dashboard.get_time_tracking_summary()
        
        scrollable_frame = dashboard_win.scrollable_frame
        
        # Rebuild all sections - COMPLETE REBUILD
        # === SECTION 1: Modern Metric Cards ===
        cards_container = tk.Frame(scrollable_frame, bg='#f5f6fa')
        cards_container.pack(fill=tk.X, padx=20, pady=20)
        
        def create_modern_card(parent, title, value, icon, color, subtitle=""):
            card = tk.Frame(parent, bg='white', relief=tk.FLAT, bd=0)
            card.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=8)
            card.configure(highlightbackground='#dfe6e9', highlightthickness=1)
            top_bar = tk.Frame(card, bg=color, height=4)
            top_bar.pack(fill=tk.X)
            content = tk.Frame(card, bg='white')
            content.pack(fill=tk.BOTH, expand=True, padx=20, pady=15)
            header = tk.Frame(content, bg='white')
            header.pack(fill=tk.X, pady=(0, 10))
            tk.Label(header, text=icon, font=('Segoe UI Emoji', 20), 
                    bg='white', fg=color).pack(side=tk.LEFT, padx=(0, 10))
            tk.Label(header, text=title, font=('Segoe UI', 10), 
                    bg='white', fg='#636e72').pack(side=tk.LEFT, anchor='w')
            tk.Label(content, text=str(value), font=('Segoe UI', 32, 'bold'), 
                    bg='white', fg='#2d3436').pack(anchor='w')
            if subtitle:
                tk.Label(content, text=subtitle, font=('Segoe UI', 9), 
                        bg='white', fg='#b2bec3').pack(anchor='w', pady=(5, 0))
        
        total_tasks = summary.get('total_tasks', 0)
        completed = summary.get('completed_tasks', 0)
        completion_rate = f"{(completed / total_tasks * 100):.0f}%" if total_tasks > 0 else "0%"
        
        create_modern_card(cards_container, "TOTAL TASKS", total_tasks, "📋", "#3498db", "All tasks in system")
        create_modern_card(cards_container, "ACTIVE", summary.get('active_tasks', 0), "🔵", "#2ecc71", "In progress")
        create_modern_card(cards_container, "COMPLETED", completed, "✅", "#27ae60", f"{completion_rate} success rate")
        create_modern_card(cards_container, "OVERDUE", summary.get('overdue_tasks', 0), "⚠️", "#e74c3c", "Need attention")
        
        # === SECTION 2: Charts Row ===
        charts_row = tk.Frame(scrollable_frame, bg='#f5f6fa')
        charts_row.pack(fill=tk.X, padx=20, pady=(0, 20))
        
        # Left: Priority Distribution
        priority_card = tk.Frame(charts_row, bg='white', highlightbackground='#dfe6e9', highlightthickness=1)
        priority_card.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 10))
        
        tk.Label(priority_card, text="📊 Priority Distribution", 
                font=('Segoe UI', 12, 'bold'), bg='white', fg='#2d3436').pack(anchor='w', padx=20, pady=(15, 10))
        
        chart_frame = tk.Frame(priority_card, bg='white')
        chart_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=(0, 15))
        
        high_count = summary.get('high_priority_active', 0)
        medium_count = summary.get('medium_priority_active', 0)
        low_count = summary.get('low_priority_active', 0)
        total_active = high_count + medium_count + low_count
        
        if total_active > 0:
            max_width = 300
            
            def create_bar(parent, label, value, total, color):
                row = tk.Frame(parent, bg='white')
                row.pack(fill=tk.X, pady=8)
                label_text = tk.Label(row, text=label, font=('Segoe UI', 10), 
                                     bg='white', fg='#636e72', width=15, anchor='w')
                label_text.pack(side=tk.LEFT)
                bar_container = tk.Frame(row, bg='#ecf0f1', height=25, width=max_width)
                bar_container.pack(side=tk.LEFT, padx=10)
                bar_container.pack_propagate(False)
                if value > 0:
                    bar_width = int((value / total) * max_width)
                    bar = tk.Frame(bar_container, bg=color, height=25, width=bar_width)
                    bar.pack(side=tk.LEFT)
                    bar.pack_propagate(False)
                percent = (value / total * 100) if total > 0 else 0
                count_label = tk.Label(row, text=f"{value} ({percent:.0f}%)", 
                                      font=('Segoe UI', 10, 'bold'), bg='white', fg=color)
                count_label.pack(side=tk.LEFT, padx=10)
            
            create_bar(chart_frame, "🔴 High Priority", high_count, total_active, '#e74c3c')
            create_bar(chart_frame, "🟠 Medium Priority", medium_count, total_active, '#f39c12')
            create_bar(chart_frame, "🟢 Low Priority", low_count, total_active, '#2ecc71')
        else:
            tk.Label(chart_frame, text="No active tasks", 
                    font=('Segoe UI', 10), bg='white', fg='#b2bec3').pack(pady=40)
        
        # Right: Monthly Schedule (Updated Logic)
        progress_card = tk.Frame(charts_row, bg='white', highlightbackground='#dfe6e9', highlightthickness=1)
        progress_card.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(10, 0))
        
        tk.Label(progress_card, text="📅 Monthly Schedule", 
                font=('Segoe UI', 12, 'bold'), bg='white', fg='#2d3436').pack(anchor='w', padx=20, pady=(15, 10))
        
        progress_content = tk.Frame(progress_card, bg='white')
        progress_content.pack(fill=tk.BOTH, expand=True, padx=20, pady=(0, 15))
        
        canvas_size = 200
        progress_canvas = tk.Canvas(progress_content, width=canvas_size, height=canvas_size, 
                                    bg='white', highlightthickness=0)
        progress_canvas.pack(pady=20)
        
        center = canvas_size / 2
        radius = 70
        width = 15
        
        # Monthly Data
        monthly_completed = summary.get('monthly_due_completed', 0)
        monthly_total = summary.get('monthly_due_total', 0)
        
        progress_canvas.create_oval(center-radius, center-radius, center+radius, center+radius,
                                    outline='#ecf0f1', width=width)
        
        if monthly_total > 0:
            progress_percent = (monthly_completed / monthly_total) * 100
            extent = (progress_percent / 100) * 360
            
            if progress_percent >= 75:
                arc_color = '#2ecc71'
            elif progress_percent >= 50:
                arc_color = '#f39c12'
            else:
                arc_color = '#3498db'
            
            progress_canvas.create_arc(center-radius, center-radius, center+radius, center+radius,
                                      start=90, extent=-extent, outline=arc_color, width=width, style=tk.ARC)
            
            progress_canvas.create_text(center, center-10, text=f"{progress_percent:.0f}%",
                                       font=('Segoe UI', 28, 'bold'), fill='#2d3436')
            progress_canvas.create_text(center, center+20, text=f"{monthly_completed}/{monthly_total}",
                                       font=('Segoe UI', 11), fill='#636e72')
        else:
            progress_canvas.create_text(center, center, text="No Schedule",
                                       font=('Segoe UI', 14), fill='#b2bec3')
            progress_canvas.create_text(center, center+20, text="This Month",
                                       font=('Segoe UI', 10), fill='#b2bec3')
        
        # === SECTION 3: Bottom Row (Workload & Upcoming) ===
        bottom_row = tk.Frame(scrollable_frame, bg='#f5f6fa')
        bottom_row.pack(fill=tk.X, padx=20, pady=(0, 20))
        
        # Left: User Workload
        workload_card = tk.Frame(bottom_row, bg='white', highlightbackground='#dfe6e9', highlightthickness=1)
        workload_card.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 10))
        
        tk.Label(workload_card, text="👥 User Workload", 
                font=('Segoe UI', 12, 'bold'), bg='white', fg='#2d3436').pack(anchor='w', padx=20, pady=(15, 10))
        
        tree_frame = tk.Frame(workload_card, bg='white')
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=(0, 15))
        
        style = ttk.Style()
        style.configure("Dashboard.Treeview", 
                       background="white", foreground="#2d3436", fieldbackground="white",
                       font=('Segoe UI', 10))
        style.configure("Dashboard.Treeview.Heading",
                       font=('Segoe UI', 10, 'bold'), foreground='#636e72')
        
        workload_tree = ttk.Treeview(tree_frame, 
                                     columns=("user", "active", "overdue"), 
                                     show='headings', height=8, style="Dashboard.Treeview")
        
        workload_tree.heading("user", text="User")
        workload_tree.heading("active", text="Active Tasks")
        workload_tree.heading("overdue", text="Overdue")
        
        workload_tree.column("user", width=150)
        workload_tree.column("active", width=100, anchor='center')
        workload_tree.column("overdue", width=100, anchor='center')
        
        for user_data in user_workload:
            workload_tree.insert("", tk.END, values=(
                user_data.get('user_name', 'Unknown'),
                user_data.get('active_tasks', 0),
                user_data.get('overdue_tasks', 0)
            ))
        
        tree_scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=workload_tree.yview)
        workload_tree.configure(yscrollcommand=tree_scrollbar.set)
        
        workload_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        tree_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # Right: Upcoming Deadlines (New Feature)
        upcoming_card = tk.Frame(bottom_row, bg='white', highlightbackground='#dfe6e9', highlightthickness=1)
        upcoming_card.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(10, 0))
        
        tk.Label(upcoming_card, text="📅 Upcoming Deadlines (Next 7 Days)", 
                font=('Segoe UI', 12, 'bold'), bg='white', fg='#2d3436').pack(anchor='w', padx=20, pady=(15, 10))
        
        upcoming_frame = tk.Frame(upcoming_card, bg='white')
        upcoming_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=(0, 15))

        upcoming_tree = ttk.Treeview(upcoming_frame, 
                                     columns=("date", "priority", "title", "assigned"), 
                                     show='headings', height=8, style="Dashboard.Treeview")
        
        upcoming_tree.heading("date", text="Due Date")
        upcoming_tree.heading("priority", text="Pri")
        upcoming_tree.heading("title", text="Task")
        upcoming_tree.heading("assigned", text="Assigned To")
        
        upcoming_tree.column("date", width=90)
        upcoming_tree.column("priority", width=50, anchor='center')
        upcoming_tree.column("title", width=250)
        upcoming_tree.column("assigned", width=100)
        
        # Configure tags for priority coloring
        upcoming_tree.tag_configure('High', foreground='#e74c3c')    # Red
        upcoming_tree.tag_configure('Medium', foreground='#f39c12')  # Orange
        upcoming_tree.tag_configure('Low', foreground='#2ecc71')     # Green
        upcoming_tree.tag_configure('Default', foreground='#2d3436') # Dark Grey

        # Fetch upcoming tasks
        try:
            upcoming_tasks = self.task_manager.dashboard.get_upcoming_deadlines()
            if upcoming_tasks:
                for task in upcoming_tasks:
                    priority_val = task.get('priority', 'Low')
                    priority_icon = "🔴" if priority_val == 'High' else "🟠" if priority_val == 'Medium' else "🟢"
                    
                    # Insert with priority tag
                    upcoming_tree.insert("", tk.END, values=(
                        task.get('due_date'),
                        priority_icon,
                        task.get('title'),
                        task.get('assigned_to')
                    ), tags=(priority_val,))
            else:
                # Add placeholder if empty
                upcoming_tree.insert("", tk.END, values=("", "", "No upcoming deadlines found", ""), tags=('Default',))
        except Exception:
            pass # Handle case where metod might not exist yet if hot-reloading issues occur
            
        upcoming_scrollbar = ttk.Scrollbar(upcoming_frame, orient="vertical", command=upcoming_tree.yview)
        upcoming_tree.configure(yscrollcommand=upcoming_scrollbar.set)
        
        upcoming_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        upcoming_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # === SECTION 4: Time Tracking Summary ===
        if time_tracking:
            time_card = tk.Frame(scrollable_frame, bg='white', highlightbackground='#dfe6e9', highlightthickness=1)
            time_card.pack(fill=tk.BOTH, expand=True, padx=20, pady=(0, 20))
            
            tk.Label(time_card, text="⏱️ Time Tracking Summary (Top 10)", 
                    font=('Segoe UI', 12, 'bold'), bg='white', fg='#2d3436').pack(anchor='w', padx=20, pady=(15, 10))
            
            time_tree_frame = tk.Frame(time_card, bg='white')
            time_tree_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=(0, 15))
            
            time_tree = ttk.Treeview(time_tree_frame, 
                                    columns=("task_id", "title", "user", "hours"), 
                                    show='headings', height=10, style="Dashboard.Treeview")
            
            time_tree.heading("task_id", text="Task ID")
            time_tree.heading("title", text="Task Title")
            time_tree.heading("user", text="Assigned To")
            time_tree.heading("hours", text="Hours Logged")
            
            time_tree.column("task_id", width=80, anchor='center')
            time_tree.column("title", width=400)
            time_tree.column("user", width=150)
            time_tree.column("hours", width=120, anchor='center')
            
            for task_data in time_tracking[:10]:
                total_hours = task_data.get('total_hours_logged', 0)
                time_tree.insert("", tk.END, values=(
                    task_data.get('task_id', ''),
                    task_data.get('title', '')[:60],
                    task_data.get('assigned_to', 'Unassigned'),
                    f"{total_hours:.1f}h"
                ))
            
            time_scrollbar = ttk.Scrollbar(time_tree_frame, orient="vertical", command=time_tree.yview)
            time_tree.configure(yscrollcommand=time_scrollbar.set)
            
            time_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            time_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # === SECTION 5: Performance Metrics ===
        metrics_card = tk.Frame(scrollable_frame, bg='white', highlightbackground='#dfe6e9', highlightthickness=1)
        metrics_card.pack(fill=tk.X, padx=20, pady=(0, 20))
        
        tk.Label(metrics_card, text="📊 Performance Metrics", 
                font=('Segoe UI', 12, 'bold'), bg='white', fg='#2d3436').pack(anchor='w', padx=20, pady=(15, 10))
        
        metrics_content = tk.Frame(metrics_card, bg='white')
        metrics_content.pack(fill=tk.X, padx=20, pady=(0, 15))
        
        avg_completion = summary.get('avg_completion_days', 0)
        total_estimated = summary.get('total_estimated_mhr', 0)
        total_actual = summary.get('total_actual_mhr', 0)
        
        metrics_data = [
            ("⏳ Average Completion Time", f"{avg_completion:.1f} days" if avg_completion else "N/A", '#3498db'),
            ("📋 Total Estimated MHR", f"{total_estimated:,.0f}", '#9b59b6'),
            ("✅ Total Actual MHR", f"{total_actual:,.0f}", '#2ecc71'),
            ("🎯 Completion Rate", f"{completion_rate}", '#27ae60')
        ]
        
        metrics_grid = tk.Frame(metrics_content, bg='white')
        metrics_grid.pack(fill=tk.X)
        
        for i, (label, value, color) in enumerate(metrics_data):
            metric_box = tk.Frame(metrics_grid, bg='white')
            if i < 2:
                metric_box.grid(row=0, column=i, sticky='ew', padx=10, pady=10)
            else:
                metric_box.grid(row=1, column=i-2, sticky='ew', padx=10, pady=10)
            
            metrics_grid.columnconfigure(i % 2, weight=1)
            
            tk.Label(metric_box, text=label, font=('Segoe UI', 10), 
                    bg='white', fg='#636e72').pack(anchor='w')
            tk.Label(metric_box, text=value, font=('Segoe UI', 18, 'bold'), 
                    bg='white', fg=color).pack(anchor='w', pady=(5, 0))
        
        # Restore window geometry and scroll position
        dashboard_win.geometry(geometry)
        dashboard_win.update_idletasks()
        dashboard_win.canvas.yview_moveto(scroll_pos)
        
        # Show success message in status bar style (no popup to avoid interruption)
        logging.info("Dashboard refreshed successfully")
        
    except Exception as e:
        logging.error(f"Error refreshing dashboard: {e}", exc_info=True)
        messagebox.showerror("Refresh Error", f"Failed to refresh dashboard:\n{str(e)}", parent=dashboard_win)

def open_templates(self):
    """Open template management window."""
    if not self.task_manager or not hasattr(self.task_manager, 'templates') or not self.task_manager.templates:
        messagebox.showinfo("Feature Not Available", 
                           "Templates feature is not enabled.\n\n" +
                           "Run 'sql/create_enhanced_features.sql' first!")
        return
    
    messagebox.showinfo("Templates", 
                       "Templates feature coming soon!\n\n" +
                       "Templates will let you:\n" +
                       "- Save common task configurations\n" +
                       "- Create tasks from templates\n" +
                       "- Share templates with team\n\n" +
                       "Database tables are already created and ready!")

# Bind methods to MainMenuApp class
MainMenuApp.open_dashboard = open_dashboard
MainMenuApp._refresh_dashboard = _refresh_dashboard
MainMenuApp.open_templates = open_templates


def main():
    """Main function - Entry point of the application."""
    print(f"------------------------------")
    print(f"Task Manager Application Starting...")
    print(f"Log file: {log_file_path}")
    print(f"------------------------------")

    logging.info("Application starting in main()...")
    
    # Check for updates before starting the app
    check_for_updates()

    root = tk.Tk()
    try:
        menu_app = MainMenuApp(root)
        menu_app.run()

    except Exception as e:
        logging.error("Unhandled exception in main", exc_info=True)
        messagebox.showerror("Fatal Error", f"An unexpected error occurred: {str(e)}\nCheck task_manager.log for details.")
        try:
            root.destroy()
        except (tk.TclError, AttributeError) as destroy_err:
            logging.debug(f"Error destroying root window: {destroy_err}")
        logging.info("Flushing log handlers after general error exit...")
        for handler in logging.getLogger().handlers:
            try:
                handler.flush()
                handler.close()
            except Exception as e:
                print(f"Error closing handler {handler}: {e}")


if __name__ == "__main__":
    main()